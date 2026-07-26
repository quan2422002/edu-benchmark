"""Build and validate the type-oriented HNMU Phase 1 teacher bundle v2.

The v2 packager reuses the canonical Plan 08 loader and validation joins. It
does not rerun the experiment or audit, and it treats source_file values as
opaque provenance strings.
"""

from __future__ import annotations

import csv
import math
import re
import tempfile
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook

from edu_benchmark.dialogue_audit.hnmu_audit import lesson_code
from edu_benchmark.dialogue_audit.teacher_bundle import (
    ALLOWED_DECISIONS,
    EXPECTED_CRITERIA_PER_SAMPLE,
    EXPECTED_GRADE_COUNTS,
    BundleData,
    _append_rows,
    _display_path,
    _set_widths,
    _style_table,
    load_canonical_bundle_data,
)

STATUS_ORDER = ("pass", "need_human_review", "failed")
STATUS_LABELS = {
    "pass": "Đạt theo trạng thái tổng thể chính thức",
    "need_human_review": "Cần giáo viên xem lại",
    "failed": "Chưa nên dùng ở lượt hiện tại",
    "non_pass": "Tổng chưa đạt ngay",
}
V2_OUTPUT_NAMES = (
    "README.md",
    "01_bao_cao_tong_quan.md",
    "02_checklist_tieu_chi.xlsx",
    "03_du_lieu_tho_sau_chuan_hoa.csv",
    "04_thong_ke_pass_reject_theo_khoi.xlsx",
    "05_thong_ke_do_phu_mau_pass_theo_khoi.xlsx",
    "06_ket_qua_cham_tong_the_tung_mau.csv",
    "07_ket_qua_cham_chi_tiet_tung_tieu_chi.csv",
    "08_mau_thieu_sai_truong_du_lieu.csv",
    "09_ung_vien_trung_lap.csv",
)
NORMALIZED_FIELDS = (
    "sample_id",
    "source_file",
    "source_row_number",
    "grade",
    "grade_label",
    "stt",
    "lesson",
    "position",
    "question",
    "bloom_level",
    "answer_sgv",
    "dialogue",
)
QUALITY_FIELDS = (
    "sample_id",
    "source_file",
    "source_row_number",
    "grade",
    "lesson",
    "quality_decision",
    "confidence_score",
    "failure_reasons",
    "blocking_criterion_ids",
    "suggested_reviewer_action",
    "needs_hnmu_review",
    "needs_learning_resource_review",
    "needs_sgv_verification",
    "evidence_fragment_ids",
    "checked_by",
    "checked_at",
    "source_shard",
)
CHECKLIST_FIELDS = (
    "sample_id",
    "grade",
    "source_file",
    "source_row_number",
    "criterion_id",
    "criterion_group",
    "criterion_name",
    "result",
    "confidence_score",
    "evidence_fragment_id",
    "evidence_source",
    "evidence_match_reason",
    "reason",
    "suggested_reviewer_action",
    "checked_by",
    "checked_at",
    "shard_id",
)
MISSING_FIELDS = (
    "sample_id",
    "grade",
    "source_file",
    "source_row_number",
    "issue_type",
    "field",
    "severity",
    "message",
)
DUPLICATE_FIELDS = (
    "grade",
    "sample_id_a",
    "source_file_a",
    "source_row_number_a",
    "sample_id_b",
    "source_file_b",
    "source_row_number_b",
    "duplicate_type",
    "similarity",
    "note",
)
STATUS_HEADERS = (
    "grade",
    "grade_label",
    "status",
    "status_label",
    "count",
    "percentage",
    "total_samples",
)
COVERAGE_HEADERS = (
    "grade",
    "grade_label",
    "dimension",
    "dimension_label",
    "value_id",
    "value_label",
    "count",
    "percentage_among_pass",
    "grade_pass_total",
)
CRITERIA_HEADERS = (
    "criterion_id",
    "criterion_group",
    "criterion_name",
    "required_per_sample",
    "sample_count",
)


PERCENTAGE_QUANTUM = Decimal("0.01")


def _percentage_decimal_half_up(
    numerator: int | float | Decimal,
    denominator: int | float | Decimal,
) -> Decimal:
    """Return a percentage rounded half up to two decimal places."""

    denominator_decimal = Decimal(str(denominator))
    if denominator_decimal == 0:
        raise ZeroDivisionError("Cannot calculate a percentage with a zero denominator")
    return (
        Decimal(str(numerator)) * Decimal("100") / denominator_decimal
    ).quantize(PERCENTAGE_QUANTUM, rounding=ROUND_HALF_UP)


def format_percentage_half_up(
    numerator: int | float | Decimal,
    denominator: int | float | Decimal,
) -> str:
    """Format a count ratio as a two-decimal ROUND_HALF_UP percentage."""

    return f"{_percentage_decimal_half_up(numerator, denominator):.2f}%"


def round_percentage_fraction_half_up(
    numerator: int | float | Decimal,
    denominator: int | float | Decimal,
) -> float:
    """Return the fraction underlying a two-decimal ROUND_HALF_UP percentage."""

    return float(_percentage_decimal_half_up(numerator, denominator) / Decimal("100"))


def format_fraction_percentage_half_up(value: int | float | Decimal) -> str:
    """Format an existing fraction as a two-decimal ROUND_HALF_UP percentage."""

    percentage = (Decimal(str(value)) * Decimal("100")).quantize(
        PERCENTAGE_QUANTUM,
        rounding=ROUND_HALF_UP,
    )
    return f"{percentage:.2f}%"


@dataclass(frozen=True)
class ChiSquareResult:
    statistic: float
    degrees_of_freedom: int
    p_value: float
    cramers_v: float


@dataclass(frozen=True)
class V2BuildSummary:
    output_paths: tuple[Path, ...]
    grade_counts: dict[str, int]
    status_counts: dict[str, dict[str, int]]
    chi_square: ChiSquareResult
    source_hashes: dict[str, str]
    read_paths: tuple[str, ...]


def _sample_sort_key(data: BundleData, sample_id: str) -> tuple[int, int, str]:
    row = data.normalized_by_id[sample_id]
    source_row = row.get("source_row_number", "")
    return (
        int(row["grade"]),
        int(source_row) if source_row.isdigit() else 10**9,
        sample_id,
    )


def _sample_ids(data: BundleData) -> list[str]:
    return sorted(data.normalized_by_id, key=lambda item: _sample_sort_key(data, item))


def _normalized_rows(data: BundleData) -> list[dict[str, str]]:
    return [
        {field: data.normalized_by_id[sample_id].get(field, "") for field in NORMALIZED_FIELDS}
        for sample_id in _sample_ids(data)
    ]


def _quality_rows(data: BundleData) -> list[dict[str, str]]:
    return [
        {field: data.quality_by_id[sample_id].get(field, "") for field in QUALITY_FIELDS}
        for sample_id in _sample_ids(data)
    ]


def _checklist_rows(data: BundleData) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for sample_id in _sample_ids(data):
        normalized = data.normalized_by_id[sample_id]
        for source in sorted(data.checklist_by_id[sample_id], key=lambda row: row["criterion_id"]):
            output.append(
                {
                    "sample_id": sample_id,
                    "grade": normalized["grade"],
                    "source_file": normalized["source_file"],
                    "source_row_number": normalized["source_row_number"],
                    **{field: source.get(field, "") for field in CHECKLIST_FIELDS[4:]},
                }
            )
    return output


def _missing_rows(data: BundleData) -> list[dict[str, str]]:
    output = []
    for source in data.missing_rows:
        sample_id = source["sample_id"]
        normalized = data.normalized_by_id[sample_id]
        output.append(
            {
                "sample_id": sample_id,
                "grade": normalized["grade"],
                "source_file": normalized["source_file"],
                "source_row_number": normalized["source_row_number"],
                "issue_type": source.get("issue_type", ""),
                "field": source.get("field", ""),
                "severity": source.get("severity", ""),
                "message": source.get("message", ""),
            }
        )
    return sorted(output, key=lambda row: _sample_sort_key(data, row["sample_id"]))


def _duplicate_rows(data: BundleData) -> list[dict[str, str]]:
    output = []
    for source in data.duplicate_rows:
        sample_a = source["sample_id_a"]
        sample_b = source["sample_id_b"]
        normalized_a = data.normalized_by_id[sample_a]
        normalized_b = data.normalized_by_id[sample_b]
        if normalized_a["grade"] != normalized_b["grade"]:
            raise ValueError(f"Duplicate candidate crosses grades: {sample_a}, {sample_b}")
        output.append(
            {
                "grade": normalized_a["grade"],
                "sample_id_a": sample_a,
                "source_file_a": normalized_a["source_file"],
                "source_row_number_a": normalized_a["source_row_number"],
                "sample_id_b": sample_b,
                "source_file_b": normalized_b["source_file"],
                "source_row_number_b": normalized_b["source_row_number"],
                "duplicate_type": source.get("duplicate_type", ""),
                "similarity": source.get("similarity", ""),
                "note": source.get("note", ""),
            }
        )
    return sorted(output, key=lambda row: (int(row["grade"]), row["sample_id_a"], row["sample_id_b"]))


def _criteria_rows(data: BundleData) -> list[list[object]]:
    definitions: dict[str, tuple[str, str]] = {}
    counts: Counter[str] = Counter()
    for rows in data.checklist_by_id.values():
        for row in rows:
            criterion_id = row["criterion_id"]
            definition = (row.get("criterion_group", ""), row.get("criterion_name", ""))
            previous = definitions.setdefault(criterion_id, definition)
            if previous != definition:
                raise ValueError(f"Conflicting criterion definition: {criterion_id}")
            counts[criterion_id] += 1
    if len(definitions) != EXPECTED_CRITERIA_PER_SAMPLE:
        raise ValueError(f"Expected 18 criterion definitions, found {len(definitions)}")
    expected_samples = len(data.normalized_by_id)
    if any(count != expected_samples for count in counts.values()):
        raise ValueError("Criterion definitions do not cover every sample")
    return [
        [criterion_id, *definitions[criterion_id], "Có", counts[criterion_id]]
        for criterion_id in sorted(definitions)
    ]


def _status_counts(data: BundleData) -> dict[str, dict[str, int]]:
    result: dict[str, dict[str, int]] = {}
    for grade in EXPECTED_GRADE_COUNTS:
        counts = Counter(
            data.quality_by_id[sample_id]["quality_decision"]
            for sample_id, row in data.normalized_by_id.items()
            if row["grade"] == grade
        )
        result[grade] = {status: counts[status] for status in STATUS_ORDER}
    result["all"] = {
        status: sum(result[grade][status] for grade in EXPECTED_GRADE_COUNTS)
        for status in STATUS_ORDER
    }
    return result


def _status_rows(data: BundleData) -> list[list[object]]:
    counts = _status_counts(data)
    output: list[list[object]] = []
    for grade in (*EXPECTED_GRADE_COUNTS, "all"):
        total = sum(counts[grade].values())
        grade_label = f"Lớp {grade}" if grade != "all" else "Toàn bộ"
        for status in STATUS_ORDER:
            count = counts[grade][status]
            output.append([grade, grade_label, status, STATUS_LABELS[status], count, round_percentage_fraction_half_up(count, total), total])
        non_pass = counts[grade]["need_human_review"] + counts[grade]["failed"]
        output.append(
            [grade, grade_label, "non_pass", STATUS_LABELS["non_pass"], non_pass, round_percentage_fraction_half_up(non_pass, total), total]
        )
    return output


def _chi_square_result(data: BundleData) -> ChiSquareResult:
    counts = _status_counts(data)
    matrix = [[counts[grade][status] for status in STATUS_ORDER] for grade in EXPECTED_GRADE_COUNTS]
    row_totals = [sum(row) for row in matrix]
    column_totals = [sum(row[index] for row in matrix) for index in range(len(STATUS_ORDER))]
    total = sum(row_totals)
    statistic = 0.0
    for row_index, row in enumerate(matrix):
        for column_index, observed in enumerate(row):
            expected = row_totals[row_index] * column_totals[column_index] / total
            if expected <= 0:
                raise ValueError("Chi-square expected count must be positive")
            statistic += (observed - expected) ** 2 / expected
    degrees = (len(matrix) - 1) * (len(STATUS_ORDER) - 1)
    if degrees % 2:
        raise ValueError("The built-in chi-square survival function expects even degrees of freedom")
    half = statistic / 2.0
    p_value = math.exp(-half) * sum(
        half**index / math.factorial(index) for index in range(degrees // 2)
    )
    cramers_v = math.sqrt(
        statistic / (total * min(len(matrix) - 1, len(STATUS_ORDER) - 1))
    )
    return ChiSquareResult(statistic, degrees, p_value, cramers_v)


def _coverage_code(lesson_id: str) -> str:
    match = re.search(r"-B0*(\d+)([AB]?)$", lesson_id, flags=re.IGNORECASE)
    if not match:
        raise ValueError(f"Cannot extract canonical lesson code from {lesson_id!r}")
    return f"{int(match.group(1))}{match.group(2).upper()}"


def _bloom_value(value: str) -> str:
    folded = value.strip().casefold()
    for label in ("Nhận biết", "Thông hiểu", "Vận dụng"):
        if folded.startswith(label.casefold()):
            return label
    return "Không rõ"


def _pass_coverage_rows(data: BundleData) -> list[list[object]]:
    lesson_metadata: dict[tuple[str, str], dict[str, str]] = {}
    for grade, rows in data.coverage_by_grade.items():
        for row in rows:
            key = (grade, _coverage_code(row["lesson_id"]))
            if key in lesson_metadata:
                raise ValueError(f"Duplicate canonical lesson metadata: {key}")
            lesson_metadata[key] = row

    output: list[list[object]] = []
    for grade in EXPECTED_GRADE_COUNTS:
        pass_ids = [
            sample_id
            for sample_id, normalized in data.normalized_by_id.items()
            if normalized["grade"] == grade
            and data.quality_by_id[sample_id]["quality_decision"] == "pass"
        ]
        pass_total = len(pass_ids)
        dimensions: dict[str, Counter[tuple[str, str]]] = {
            "topic": Counter(),
            "lesson": Counter(),
            "bloom_band": Counter(),
        }
        for sample_id in pass_ids:
            normalized = data.normalized_by_id[sample_id]
            code = lesson_code(normalized["lesson"])
            metadata = lesson_metadata.get((grade, code))
            if metadata is None:
                raise ValueError(f"No canonical coverage metadata joins to {sample_id}: {grade}/{code}")
            dimensions["topic"][(metadata["topic_id"], metadata["topic_label"])] += 1
            dimensions["lesson"][(metadata["lesson_id"], metadata["lesson_label"])] += 1
            bloom = _bloom_value(normalized.get("bloom_level", ""))
            dimensions["bloom_band"][(bloom, bloom)] += 1

        dimension_labels = {
            "topic": "Chủ đề",
            "lesson": "Bài học",
            "bloom_band": "Mức nhận thức",
        }
        for dimension in ("topic", "lesson", "bloom_band"):
            counts = dimensions[dimension]
            if sum(counts.values()) != pass_total:
                raise ValueError(f"Pass coverage does not sum to pass total: grade {grade}/{dimension}")
            for (value_id, value_label), count in sorted(counts.items()):
                output.append(
                    [
                        grade,
                        f"Lớp {grade}",
                        dimension,
                        dimension_labels[dimension],
                        value_id,
                        value_label,
                        count,
                        round_percentage_fraction_half_up(count, pass_total) if pass_total else 0.0,
                        pass_total,
                    ]
                )
    return output


def _write_csv(path: Path, fields: Sequence[str], rows: Iterable[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="raise", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def _write_single_sheet_workbook(
    path: Path,
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    widths: dict[int, float],
    percentage_columns: Sequence[int] = (),
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    _append_rows(worksheet, [headers])
    _append_rows(worksheet, rows)
    _style_table(worksheet)
    _set_widths(worksheet, widths)
    for column in percentage_columns:
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(row_number, column).number_format = "0.00%"
    workbook.save(path)
    workbook.close()


def _readme_text() -> str:
    return """# Bộ kết quả rà soát Phase 1 gửi HNMU

Bộ hồ sơ này trình bày kết quả rà soát 1.050 mẫu hội thoại Tin học lớp 6–9 theo từng loại tài liệu.

## Nên đọc theo thứ tự nào?

1. Đọc `01_bao_cao_tong_quan.md` để xem kết quả chung và so sánh giữa các khối.
2. Mở `04_thong_ke_pass_reject_theo_khoi.xlsx` để xem số lượng và tỷ lệ từng trạng thái.
3. Mở `06_ket_qua_cham_tong_the_tung_mau.csv` để tra trạng thái của từng mẫu.
4. Dùng `03_du_lieu_tho_sau_chuan_hoa.csv` để đọc đầy đủ câu hỏi, đáp án SGV và hội thoại.
5. Khi cần xem sâu hơn, dùng các file 07–09 theo cùng mã mẫu.

## Ý nghĩa các file

- `02_checklist_tieu_chi.xlsx`: danh sách 18 tiêu chí dùng cho mọi mẫu.
- `03_du_lieu_tho_sau_chuan_hoa.csv`: toàn bộ 1.050 mẫu sau chuẩn hóa, giữ nguyên nội dung và thông tin truy vết.
- `04_thong_ke_pass_reject_theo_khoi.xlsx`: số lượng và tỷ lệ trạng thái theo từng khối và toàn bộ dữ liệu.
- `05_thong_ke_do_phu_mau_pass_theo_khoi.xlsx`: độ phủ được tính lại riêng từ các mẫu có trạng thái `pass`.
- `06_ket_qua_cham_tong_the_tung_mau.csv`: kết quả chấm tổng thể của từng mẫu.
- `07_ket_qua_cham_chi_tiet_tung_tieu_chi.csv`: thành phần bổ sung thứ nhất của kết quả chấm chi tiết, gồm kết quả agent chấm theo từng tiêu chí.
- `08_mau_thieu_sai_truong_du_lieu.csv`: thành phần bổ sung thứ hai, gồm các cảnh báo thiếu hoặc sai trường dữ liệu.
- `09_ung_vien_trung_lap.csv`: thành phần bổ sung thứ ba, gồm các cặp mẫu cần xem lại khả năng trùng lặp.

Ba file 07–09 bổ sung cho kết quả tổng thể trong file 06. Chúng không phải ba lần chấm độc lập.

## Cách truy vết

Ví dụ đúng: tìm `sample_id` trong file 06, rồi dùng cùng mã đó để mở nội dung mẫu ở file 03 và xem các tiêu chí liên quan ở file 07.

Ví dụ không đúng: coi trạng thái `failed` là quyết định loại bỏ vĩnh viễn. Trạng thái này chỉ cho biết mẫu chưa nên dùng ở lượt hiện tại; giáo viên HNMU/UET vẫn giữ quyền xác nhận chuyên môn.

Khi cần phản hồi, ghi rõ `sample_id`, `grade`, nội dung cần sửa và lý do. Các trường `source_file` và `source_row_number` giúp đối chiếu với dữ liệu đã cung cấp.
"""


def _overview_report(data: BundleData) -> str:
    counts = _status_counts(data)
    test = _chi_square_result(data)
    lines = [
        "# Báo cáo tổng quan kết quả rà soát Phase 1",
        "",
        "## Quy mô và trạng thái",
        "",
        "| Khối | Tổng mẫu | Pass | Need human review | Failed | Non-pass |",
        "|---:|---:|---:|---:|---:|---:|",
    ]
    for grade in EXPECTED_GRADE_COUNTS:
        total = sum(counts[grade].values())
        non_pass = counts[grade]["need_human_review"] + counts[grade]["failed"]
        lines.append(
            f"| {grade} | {total} | {counts[grade]['pass']} ({format_percentage_half_up(counts[grade]['pass'], total)}) "
            f"| {counts[grade]['need_human_review']} ({format_percentage_half_up(counts[grade]['need_human_review'], total)}) "
            f"| {counts[grade]['failed']} ({format_percentage_half_up(counts[grade]['failed'], total)}) "
            f"| {non_pass} ({format_percentage_half_up(non_pass, total)}) |"
        )
    total = sum(counts["all"].values())
    non_pass = counts["all"]["need_human_review"] + counts["all"]["failed"]
    lines.extend(
        [
            f"| **Toàn bộ** | **{total}** | **{counts['all']['pass']} ({format_percentage_half_up(counts['all']['pass'], total)})** "
            f"| **{counts['all']['need_human_review']} ({format_percentage_half_up(counts['all']['need_human_review'], total)})** "
            f"| **{counts['all']['failed']} ({format_percentage_half_up(counts['all']['failed'], total)})** "
            f"| **{non_pass} ({format_percentage_half_up(non_pass, total)})** |",
            "",
        ]
    )
    pass_rates = {
        grade: counts[grade]["pass"] / sum(counts[grade].values())
        for grade in EXPECTED_GRADE_COUNTS
    }
    highest = max(pass_rates, key=pass_rates.get)
    lowest = min(pass_rates, key=pass_rates.get)
    lines.extend(
        [
            "Tỷ lệ trạng thái không đồng đều giữa các khối. "
            f"Khối {highest} có tỷ lệ pass cao nhất ({format_fraction_percentage_half_up(pass_rates[highest])}); "
            f"khối {lowest} có tỷ lệ pass thấp nhất ({format_fraction_percentage_half_up(pass_rates[lowest])}).",
            "",
            "## Kiểm định sự khác biệt giữa các khối",
            "",
            "Kiểm định chi-square được thực hiện trên bảng 4 khối × 3 trạng thái loại trừ nhau: "
            "`pass`, `need_human_review` và `failed`. Chỉ số `non_pass` không được đưa thêm "
            "vào bảng kiểm định vì đây là tổng của hai trạng thái sau.",
            "",
            f"- Chi-square: {test.statistic:.6f}",
            f"- Bậc tự do: {test.degrees_of_freedom}",
            f"- p-value: {test.p_value:.8g}",
            f"- Cramér’s V: {test.cramers_v:.6f}",
            "",
            (
                "Kết quả cho thấy có bằng chứng thống kê về mối liên hệ giữa khối lớp và trạng thái rà soát."
                if test.p_value < 0.05
                else "Kết quả chưa cho thấy bằng chứng thống kê về mối liên hệ giữa khối lớp và trạng thái rà soát."
            ),
            "Cramér’s V mô tả độ mạnh của mối liên hệ; chỉ số này không thay thế nhận định chuyên môn của giáo viên.",
            "",
            "## Độ phủ của nhóm pass",
            "",
            "Độ phủ trong file 05 được tính lại từ phép nối theo `sample_id`: chỉ giữ các mẫu có "
            "`quality_decision = pass`, sau đó đếm riêng theo khối, chủ đề, bài học và mức nhận thức. "
            "Không sử dụng tỷ lệ độ phủ của toàn bộ dữ liệu để đại diện cho nhóm pass.",
            "",
            "## Nguồn canonical đã dùng",
            "",
            "| STT | Đường dẫn | Số bản ghi/dòng | SHA-256 |",
            "|---:|---|---:|---|",
        ]
    )
    for index, (path, digest) in enumerate(data.source_hashes.items(), start=1):
        lines.append(
            f"| {index} | `{_display_path(path)}` | {data.source_record_counts[path]} | `{digest}` |"
        )
    lines.extend(
        [
            "",
            "## Phụ lục — Nội dung báo cáo canonical trước khi đóng gói lại",
            "",
            data.teacher_report_text.rstrip(),
            "",
        ]
    )
    return "\n".join(lines)


def _build_files(data: BundleData, directory: Path) -> None:
    directory.mkdir(parents=True, exist_ok=False)
    (directory / "README.md").write_text(_readme_text(), encoding="utf-8")
    (directory / "01_bao_cao_tong_quan.md").write_text(_overview_report(data), encoding="utf-8")

    _write_single_sheet_workbook(
        directory / "02_checklist_tieu_chi.xlsx",
        "checklist_tieu_chi",
        CRITERIA_HEADERS,
        _criteria_rows(data),
        {1: 22, 2: 24, 3: 55, 4: 22, 5: 16},
    )
    _write_csv(directory / "03_du_lieu_tho_sau_chuan_hoa.csv", NORMALIZED_FIELDS, _normalized_rows(data))
    _write_single_sheet_workbook(
        directory / "04_thong_ke_pass_reject_theo_khoi.xlsx",
        "pass_reject_theo_khoi",
        STATUS_HEADERS,
        _status_rows(data),
        {1: 10, 2: 16, 3: 24, 4: 34, 5: 14, 6: 16, 7: 18},
        percentage_columns=(6,),
    )
    _write_single_sheet_workbook(
        directory / "05_thong_ke_do_phu_mau_pass_theo_khoi.xlsx",
        "do_phu_mau_pass",
        COVERAGE_HEADERS,
        _pass_coverage_rows(data),
        {1: 10, 2: 16, 3: 20, 4: 22, 5: 22, 6: 65, 7: 14, 8: 22, 9: 20},
        percentage_columns=(8,),
    )
    _write_csv(directory / "06_ket_qua_cham_tong_the_tung_mau.csv", QUALITY_FIELDS, _quality_rows(data))
    _write_csv(
        directory / "07_ket_qua_cham_chi_tiet_tung_tieu_chi.csv",
        CHECKLIST_FIELDS,
        _checklist_rows(data),
    )
    _write_csv(directory / "08_mau_thieu_sai_truong_du_lieu.csv", MISSING_FIELDS, _missing_rows(data))
    _write_csv(directory / "09_ung_vien_trung_lap.csv", DUPLICATE_FIELDS, _duplicate_rows(data))


def _read_csv(path: Path, fields: Sequence[str]) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if tuple(reader.fieldnames or ()) != tuple(fields):
            raise ValueError(f"Unexpected columns in {path.name}: {reader.fieldnames}")
        return list(reader)


def _validate_csv_exact(
    path: Path,
    fields: Sequence[str],
    expected: Sequence[dict[str, str]],
) -> list[dict[str, str]]:
    actual = _read_csv(path, fields)
    if actual != list(expected):
        raise ValueError(f"{path.name} differs from the canonical joined rows")
    return actual


def _workbook_rows(path: Path, expected_sheet: str, headers: Sequence[str]) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if workbook.sheetnames != [expected_sheet]:
        raise ValueError(f"{path.name} must contain exactly one sheet named {expected_sheet}")
    worksheet = workbook[expected_sheet]
    actual_headers = tuple(cell.value for cell in worksheet[1])
    if actual_headers != tuple(headers):
        raise ValueError(f"Unexpected columns in {path.name}: {actual_headers}")
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    workbook.close()
    return rows


def _assert_workbook_rows(
    path: Path,
    sheet: str,
    headers: Sequence[str],
    expected: Sequence[Sequence[object]],
) -> None:
    actual = _workbook_rows(path, sheet, headers)
    expected_rows = [tuple(row) for row in expected]
    if len(actual) != len(expected_rows):
        raise ValueError(f"{path.name} has {len(actual)} rows; expected {len(expected_rows)}")
    for row_index, (actual_row, expected_row) in enumerate(zip(actual, expected_rows), start=2):
        for column_index, (actual_value, expected_value) in enumerate(zip(actual_row, expected_row), start=1):
            if isinstance(expected_value, float):
                if not isinstance(actual_value, (int, float)) or not math.isclose(
                    float(actual_value), expected_value, rel_tol=1e-12, abs_tol=1e-12
                ):
                    raise ValueError(f"{path.name} changed numeric cell {row_index},{column_index}")
            elif actual_value != expected_value:
                raise ValueError(f"{path.name} changed cell {row_index},{column_index}")


def _validate_flat_phase1_teacher_bundle_v2(
    experiment_dir: Path,
    bundle_dir: Path,
    *,
    expected_source_hashes: dict[Path, str] | None = None,
) -> dict[str, object]:
    """Validate every v2 deliverable against the canonical joined data."""

    data = load_canonical_bundle_data(experiment_dir)
    if expected_source_hashes is not None and data.source_hashes != expected_source_hashes:
        raise ValueError("Canonical source hashes changed during v2 build")
    if not bundle_dir.is_dir():
        raise FileNotFoundError(f"Missing v2 bundle directory: {bundle_dir}")
    actual_names = {path.name for path in bundle_dir.iterdir()}
    if actual_names != set(V2_OUTPUT_NAMES):
        raise ValueError(
            f"V2 bundle file set differs: missing={sorted(set(V2_OUTPUT_NAMES) - actual_names)}, "
            f"extra={sorted(actual_names - set(V2_OUTPUT_NAMES))}"
        )

    normalized = _validate_csv_exact(
        bundle_dir / V2_OUTPUT_NAMES[3], NORMALIZED_FIELDS, _normalized_rows(data)
    )
    quality = _validate_csv_exact(
        bundle_dir / V2_OUTPUT_NAMES[6], QUALITY_FIELDS, _quality_rows(data)
    )
    checklist = _validate_csv_exact(
        bundle_dir / V2_OUTPUT_NAMES[7], CHECKLIST_FIELDS, _checklist_rows(data)
    )
    missing = _validate_csv_exact(
        bundle_dir / V2_OUTPUT_NAMES[8], MISSING_FIELDS, _missing_rows(data)
    )
    duplicates = _validate_csv_exact(
        bundle_dir / V2_OUTPUT_NAMES[9], DUPLICATE_FIELDS, _duplicate_rows(data)
    )

    normalized_ids = [row["sample_id"] for row in normalized]
    quality_ids = [row["sample_id"] for row in quality]
    expected_ids = set(data.normalized_by_id)
    if len(normalized_ids) != 1050 or len(set(normalized_ids)) != 1050:
        raise ValueError("Normalized deliverable must contain 1,050 unique sample_id values")
    if set(normalized_ids) != expected_ids or set(quality_ids) != expected_ids or len(quality_ids) != 1050:
        raise ValueError("Sample-level v2 files lose, add, or duplicate sample_id values")
    for row in quality:
        if row["grade"] != data.normalized_by_id[row["sample_id"]]["grade"]:
            raise ValueError(f"Quality grade mismatch: {row['sample_id']}")
        if row["quality_decision"] not in ALLOWED_DECISIONS:
            raise ValueError(f"Unexpected quality decision: {row['sample_id']}")

    pairs = [(row["sample_id"], row["criterion_id"]) for row in checklist]
    if len(pairs) != 1050 * EXPECTED_CRITERIA_PER_SAMPLE or len(set(pairs)) != len(pairs):
        raise ValueError("Detailed criterion deliverable has missing or duplicate keys")
    if {sample_id for sample_id, _ in pairs} != expected_ids:
        raise ValueError("Detailed criterion deliverable has a different sample set")
    if any(row["grade"] != data.normalized_by_id[row["sample_id"]]["grade"] for row in checklist):
        raise ValueError("Detailed criterion deliverable has a grade mismatch")
    if any(row["sample_id"] not in expected_ids for row in missing):
        raise ValueError("Missing-field deliverable contains an unknown sample")
    for row in duplicates:
        if row["sample_id_a"] not in expected_ids or row["sample_id_b"] not in expected_ids:
            raise ValueError("Duplicate deliverable contains an unknown sample")
        if row["grade"] != data.normalized_by_id[row["sample_id_a"]]["grade"]:
            raise ValueError("Duplicate deliverable has a grade mismatch")

    _assert_workbook_rows(
        bundle_dir / V2_OUTPUT_NAMES[2],
        "checklist_tieu_chi",
        CRITERIA_HEADERS,
        _criteria_rows(data),
    )
    _assert_workbook_rows(
        bundle_dir / V2_OUTPUT_NAMES[4],
        "pass_reject_theo_khoi",
        STATUS_HEADERS,
        _status_rows(data),
    )
    coverage_rows = _pass_coverage_rows(data)
    _assert_workbook_rows(
        bundle_dir / V2_OUTPUT_NAMES[5],
        "do_phu_mau_pass",
        COVERAGE_HEADERS,
        coverage_rows,
    )

    for grade in EXPECTED_GRADE_COUNTS:
        expected_pass = sum(
            row["grade"] == grade and data.quality_by_id[sample_id]["quality_decision"] == "pass"
            for sample_id, row in data.normalized_by_id.items()
        )
        for dimension in ("topic", "lesson", "bloom_band"):
            actual_count = sum(
                int(row[6])
                for row in coverage_rows
                if row[0] == grade and row[2] == dimension
            )
            if actual_count != expected_pass:
                raise ValueError(f"Pass coverage mismatch: grade {grade}/{dimension}")

    readme = (bundle_dir / V2_OUTPUT_NAMES[0]).read_text(encoding="utf-8")
    if not all(name in readme for name in V2_OUTPUT_NAMES[1:]):
        raise ValueError("README does not explain every v2 deliverable")
    if "Ba file 07–09 bổ sung" not in readme:
        raise ValueError("README does not explain the supplementary 07–09 relationship")
    overview = (bundle_dir / V2_OUTPUT_NAMES[1]).read_text(encoding="utf-8")
    for marker in ("Chi-square:", "Cramér’s V:", "Phụ lục — Nội dung báo cáo canonical"):
        if marker not in overview:
            raise ValueError(f"Overview report is missing: {marker}")

    status_counts = _status_counts(data)
    return {
        "status": "ok",
        "output_paths": [(bundle_dir / name).as_posix() for name in V2_OUTPUT_NAMES],
        "grade_counts": dict(EXPECTED_GRADE_COUNTS),
        "status_counts": status_counts,
        "checklist_row_count": len(checklist),
        "missing_row_count": len(missing),
        "duplicate_row_count": len(duplicates),
        "chi_square": _chi_square_result(data).__dict__,
        "source_count": len(data.source_hashes),
        "source_hashes": {_display_path(path): digest for path, digest in data.source_hashes.items()},
        "read_paths": sorted(_display_path(path) for path in data.read_paths),
    }


def _build_flat_phase1_teacher_bundle_v2(experiment_dir: Path, bundle_dir: Path) -> V2BuildSummary:
    """Build the v2 bundle atomically and refuse every overwrite."""

    if bundle_dir.exists():
        raise FileExistsError(f"V2 bundle already exists; refusing to overwrite: {bundle_dir}")
    data = load_canonical_bundle_data(experiment_dir)
    initial_hashes = dict(data.source_hashes)
    bundle_dir.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix=".hnmu-teacher-bundle-v2-", dir=bundle_dir.parent) as temp:
        staged = Path(temp) / "bundle"
        _build_files(data, staged)
        _validate_flat_phase1_teacher_bundle_v2(
            experiment_dir,
            staged,
            expected_source_hashes=initial_hashes,
        )
        staged.replace(bundle_dir)

    validation = _validate_flat_phase1_teacher_bundle_v2(
        experiment_dir,
        bundle_dir,
        expected_source_hashes=initial_hashes,
    )
    return V2BuildSummary(
        output_paths=tuple(bundle_dir / name for name in V2_OUTPUT_NAMES),
        grade_counts=dict(EXPECTED_GRADE_COUNTS),
        status_counts=validation["status_counts"],
        chi_square=ChiSquareResult(**validation["chi_square"]),
        source_hashes=validation["source_hashes"],
        read_paths=tuple(validation["read_paths"]),
    )


def validate_phase1_teacher_bundle_v2(
    experiment_dir: Path,
    bundle_dir: Path,
    *,
    expected_source_hashes: dict[Path, str] | None = None,
) -> dict[str, object]:
    """Validate the grade-partitioned public v2 bundle."""

    from edu_benchmark.dialogue_audit.teacher_bundle_v2_partitioned import (
        validate_partitioned_phase1_teacher_bundle_v2,
    )

    return validate_partitioned_phase1_teacher_bundle_v2(
        experiment_dir,
        bundle_dir,
        expected_source_hashes=expected_source_hashes,
    )


def build_phase1_teacher_bundle_v2(experiment_dir: Path, bundle_dir: Path) -> V2BuildSummary:
    """Build the grade-partitioned public v2 bundle."""

    from edu_benchmark.dialogue_audit.teacher_bundle_v2_partitioned import (
        build_partitioned_phase1_teacher_bundle_v2,
    )

    return build_partitioned_phase1_teacher_bundle_v2(experiment_dir, bundle_dir)
