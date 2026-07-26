"""HNMU summaries for focused root reading and grade-level comparison."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from edu_benchmark.dialogue_audit.fragment_analysis_hnmu import (
    GRADE_ANALYSIS_NAME,
    GRADE_APPENDIX_NAME,
    GRADE_APPENDIX_SHEET,
    ROOT_ANALYSIS_NAME,
    ROOT_APPENDIX_NAME,
    ROOT_APPENDIX_SHEET,
    TECHNICAL_DATA_HEADER_ROW,
    TECHNICAL_FIELDS,
    analysis_key,
    expected_technical_display_rows,
    read_technical_appendix_workbook,
    validate_summary_keys,
    write_technical_appendix_workbook,
)
from edu_benchmark.dialogue_audit.fragment_score_analysis import (
    FRAGMENT_METRICS,
    _summary_row,
)

ROOT_SUMMARY_SHEET = "tom_tat_giua_cac_khoi"
GRADE_SUMMARY_SHEET = "tom_tat_lop"
ROOT_FOCUS_TITLE = "FRAGMENT ĐẦY ĐỦ HƠN CÓ ĐI KÈM TỶ LỆ ĐẠT CAO HƠN KHÔNG?"
ROOT_FOCUS_QUESTION = (
    "Các mẫu được đối chiếu với tỷ lệ tiêu chí có dẫn fragment cao hơn "
    "có tỷ lệ đạt chính thức cao hơn không?"
)
SUMMARY_HEADERS = (
    "Nội dung về fragment",
    "Kết quả chấm được xem xét",
    "Khi xem tất cả mẫu",
    "Khi so các mẫu trong cùng nhóm chấm",
    "Kết luận dễ hiểu",
)
SUMMARY_BLOCK_A_HEADER_ROW = 7
SUMMARY_BLOCK_A_START_ROW = 8
SUMMARY_BLOCK_B_HEADER_ROW = 14
SUMMARY_BLOCK_B_START_ROW = 15
SUMMARY_LAST_ROW = 18
SUMMARY_DATA_HEADER_ROW = SUMMARY_BLOCK_A_HEADER_ROW

SUMMARY_TITLE = (
    "TÓM TẮT MỐI LIÊN HỆ GIỮA VIỆC SỬ DỤNG DẪN CHỨNG HỌC LIỆU VÀ KẾT QUẢ CHẤM"
)
METRIC_LABELS = {
    "fragment_row_count": "Số tiêu chí có dẫn fragment",
    "fragment_reference_count": "Tổng lượt dẫn fragment",
    "unique_fragment_count": "Số fragment khác nhau",
    "fragment_criterion_coverage": "Tỷ lệ tiêu chí có dẫn fragment",
}
OUTCOME_LABELS = {
    "official_pass": "Trạng thái đạt chính thức",
    "checklist_pass_rate": "Tỷ lệ tiêu chí đạt",
}


def _estimable(row: dict[str, object] | None) -> bool:
    return bool(
        row
        and row.get("estimable")
        and isinstance(row.get("statistic_value"), (int, float))
    )


def _has_evidence(row: dict[str, object] | None) -> bool:
    return bool(
        _estimable(row)
        and isinstance(row.get("p_value"), (int, float))
        and float(row["p_value"]) < 0.05
    )


def _near_zero(row: dict[str, object] | None) -> bool:
    return bool(_estimable(row) and abs(float(row["statistic_value"])) <= 0.05)


def _combined_result(row: dict[str, object] | None) -> str:
    if not _estimable(row):
        return "Không thể ước lượng do dữ liệu không đủ biến thiên"
    value = float(row["statistic_value"])
    absolute = abs(value)
    if absolute <= 0.05:
        association = "Gần như không có liên hệ"
    else:
        direction = "dương" if value > 0 else "âm"
        if absolute < 0.20:
            strength = "rất yếu"
        elif absolute < 0.40:
            strength = "yếu"
        elif absolute < 0.60:
            strength = "vừa"
        else:
            strength = "mạnh"
        association = f"Liên hệ {direction} {strength}"
    evidence = "có bằng chứng" if _has_evidence(row) else "chưa có bằng chứng"
    return f"{association}; {evidence} (r = {value:.2f})"


def _plain_interpretation(
    crude: dict[str, object],
    adjusted: dict[str, object],
) -> str:
    if not _estimable(adjusted):
        return "Không thể ước lượng sau điều chỉnh; kết quả trước điều chỉnh không chứng minh mối liên hệ độc lập."
    crude_evidence = _has_evidence(crude)
    adjusted_evidence = _has_evidence(adjusted)
    if crude_evidence and not adjusted_evidence:
        return "Mối liên hệ quan sát được không còn sau điều chỉnh."
    if not crude_evidence and not adjusted_evidence:
        return "Không thấy bằng chứng về mối liên hệ ở cả hai phân tích."
    if not crude_evidence and adjusted_evidence:
        return "Mối liên hệ chỉ xuất hiện sau điều chỉnh và cần được diễn giải thận trọng."
    reversed_direction = (
        not _near_zero(crude)
        and not _near_zero(adjusted)
        and float(crude["statistic_value"]) * float(adjusted["statistic_value"]) < 0
    )
    if reversed_direction:
        return "Mối liên hệ đổi chiều sau điều chỉnh nên kết quả chưa ổn định."
    return "Mối liên hệ vẫn được ghi nhận sau điều chỉnh, nhưng không chứng minh quan hệ nhân quả."


def build_hnmu_summary_rows(
    rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None = None,
) -> list[dict[str, object]]:
    adjustment = (
        "adjusted_for_grade_and_auditor_group"
        if pooled
        else "adjusted_for_auditor_group"
    )
    output: list[dict[str, object]] = []
    for family, outcome in (
        ("fragment_vs_official_pass", "official_pass"),
        ("fragment_vs_checklist_pass_rate", "checklist_pass_rate"),
    ):
        for metric in FRAGMENT_METRICS:
            crude = _summary_row(rows, family, metric, "crude")
            adjusted = _summary_row(rows, family, metric, adjustment)
            if crude is None or adjusted is None:
                raise ValueError(f"Missing summary pair for {outcome}/{metric}")
            output.append(
                {
                    "outcome": outcome,
                    "outcome_label": OUTCOME_LABELS[outcome],
                    "metric": metric,
                    "metric_label": METRIC_LABELS[metric],
                    "crude_sample_count": int(crude["sample_count"]),
                    "adjusted_sample_count": int(adjusted["sample_count"]),
                    "crude_statistic": crude["statistic_value"],
                    "adjusted_statistic": adjusted["statistic_value"],
                    "crude_evidence": _has_evidence(crude),
                    "adjusted_evidence": _has_evidence(adjusted),
                    "adjusted_estimable": _estimable(adjusted),
                    "crude_result": _combined_result(crude),
                    "adjusted_result": _combined_result(adjusted),
                    "plain_interpretation": _plain_interpretation(crude, adjusted),
                    "controlled_factors": (
                        "Khối lớp và nhóm chấm"
                        if pooled
                        else f"Nhóm chấm trong lớp {grade}"
                    ),
                    "analysis_key": analysis_key(outcome, metric),
                }
            )
    return output



def _summary_result_is_clear(row: dict[str, object], prefix: str) -> bool:
    if prefix == "adjusted" and not row.get("adjusted_estimable"):
        return False
    value = row.get(f"{prefix}_statistic")
    return bool(
        row.get(f"{prefix}_evidence")
        and isinstance(value, (int, float))
        and abs(float(value)) > 0.05
    )


def _grade_trend_label(row: dict[str, object], prefix: str) -> str:
    if prefix == "adjusted" and not row.get("adjusted_estimable"):
        return "Không đủ dữ liệu để kết luận"
    if not _summary_result_is_clear(row, prefix):
        return "Không thấy khác biệt rõ ràng"
    return (
        "Có xu hướng cao hơn"
        if float(row[f"{prefix}_statistic"]) > 0
        else "Có xu hướng thấp hơn"
    )


def _grade_plain_conclusion(row: dict[str, object]) -> str:
    if not row.get("adjusted_estimable"):
        return "Không đủ dữ liệu phù hợp để kết luận."
    crude_clear = _summary_result_is_clear(row, "crude")
    adjusted_clear = _summary_result_is_clear(row, "adjusted")
    if not crude_clear and not adjusted_clear:
        return "Chưa thấy mối liên hệ rõ ràng."
    if crude_clear and not adjusted_clear:
        return (
            "Sự khác biệt không còn rõ ràng khi so sánh các mẫu trong cùng "
            "nhóm chấm."
        )
    if not crude_clear and adjusted_clear:
        return (
            "Xu hướng chỉ xuất hiện khi so sánh trong cùng nhóm chấm và cần "
            "được diễn giải thận trọng."
        )
    crude = float(row["crude_statistic"])
    adjusted = float(row["adjusted_statistic"])
    if crude * adjusted < 0:
        return (
            "Hai cách so sánh cho kết quả trái chiều nên chưa thể đưa ra "
            "kết luận ổn định."
        )
    return (
        "Xu hướng vẫn xuất hiện khi so sánh các mẫu trong cùng nhóm chấm, "
        "nhưng không chứng minh quan hệ nguyên nhân."
    )


def grade_summary_display_rows(
    rows: Iterable[dict[str, object]],
) -> list[tuple[object, ...]]:
    output: list[tuple[object, ...]] = []
    for row in rows:
        conclusion = _grade_plain_conclusion(row)
        if len(conclusion.split()) > 25:
            raise ValueError(
                "Grade summary conclusion exceeds 25 words: "
                f"{row['analysis_key']}"
            )
        output.append(
            (
                METRIC_LABELS[str(row["metric"])],
                OUTCOME_LABELS[str(row["outcome"])],
                _grade_trend_label(row, "crude"),
                _grade_trend_label(row, "adjusted"),
                conclusion,
            )
        )
    return output

def _same_direction_with_evidence(row: dict[str, object]) -> bool:
    if not _summary_result_is_clear(row, "crude"):
        return False
    if not _summary_result_is_clear(row, "adjusted"):
        return False
    crude = float(row["crude_statistic"])
    adjusted = float(row["adjusted_statistic"])
    return crude * adjusted > 0


def hnmu_main_conclusion(
    summary_rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None = None,
) -> str:
    if len(summary_rows) != 8:
        raise ValueError(f"Expected 8 HNMU summary rows, found {len(summary_rows)}")
    comparison = (
        "cùng khối lớp và cùng nhóm chấm"
        if pooled
        else "cùng nhóm chấm"
    )
    scope = "toàn bộ dữ liệu" if pooled else f"lớp {grade}"
    if all(not row["adjusted_estimable"] for row in summary_rows):
        return (
            f"Trong {scope}, không đủ dữ liệu phù hợp để kết luận khi so sánh "
            f"các mẫu trong {comparison}."
        )
    stable = [row for row in summary_rows if _same_direction_with_evidence(row)]
    if stable:
        return (
            f"Trong {scope}, một số xu hướng vẫn xuất hiện khi so sánh các mẫu "
            f"trong {comparison}; kết quả không chứng minh quan hệ nguyên nhân."
        )
    if any(
        _summary_result_is_clear(row, "crude")
        and not _summary_result_is_clear(row, "adjusted")
        for row in summary_rows
    ):
        return (
            f"Trong {scope}, một số khác biệt quan sát được không còn rõ ràng "
            f"khi so sánh các mẫu trong {comparison}."
        )
    return f"Trong {scope}, chưa thấy mối liên hệ rõ ràng."


def _noteworthy_text(summary_rows: Sequence[dict[str, object]]) -> str:
    stable = [row for row in summary_rows if _same_direction_with_evidence(row)]
    if not stable:
        return (
            "Chưa có kết quả nào cho thấy cùng một xu hướng rõ ràng ở cả hai "
            "cách so sánh."
        )
    return (
        f"Có {len(stable)}/8 kết quả cho thấy cùng một xu hướng rõ ràng ở cả "
        "hai cách so sánh."
    )


def _sample_limit_text(
    summary_rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None,
) -> str:
    totals = {int(row["crude_sample_count"]) for row in summary_rows}
    if len(totals) != 1:
        raise ValueError("Crude sample counts are inconsistent within HNMU summary")
    total = totals.pop()
    total_label = f"{total:,}".replace(",", ".")
    counts_by_metric: dict[str, int] = {}
    for metric in FRAGMENT_METRICS:
        counts = {
            int(row["adjusted_sample_count"])
            for row in summary_rows
            if row["metric"] == metric
        }
        if len(counts) != 1:
            raise ValueError(f"Adjusted sample count differs by outcome for {metric}")
        counts_by_metric[metric] = counts.pop()
    count_frequency = Counter(counts_by_metric.values())
    common_count = count_frequency.most_common(1)[0][0]
    exceptions = [
        (METRIC_LABELS[metric], count)
        for metric, count in counts_by_metric.items()
        if count != common_count
    ]
    comparison = (
        "cùng khối lớp và cùng nhóm chấm"
        if pooled
        else "cùng nhóm chấm"
    )
    text = (
        f"Khi so sánh các mẫu trong {comparison}, phần lớn cách đo sử dụng "
        f"{common_count}/{total_label} mẫu"
    )
    if exceptions:
        details = "; ".join(
            f"{label} sử dụng {count}/{total_label} mẫu"
            for label, count in exceptions
        )
        text += f"; riêng {details}"
    return text + ". Kết quả chỉ mô tả các yếu tố đi cùng nhau."


def summary_intro_blocks(
    summary_rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None = None,
) -> tuple[tuple[str, str], ...]:
    return (
        (
            "KẾT LUẬN CHUNG",
            hnmu_main_conclusion(summary_rows, pooled=pooled, grade=grade),
        ),
        ("KẾT QUẢ ĐÁNG CHÚ Ý", _noteworthy_text(summary_rows)),
        (
            "GIỚI HẠN KHI DIỄN GIẢI",
            _sample_limit_text(summary_rows, pooled=pooled, grade=grade),
        ),
    )


def _root_focus_row(
    summary_rows: Sequence[dict[str, object]],
) -> dict[str, object]:
    matches = [
        row
        for row in summary_rows
        if row.get("outcome") == "official_pass"
        and row.get("metric") == "fragment_criterion_coverage"
    ]
    if len(matches) != 1:
        raise ValueError(
            "Root HNMU summary requires official_pass × fragment_criterion_coverage"
        )
    row = matches[0]
    if row.get("analysis_key") != "FRG-OP-04":
        raise ValueError("Unexpected analysis key for root HNMU focus question")
    return row


def _root_focus_conclusion(row: dict[str, object]) -> str:
    if not row["adjusted_estimable"]:
        return (
            "Kết luận: Chưa thể trả lời liệu mẫu được dẫn fragment đầy đủ hơn có tỷ lệ đạt cao hơn hay không. "
            "Dữ liệu hiện có chưa cho phép so sánh rõ các mẫu cùng khối lớp và cùng nhóm chấm."
        )
    if row["adjusted_evidence"] and float(row["adjusted_statistic"]) > 0:
        return (
            "Kết luận: Có dấu hiệu cho thấy mẫu được dẫn fragment đầy đủ hơn có tỷ lệ đạt cao hơn. "
            "Xu hướng này vẫn được ghi nhận khi so sánh các mẫu cùng khối lớp và cùng nhóm chấm."
        )
    return (
        "Kết luận: Chưa thể khẳng định mẫu được dẫn fragment đầy đủ hơn có tỷ lệ đạt cao hơn. "
        "Khi so sánh các mẫu cùng khối lớp và cùng nhóm chấm, không còn thấy sự khác biệt rõ ràng."
    )


def _root_focus_explanation(row: dict[str, object]) -> str:
    if not row["adjusted_estimable"]:
        return (
            "Dữ liệu ban đầu có thể cho thấy một xu hướng giữa độ đầy đủ của fragment và tỷ lệ đạt. "
            "Tuy nhiên, dữ liệu trong các nhóm có thể so sánh trực tiếp chưa đủ để xác nhận xu hướng đó. "
            "Vì vậy, câu hỏi này cần thêm dữ liệu phù hợp trước khi đưa ra kết luận."
        )
    crude_positive = bool(
        row["crude_evidence"]
        and isinstance(row["crude_statistic"], (int, float))
        and float(row["crude_statistic"]) > 0
    )
    adjusted_positive = bool(
        row["adjusted_evidence"]
        and isinstance(row["adjusted_statistic"], (int, float))
        and float(row["adjusted_statistic"]) > 0
    )
    if adjusted_positive:
        opening = (
            "Ban đầu, nhóm mẫu có fragment đầy đủ hơn có xu hướng đạt nhiều hơn. "
            if crude_positive
            else "Khi xem toàn bộ dữ liệu, xu hướng đạt cao hơn chưa rõ ràng. "
        )
        return (
            opening
            + "Khi so sánh các mẫu cùng khối lớp và cùng nhóm chấm, xu hướng đạt cao hơn được ghi nhận. "
            "Kết quả này vẫn cần được hiểu là sự đi kèm, không phải quan hệ nguyên nhân."
        )
    opening = (
        "Ban đầu, nhóm mẫu có fragment đầy đủ hơn có xu hướng đạt nhiều hơn. "
        if crude_positive
        else "Ban đầu, chưa thấy nhóm mẫu có fragment đầy đủ hơn đạt nhiều hơn một cách rõ ràng. "
    )
    return (
        opening
        + "Tuy nhiên, khi so sánh các mẫu cùng khối lớp và cùng nhóm chấm, xu hướng này không còn rõ ràng. "
        "Vì vậy, chưa thể khẳng định mức độ đầy đủ của fragment đi kèm tỷ lệ đạt cao hơn."
    )


def root_hnmu_focus_content(
    summary_rows: Sequence[dict[str, object]],
) -> dict[str, str]:
    row = _root_focus_row(summary_rows)
    return {
        "title": ROOT_FOCUS_TITLE,
        "question": ROOT_FOCUS_QUESTION,
        "conclusion": _root_focus_conclusion(row),
        "explanation": _root_focus_explanation(row),
        "note_1": (
            "Phân tích này chỉ xem xét mối liên hệ, không chứng minh fragment đầy đủ hơn "
            "là nguyên nhân làm tăng tỷ lệ đạt."
        ),
        "note_2": "Chi tiết số liệu và phương pháp được lưu trong phụ lục kỹ thuật.",
    }


def write_root_hnmu_focus_workbook(
    path: Path,
    sheet_name: str,
    summary_rows: Sequence[dict[str, object]],
) -> None:
    content = root_hnmu_focus_content(summary_rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    blocks = (
        (1, content["title"], "title"),
        (3, "CÂU HỎI PHÂN TÍCH", "section"),
        (4, content["question"], "body"),
        (6, "KẾT LUẬN", "section"),
        (7, content["conclusion"], "conclusion"),
        (9, "GIẢI THÍCH NGẮN", "section"),
        (10, content["explanation"], "body"),
        (12, "LƯU Ý", "section"),
        (13, content["note_1"], "note"),
        (14, content["note_2"], "note"),
    )
    navy = "1F4E78"
    green = "548235"
    pale_green = "E2F0D9"
    pale_blue = "DDEBF7"
    for row_number, text, kind in blocks:
        worksheet.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=3,
        )
        cell = worksheet.cell(row_number, 1, text)
        cell.alignment = Alignment(
            horizontal="center" if kind in {"title", "section"} else "left",
            vertical="center" if kind in {"title", "section"} else "top",
            wrap_text=True,
        )
        if kind == "title":
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True, size=16)
        elif kind == "section":
            cell.fill = PatternFill("solid", fgColor=green)
            cell.font = Font(color="FFFFFF", bold=True, size=12)
        elif kind == "conclusion":
            cell.fill = PatternFill("solid", fgColor=pale_green)
            cell.font = Font(color="1F1F1F", bold=True, size=12)
        elif kind == "note":
            cell.fill = PatternFill("solid", fgColor=pale_blue)
            cell.font = Font(color="1F1F1F", italic=True, size=11)
        else:
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
            cell.font = Font(color="1F1F1F", size=11)

    for column in ("A", "B", "C"):
        worksheet.column_dimensions[column].width = 34
    heights = {
        1: 48, 3: 26, 4: 44, 6: 26, 7: 62,
        9: 26, 10: 74, 12: 26, 13: 48, 14: 36,
    }
    for row_number, height in heights.items():
        worksheet.row_dimensions[row_number].height = height
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 100
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = "A1:C14"
    worksheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    workbook.save(path)
    workbook.close()


def read_root_hnmu_focus_workbook(
    path: Path,
    sheet_name: str,
) -> dict[str, str]:
    workbook = load_workbook(path, data_only=True)
    if workbook.sheetnames != [sheet_name]:
        raise ValueError(f"{path.name} must have exactly one sheet named {sheet_name}")
    worksheet = workbook[sheet_name]
    content = {
        "title": str(worksheet["A1"].value or ""),
        "question": str(worksheet["A4"].value or ""),
        "conclusion": str(worksheet["A7"].value or ""),
        "explanation": str(worksheet["A10"].value or ""),
        "note_1": str(worksheet["A13"].value or ""),
        "note_2": str(worksheet["A14"].value or ""),
    }
    if worksheet.max_column > 3 or worksheet.max_row > 14:
        raise ValueError(f"{path.name} exceeds the compact root layout")
    if worksheet.tables or worksheet.auto_filter.ref or worksheet.freeze_panes:
        raise ValueError(f"{path.name} must use text blocks rather than a result table")
    workbook.close()
    return content

def _append_result_block(
    worksheet,
    *,
    title_row: int,
    header_row: int,
    start_row: int,
    title: str,
    rows: Sequence[Sequence[object]],
    table_name: str,
) -> None:
    worksheet.merge_cells(
        start_row=title_row,
        start_column=1,
        end_row=title_row,
        end_column=len(SUMMARY_HEADERS),
    )
    worksheet.cell(title_row, 1, title)
    for column, header in enumerate(SUMMARY_HEADERS, start=1):
        worksheet.cell(header_row, column, header)
    for offset, row in enumerate(rows):
        row_number = start_row + offset
        for column, value in enumerate(row, start=1):
            worksheet.cell(row_number, column, value)
    table = Table(
        displayName=table_name,
        ref=(
            f"A{header_row}:"
            f"{chr(64 + len(SUMMARY_HEADERS))}{start_row + len(rows) - 1}"
        ),
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def write_hnmu_summary_workbook(
    path: Path,
    sheet_name: str,
    summary_rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None = None,
) -> None:
    if len(summary_rows) != 8:
        raise ValueError("HNMU summary workbook requires exactly 8 result rows")
    if pooled or grade is None:
        raise ValueError("This writer is reserved for one fixed grade")
    display_rows = grade_summary_display_rows(summary_rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.merge_cells("A1:E1")
    worksheet["A1"] = SUMMARY_TITLE
    for row_number, (label, text) in enumerate(
        summary_intro_blocks(summary_rows, pooled=False, grade=grade),
        start=2,
    ):
        worksheet.cell(row_number, 1, label)
        worksheet.merge_cells(
            start_row=row_number,
            start_column=2,
            end_row=row_number,
            end_column=5,
        )
        worksheet.cell(row_number, 2, text)

    _append_result_block(
        worksheet,
        title_row=6,
        header_row=SUMMARY_BLOCK_A_HEADER_ROW,
        start_row=SUMMARY_BLOCK_A_START_ROW,
        title="A. TRẠNG THÁI ĐẠT CHÍNH THỨC",
        rows=display_rows[:4],
        table_name="TomTatTrangThaiDat",
    )
    _append_result_block(
        worksheet,
        title_row=13,
        header_row=SUMMARY_BLOCK_B_HEADER_ROW,
        start_row=SUMMARY_BLOCK_B_START_ROW,
        title="B. TỶ LỆ TIÊU CHÍ ĐẠT",
        rows=display_rows[4:],
        table_name="TomTatTyLeTieuChi",
    )

    navy = "1F4E78"
    green = "548235"
    pale_green = "E2F0D9"
    pale_blue = "DDEBF7"
    worksheet["A1"].fill = PatternFill("solid", fgColor=navy)
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    for row_number in range(2, 5):
        worksheet.cell(row_number, 1).font = Font(bold=True, color="1F1F1F")
        worksheet.cell(row_number, 1).fill = PatternFill("solid", fgColor=pale_blue)
        worksheet.cell(row_number, 1).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        worksheet.cell(row_number, 2).fill = PatternFill("solid", fgColor="FFFFFF")
        worksheet.cell(row_number, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
    for title_row in (6, 13):
        worksheet.cell(title_row, 1).fill = PatternFill("solid", fgColor=green)
        worksheet.cell(title_row, 1).font = Font(color="FFFFFF", bold=True)
        worksheet.cell(title_row, 1).alignment = Alignment(vertical="center")
    for header_row in (SUMMARY_BLOCK_A_HEADER_ROW, SUMMARY_BLOCK_B_HEADER_ROW):
        for cell in worksheet[header_row]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_number in (*range(8, 12), *range(15, 19)):
        for cell in worksheet[row_number]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.cell(row_number, 5).fill = PatternFill("solid", fgColor=pale_green)

    widths = {"A": 27, "B": 25, "C": 23, "D": 33, "E": 48}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.row_dimensions[1].height = 42
    worksheet.row_dimensions[2].height = 45
    worksheet.row_dimensions[3].height = 48
    worksheet.row_dimensions[4].height = 62
    for row_number in (6, 13):
        worksheet.row_dimensions[row_number].height = 24
    for row_number in (7, 14):
        worksheet.row_dimensions[row_number].height = 42
    for row_number in (*range(8, 12), *range(15, 19)):
        worksheet.row_dimensions[row_number].height = 58

    worksheet.freeze_panes = "A8"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 85
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = "A1:E18"
    worksheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    workbook.save(path)
    workbook.close()


def read_hnmu_summary_workbook(
    path: Path,
    sheet_name: str,
) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, data_only=True)
    if workbook.sheetnames != [sheet_name]:
        raise ValueError(f"{path.name} must have exactly one sheet named {sheet_name}")
    worksheet = workbook[sheet_name]
    if worksheet["A1"].value != SUMMARY_TITLE:
        raise ValueError(f"Unexpected HNMU summary title in {path.name}")
    for header_row in (SUMMARY_BLOCK_A_HEADER_ROW, SUMMARY_BLOCK_B_HEADER_ROW):
        headers = tuple(worksheet.cell(header_row, column).value for column in range(1, len(SUMMARY_HEADERS) + 1))
        if headers != SUMMARY_HEADERS:
            raise ValueError(f"Unexpected HNMU summary headers in {path.name}")
    rows = [
        tuple(worksheet.cell(row, column).value for column in range(1, len(SUMMARY_HEADERS) + 1))
        for row in (*range(8, 12), *range(15, 19))
    ]
    if worksheet.freeze_panes != "A8" or len(worksheet.tables) != 2:
        raise ValueError(f"{path.name} misses freeze pane or two result tables")
    workbook.close()
    return rows


def expected_summary_display_rows(
    rows: Iterable[dict[str, object]],
) -> list[tuple[object, ...]]:
    return grade_summary_display_rows(rows)
