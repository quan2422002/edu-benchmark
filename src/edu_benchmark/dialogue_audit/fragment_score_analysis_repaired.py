"""Repaired-checklist fragment analysis and teacher-readable workbook rendering."""

from __future__ import annotations

import math
from collections import Counter
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from edu_benchmark.dialogue_audit.fragment_score_analysis import (
    ANALYSIS_HEADERS,
    FRAGMENT_METRICS,
    FragmentAnalysisData,
    SampleAnalysisRecord,
    _summary_row,
    build_analysis_rows,
)
from edu_benchmark.dialogue_audit.teacher_bundle import (
    EXPECTED_CRITERIA_PER_SAMPLE,
    EXPECTED_GRADE_COUNTS,
    BundleData,
)
from edu_benchmark.dialogue_audit.teacher_bundle_v2 import (
    format_fraction_percentage_half_up,
)

ROOT_ANALYSIS_NAME = "05_phan_tich_fragment_va_ket_qua_cham_giua_cac_khoi.xlsx"
GRADE_ANALYSIS_NAME = "07_phan_tich_fragment_va_ket_qua_cham.xlsx"
ROOT_SHEET_NAME = "fragment_giua_cac_khoi"
GRADE_SHEET_NAME = "fragment_va_ket_qua_cham"
DATA_HEADER_ROW = 17

METRIC_LABELS = {
    "fragment_row_count": "số tiêu chí có ghi fragment",
    "fragment_reference_count": "tổng số lượt tham chiếu fragment",
    "unique_fragment_count": "số fragment phân biệt",
    "fragment_criterion_coverage": "tỷ lệ tiêu chí có ghi fragment",
}
OUTCOME_LABELS = {
    "official_pass": "trạng thái pass chính thức",
    "checklist_pass_rate": "tỷ lệ tiêu chí đạt",
}
HEADER_LABELS = {
    "grade": "Phạm vi (grade_or_scope)",
    "analysis_family": "Nhóm phân tích (analysis_family)",
    "outcome": "Kết quả được so sánh (outcome)",
    "fragment_metric": "Cách đo fragment (fragment_metric)",
    "adjustment": "Mức điều chỉnh (adjustment)",
    "grouping_or_bucket": "Nhóm giá trị (grouping_or_bucket)",
    "sample_count": "Số mẫu (sample_count)",
    "pass_count": "Số pass (pass_count)",
    "non_pass_count": "Số non-pass (non_pass_count)",
    "pass_rate": "Tỷ lệ pass (pass_rate)",
    "pass_rate_difference_from_lowest_bucket": "Chênh lệch với nhóm thấp nhất",
    "statistic_name": "Chỉ số thống kê (statistic_name)",
    "statistic_value": "Giá trị chỉ số (statistic_value)",
    "p_value": "p-value",
    "effect_size_name": "Tên effect size",
    "effect_size_value": "Giá trị effect size",
    "estimable": "Có thể ước lượng (estimable)",
    "interpretation": "Diễn giải đầy đủ",
    "warning": "Cảnh báo/giới hạn",
    "metric_min": "Nhỏ nhất",
    "metric_q1": "Tứ phân vị 25%",
    "metric_median": "Trung vị",
    "metric_q3": "Tứ phân vị 75%",
    "metric_max": "Lớn nhất",
    "metric_mean": "Trung bình",
    "strata_with_variation": "Số strata đủ biến thiên",
    "strata_total": "Tổng số strata",
    "method": "Phương pháp thực tế",
}


def build_repaired_fragment_data(data: BundleData) -> FragmentAnalysisData:
    """Aggregate the canonical repaired checklist to one record per sample."""

    records: list[SampleAnalysisRecord] = []
    pairs: set[tuple[str, str]] = set()
    criterion_distribution: Counter[int] = Counter()
    allowed_results = {"pass", "uncertain", "fail", "not_applicable"}
    for sample_id in sorted(data.normalized_by_id):
        normalized = data.normalized_by_id[sample_id]
        quality = data.quality_by_id[sample_id]
        checklist = data.checklist_by_id.get(sample_id, [])
        if not checklist:
            raise ValueError(f"Checklist repaired misses sample_id: {sample_id}")
        results = [row.get("result", "").strip().casefold() for row in checklist]
        if set(results) - allowed_results:
            raise ValueError(f"Unexpected criterion result for {sample_id}: {set(results) - allowed_results}")
        for row in checklist:
            pair = (sample_id, row.get("criterion_id", "").strip())
            if not pair[1] or pair in pairs:
                raise ValueError(f"Missing or duplicate repaired checklist key: {pair}")
            pairs.add(pair)
        if len(checklist) != EXPECTED_CRITERIA_PER_SAMPLE:
            raise ValueError(f"Repaired checklist has {len(checklist)} criteria for {sample_id}")
        criterion_distribution[len(checklist)] += 1

        checked_by_values = sorted({row.get("checked_by", "").strip() for row in checklist if row.get("checked_by", "").strip()})
        if not checked_by_values:
            raise ValueError(f"Repaired checklist misses checked_by for {sample_id}")
        checked_by = "; ".join(checked_by_values)
        row_shards = {
            row.get("shard_id", "").strip()
            for row in checklist
            if row.get("shard_id", "").strip()
        }
        quality_shard = quality.get("source_shard", "").strip()
        shard_values = sorted(row_shards | ({quality_shard} if quality_shard else set()))
        if not shard_values:
            raise ValueError(f"Cannot resolve unified_shard for {sample_id}")
        unified_shard = "; ".join(shard_values)

        status = quality["quality_decision"].strip().casefold()
        official_pass = int(status == "pass")
        applicable = sum(result in {"pass", "uncertain", "fail"} for result in results)
        if not applicable:
            raise ValueError(f"No applicable criterion for {sample_id}")
        fragments = [
            fragment.strip()
            for row in checklist
            for fragment in row.get("evidence_fragment_id", "").split(";")
            if fragment.strip()
        ]
        fragment_rows = sum(bool(row.get("evidence_fragment_id", "").strip()) for row in checklist)
        pass_count = sum(result == "pass" for result in results)
        records.append(
            SampleAnalysisRecord(
                sample_id=sample_id,
                grade=normalized["grade"],
                checked_by=checked_by,
                unified_shard=unified_shard,
                auditor_group=f"{checked_by} | {unified_shard}",
                official_overall_status=status,
                official_pass=official_pass,
                official_non_pass=1 - official_pass,
                observed_criterion_count=len(checklist),
                applicable_criterion_count=applicable,
                criterion_pass_count=pass_count,
                checklist_pass_rate=pass_count / applicable,
                fragment_row_count=fragment_rows,
                fragment_reference_count=len(fragments),
                unique_fragment_count=len(set(fragments)),
                has_any_fragment=int(bool(fragments)),
                fragment_criterion_coverage=fragment_rows / applicable,
            )
        )

    grade_counts = Counter(record.grade for record in records)
    if dict(grade_counts) != EXPECTED_GRADE_COUNTS:
        raise ValueError(f"Wrong repaired fragment grade partition: {dict(grade_counts)}")
    if len(records) != 1050 or len(pairs) != 1050 * EXPECTED_CRITERIA_PER_SAMPLE:
        raise ValueError("Repaired fragment input does not contain 1,050 × 18 unique keys")
    if criterion_distribution != Counter({18: 1050}):
        raise ValueError(f"Unexpected repaired criterion distribution: {dict(criterion_distribution)}")
    return FragmentAnalysisData(
        records=tuple(sorted(records, key=lambda row: (int(row.grade), row.sample_id))),
        source_hashes=dict(data.source_hashes),
        criterion_pair_count=len(pairs),
        criterion_count_distribution_67={18: 462},
        join_success_count=len(records),
        join_failure_count=0,
    )


def _scope_label(value: object) -> str:
    return "toàn bộ lớp 6–9" if str(value) == "all" else f"lớp {value}"


def _strength(value: float) -> str:
    absolute = abs(value)
    if absolute < 0.1:
        return "rất yếu"
    if absolute < 0.3:
        return "yếu"
    if absolute < 0.5:
        return "vừa"
    return "mạnh"


def _direction(value: float) -> str:
    if abs(value) < 1e-12:
        return "gần như bằng 0"
    return "dương" if value > 0 else "âm"


def _complete_interpretation(row: dict[str, object]) -> str:
    scope = _scope_label(row["grade"])
    metric = METRIC_LABELS.get(str(row["fragment_metric"]), str(row["fragment_metric"]))
    outcome = OUTCOME_LABELS.get(str(row["outcome"]), str(row["outcome"]))
    adjustment = str(row["adjustment"])
    grouping = str(row["grouping_or_bucket"])
    if not bool(row.get("estimable")):
        return (
            f"Trong phạm vi {scope}, không thể ước lượng mối liên hệ giữa {metric} và {outcome} "
            f"cho {grouping} vì không có đủ nhóm đối chứng hoặc biến thiên; không có p-value được tạo."
        )
    if row["analysis_family"] == "cross_grade_consistency":
        return (
            f"Khi so sánh bốn lớp và kết quả gộp, chiều hoặc độ mạnh của mối liên hệ giữa {metric} và {outcome} "
            f"được đánh giá từ kết quả thô và kết quả đã điều chỉnh; cần đọc cùng cảnh báo để tránh diễn giải nhân quả."
        )
    if grouping == "bucket_association_test":
        value = row.get("effect_size_value")
        effect = f"Cramér’s V = {float(value):.3f}" if isinstance(value, (int, float)) else "effect size không ước lượng được"
        p_value = row.get("p_value")
        significance = (
            "có khác biệt thống kê giữa các nhóm"
            if isinstance(p_value, (int, float)) and p_value < 0.05
            else "chưa thấy khác biệt thống kê rõ giữa các nhóm"
        )
        return (
            f"Trong phạm vi {scope}, bảng nhóm của {metric} so với {outcome} {significance} ({effect}); "
            "đây là mô tả mối liên hệ, không phải quan hệ nhân quả."
        )
    if row.get("statistic_name") == "bucket_pass_rate":
        rate = row.get("pass_rate")
        rate_text = (
            format_fraction_percentage_half_up(rate)
            if isinstance(rate, (int, float))
            else "không xác định"
        )
        return (
            f"Trong phạm vi {scope}, nhóm “{grouping}” có tỷ lệ pass quan sát là {rate_text}; "
            f"dòng này mô tả phân phối của {metric}, không kiểm định quan hệ nhân quả với {outcome}."
        )
    value = row.get("statistic_value")
    p_value = row.get("p_value")
    if not isinstance(value, (int, float)):
        return f"Trong phạm vi {scope}, kết quả giữa {metric} và {outcome} không có hệ số hợp lệ để diễn giải."
    adjusted = adjustment != "crude"
    mode = "sau điều chỉnh" if adjusted else "ở mức thô"
    significance = (
        "có ý nghĩa thống kê ở ngưỡng 0,05"
        if isinstance(p_value, (int, float)) and p_value < 0.05
        else "chưa có ý nghĩa thống kê ở ngưỡng 0,05"
    )
    control = ""
    if adjustment == "adjusted_for_auditor_group":
        control = " sau khi kiểm soát nhóm auditor/shard"
    elif adjustment == "adjusted_for_grade_and_auditor_group":
        control = " sau khi kiểm soát đồng thời khối lớp và nhóm auditor/shard"
    return (
        f"Trong phạm vi {scope}, {mode}{control}, mối liên hệ giữa {metric} và {outcome} có chiều "
        f"{_direction(float(value))}, mức {_strength(float(value))}, và {significance}; kết quả chỉ phản ánh liên hệ quan sát được."
    )


def prepare_analysis_rows(data: FragmentAnalysisData) -> tuple[dict[str, list[dict[str, object]]], list[dict[str, object]]]:
    grade_rows, root_rows = build_analysis_rows(data)
    for rows in [*grade_rows.values(), root_rows]:
        for row in rows:
            row["interpretation"] = _complete_interpretation(row)
    return grade_rows, root_rows


def _numeric_summary(rows: Sequence[dict[str, object]], family: str, metric: str, adjustment: str) -> dict[str, object] | None:
    return _summary_row(rows, family, metric, adjustment)


def main_conclusion(rows: Sequence[dict[str, object]], *, pooled: bool) -> str:
    adjustment = "adjusted_for_grade_and_auditor_group" if pooled else "adjusted_for_auditor_group"
    pairs: list[tuple[dict[str, object] | None, dict[str, object] | None]] = []
    for family in ("fragment_vs_official_pass", "fragment_vs_checklist_pass_rate"):
        for metric in FRAGMENT_METRICS:
            pairs.append(
                (
                    _numeric_summary(rows, family, metric, "crude"),
                    _numeric_summary(rows, family, metric, adjustment),
                )
            )
    estimable = [
        (crude, adjusted)
        for crude, adjusted in pairs
        if adjusted
        and bool(adjusted.get("estimable"))
        and isinstance(adjusted.get("statistic_value"), (int, float))
    ]
    grade = next((str(row["grade"]) for row in rows if str(row.get("grade")) != "all"), "")
    if not estimable:
        if pooled:
            return (
                "Không thể ước lượng các mối liên hệ sau điều chỉnh trong dữ liệu gộp vì không có strata "
                "grade × auditor_group nào đủ biến thiên đồng thời về outcome và metric fragment. "
                "Các kết quả thô không được xem là bằng chứng về mối liên hệ độc lập."
            )
        return (
            f"Không thể ước lượng mối liên hệ sau điều chỉnh trong lớp {grade} vì không có auditor_group nào "
            "đủ biến thiên đồng thời về outcome và metric fragment. Các kết quả thô không được xem là bằng chứng "
            "về mối liên hệ độc lập."
        )

    significant_count = sum(
        isinstance(adjusted.get("p_value"), (int, float)) and float(adjusted["p_value"]) < 0.05
        for _, adjusted in estimable
    )
    reversal_count = sum(
        bool(crude)
        and isinstance(crude.get("statistic_value"), (int, float))
        and float(crude["statistic_value"]) * float(adjusted["statistic_value"]) < 0
        for crude, adjusted in estimable
    )
    adjusted_signs = {
        1 if float(adjusted["statistic_value"]) > 0 else -1 if float(adjusted["statistic_value"]) < 0 else 0
        for _, adjusted in estimable
    }
    non_estimable_count = len(pairs) - len(estimable)
    scope = "Trong dữ liệu gộp" if pooled else f"Trong lớp {grade}"
    control = (
        "sau khi kiểm soát đồng thời grade và auditor_group"
        if pooled
        else "sau khi kiểm soát auditor_group trong lớp"
    )
    estimate_text = f"ước lượng được {len(estimable)}/8 mối liên hệ {control}"
    if non_estimable_count:
        estimate_text += (
            f"; {non_estimable_count}/8 mối liên hệ còn lại không thể ước lượng vì strata không đủ biến thiên"
        )
    if significant_count:
        evidence_text = (
            f"Trong các ước lượng được, {significant_count} mối liên hệ có p-value dưới 0,05, "
            "nhưng cần đọc cùng chiều, độ mạnh và số mẫu thực sự được dùng"
        )
    else:
        evidence_text = "Các ước lượng được chưa có ý nghĩa thống kê ở ngưỡng 0,05"
    consistency_parts = []
    if reversal_count:
        consistency_parts.append(f"{reversal_count} mối liên hệ đổi chiều so với kết quả thô")
    if len(adjusted_signs - {0}) > 1:
        consistency_parts.append("các hệ số sau điều chỉnh không cùng chiều")
    consistency_text = (
        "; ".join(consistency_parts).capitalize() + ". "
        if consistency_parts
        else ""
    )
    return (
        f"{scope}, {estimate_text}. {evidence_text}. {consistency_text}"
        "Chưa có bằng chứng về một mối liên hệ độc lập và ổn định; kết quả không chứng minh quan hệ nhân quả."
    )


def _guide_rows(conclusion: str, *, pooled: bool) -> list[tuple[str, str]]:
    adjusted_text = (
        "Trong file tổng hợp, adjusted kiểm soát đồng thời khối lớp và nhóm auditor/shard."
        if pooled
        else "Trong file lớp, adjusted kiểm soát nhóm auditor/shard trong chính lớp đó."
    )
    limit_text = (
        "Đây là phân tích quan sát; khác biệt giữa khối lớp và nhóm auditor/shard có thể chi phối kết quả."
        if pooled
        else "Đây là phân tích quan sát trong một lớp; khác biệt giữa các auditor_group có thể chi phối kết quả."
    )
    return [
        ("HƯỚNG DẪN ĐỌC BẢNG", ""),
        ("Câu hỏi", "Mức độ ghi tham chiếu fragment có đi cùng kết quả chấm hay không?"),
        ("Bốn cách đo fragment", "Số tiêu chí có fragment; tổng lượt tham chiếu; số fragment phân biệt; tỷ lệ tiêu chí có fragment."),
        ("official_pass", "Bằng 1 khi trạng thái tổng thể canonical là pass; các trạng thái còn lại được gộp thành non-pass."),
        ("checklist_pass_rate", "Tỷ lệ tiêu chí pass trên các tiêu chí áp dụng của từng sample_id."),
        ("crude", "Mối liên hệ quan sát trực tiếp, chưa kiểm soát khác biệt quy trình chấm."),
        ("adjusted", adjusted_text),
        ("statistic_value", "Cho biết chiều và độ mạnh của mối liên hệ; dấu dương/âm không có nghĩa là nguyên nhân."),
        ("p_value", "Giúp đánh giá mức tương thích với giả thuyết không có liên hệ; p-value nhỏ không chứng minh quan hệ nhân quả."),
        ("effect_size", "Cho biết độ lớn của mối liên hệ, cần đọc cùng p-value, số mẫu và cảnh báo."),
        ("estimable = Không", "Không có đủ nhóm đối chứng hoặc biến thiên; dòng đó không có p-value giả."),
        ("Giới hạn", limit_text),
        ("", ""),
        ("KẾT LUẬN CHÍNH", ""),
        ("Kết luận", conclusion),
        ("", ""),
    ]


def _display_value(field: str, value: object, row: dict[str, object]) -> object:
    if field == "estimable":
        return "Có" if bool(value) else "Không"
    if not bool(row.get("estimable")) and field in {"statistic_value", "p_value", "effect_size_value"}:
        return "Không thể ước lượng"
    if isinstance(value, float) and math.isnan(value):
        return "Không thể ước lượng"
    return value


def write_analysis_workbook(
    path: Path,
    sheet_name: str,
    rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    conclusion = main_conclusion(rows, pooled=pooled)
    for label, explanation in _guide_rows(conclusion, pooled=pooled):
        worksheet.append([label, explanation])
    worksheet.append([HEADER_LABELS[field] for field in ANALYSIS_HEADERS])
    for row in rows:
        worksheet.append([_display_value(field, row.get(field, ""), row) for field in ANALYSIS_HEADERS])

    navy = "1F4E78"
    for cell in worksheet[DATA_HEADER_ROW]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet["A1"].font = Font(bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor=navy)
    worksheet["A14"].font = Font(bold=True, color="FFFFFF")
    worksheet["A14"].fill = PatternFill("solid", fgColor="548235")
    for row_number in range(1, DATA_HEADER_ROW):
        worksheet.cell(row_number, 1).font = Font(bold=True)
        worksheet.cell(row_number, 1).alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
    for row_number in range(DATA_HEADER_ROW + 1, worksheet.max_row + 1):
        adjustment = str(worksheet.cell(row_number, 5).value or "")
        color = "DDEBF7" if adjustment == "crude" else "E2F0D9" if adjustment.startswith("adjusted") else "FFF2CC"
        for cell in worksheet[row_number]:
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(wrap_text=cell.column in {6, 18, 19, 28}, vertical="top")
        for column in (10, 11):
            if isinstance(worksheet.cell(row_number, column).value, (int, float)):
                worksheet.cell(row_number, column).number_format = "0.00%"
        for column in (13, 16, 20, 21, 22, 23, 24, 25):
            if isinstance(worksheet.cell(row_number, column).value, (int, float)):
                worksheet.cell(row_number, column).number_format = "0.0000"
        if isinstance(worksheet.cell(row_number, 14).value, (int, float)):
            worksheet.cell(row_number, 14).number_format = "0.000E+00"
    widths = {1: 22, 2: 34, 3: 28, 4: 31, 5: 39, 6: 49, 7: 13, 8: 12, 9: 15, 10: 13, 11: 18,
              12: 27, 13: 17, 14: 15, 15: 20, 16: 17, 17: 15, 18: 78, 19: 72, 20: 12, 21: 14,
              22: 13, 23: 14, 24: 12, 25: 13, 26: 20, 27: 15, 28: 72}
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.column_dimensions["B"].width = 105
    worksheet.freeze_panes = f"A{DATA_HEADER_ROW + 1}"
    worksheet.auto_filter.ref = f"A{DATA_HEADER_ROW}:{get_column_letter(len(ANALYSIS_HEADERS))}{worksheet.max_row}"
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[15].height = 54
    workbook.save(path)
    workbook.close()


def read_analysis_workbook(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, data_only=True)
    if workbook.sheetnames != [sheet_name]:
        raise ValueError(f"{path.name} must have exactly one sheet named {sheet_name}")
    worksheet = workbook[sheet_name]
    headers = tuple(cell.value for cell in worksheet[DATA_HEADER_ROW])
    expected = tuple(HEADER_LABELS[field] for field in ANALYSIS_HEADERS)
    if headers != expected:
        raise ValueError(f"Unexpected fragment-analysis headers in {path.name}")
    rows = list(worksheet.iter_rows(min_row=DATA_HEADER_ROW + 1, values_only=True))
    if worksheet.freeze_panes != f"A{DATA_HEADER_ROW + 1}" or not worksheet.auto_filter.ref:
        raise ValueError(f"{path.name} misses freeze pane or filter")
    workbook.close()
    return rows


def expected_display_rows(rows: Iterable[dict[str, object]]) -> list[tuple[object, ...]]:
    return [
        tuple(_display_value(field, row.get(field, ""), row) for field in ANALYSIS_HEADERS)
        for row in rows
    ]
