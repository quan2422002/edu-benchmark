"""Build and validate teacher-facing workbooks from canonical HNMU audit outputs.

This module packages existing Plan 04 outputs. It never reruns the audit and
treats source_file values as opaque provenance strings. Only the explicit
fifteen-file allowlist may be opened as source data.
"""

from __future__ import annotations

import csv
import hashlib
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

EXPECTED_GRADE_COUNTS = {"6": 238, "7": 224, "8": 280, "9": 308}
EXPECTED_REVIEW_COUNTS = {"6": 132, "7": 92, "8": 71, "9": 90}
EXPECTED_CRITERIA_PER_SAMPLE = 18
ALLOWED_DECISIONS = {"pass", "need_human_review", "failed"}
SHEET_NAMES = [
    "00_Huong_dan",
    "01_Tong_quan",
    "02_Can_ra_soat",
    "03_Da_dat",
    "04_Do_phu",
    "PL_Chi_tiet_tieu_chi",
    "PL_Nguon_du_lieu",
]
SAMPLE_HEADERS = [
    "Mã mẫu (sample_id)",
    "Lớp",
    "Tệp nguồn (source_file)",
    "Dòng nguồn (source_row_number)",
    "STT",
    "Bài",
    "Vị trí",
    "Câu hỏi",
    "Mức nhận thức",
    "Đáp án SGV",
    "Hội thoại",
    "Kết quả rà soát",
    "Mã kết quả (quality_decision)",
    "Độ tin cậy",
    "Lý do cần xem lại",
    "Mã tiêu chí liên quan",
    "Hành động đề nghị",
    "Mức ưu tiên",
    "Câu hỏi đề nghị HNMU xác nhận",
    "Cảnh báo cơ học",
    "Cần HNMU xem lại",
    "Cần kiểm học liệu",
    "Cần kiểm SGV",
    "Mã bằng chứng",
]
CHECKLIST_HEADERS = [
    "Mã mẫu (sample_id)",
    "Mã tiêu chí (criterion_id)",
    "Nhóm tiêu chí",
    "Tên tiêu chí",
    "Kết quả",
    "Mã kết quả (result)",
    "Độ tin cậy",
    "Mã bằng chứng",
    "Nguồn bằng chứng",
    "Lý do bằng chứng phù hợp",
    "Giải thích kết quả",
    "Hành động đề nghị",
    "Người kiểm",
    "Thời điểm kiểm",
]
DECISION_LABELS = {
    "pass": "Đạt theo checklist hiện tại",
    "need_human_review": "Cần giáo viên xem lại",
    "failed": "Chưa nên dùng ở lượt hiện tại",
}
CRITERION_LABELS = {
    "pass": "Đạt",
    "fail": "Không đạt",
    "uncertain": "Chưa chắc, cần xem lại",
    "not_applicable": "Không áp dụng",
}
PRIORITY_LABELS = {
    "high": "Cao",
    "cao": "Cao",
    "medium": "Trung bình",
    "trung bình": "Trung bình",
}


@dataclass(frozen=True)
class CanonicalSources:
    """The exact fifteen source files approved for Plan 08."""

    normalized_67: Path
    coverage_67: Path
    missing_67: Path
    duplicates_67: Path
    quality_67: Path
    review_67: Path
    checklist_67: Path
    normalized_89: Path
    coverage_89: Path
    missing_89: Path
    duplicates_89: Path
    quality_89: Path
    review_89: Path
    checklist_89: Path
    teacher_report: Path

    def all_paths(self) -> tuple[Path, ...]:
        return (
            self.normalized_67,
            self.coverage_67,
            self.missing_67,
            self.duplicates_67,
            self.quality_67,
            self.review_67,
            self.checklist_67,
            self.normalized_89,
            self.coverage_89,
            self.missing_89,
            self.duplicates_89,
            self.quality_89,
            self.review_89,
            self.checklist_89,
            self.teacher_report,
        )


@dataclass
class BatchData:
    normalized: list[dict[str, str]]
    coverage: list[dict[str, str]]
    missing: list[dict[str, str]]
    duplicates: list[dict[str, str]]
    quality: list[dict[str, str]]
    review: list[dict[str, str]]
    checklist: list[dict[str, str]]


@dataclass
class BundleData:
    normalized_by_id: dict[str, dict[str, str]]
    quality_by_id: dict[str, dict[str, str]]
    review_by_id: dict[str, dict[str, str]]
    checklist_by_id: dict[str, list[dict[str, str]]]
    alerts_by_id: dict[str, list[str]]
    coverage_by_grade: dict[str, list[dict[str, str]]]
    missing_rows: list[dict[str, str]]
    duplicate_rows: list[dict[str, str]]
    teacher_report_text: str
    source_hashes: dict[Path, str]
    source_record_counts: dict[Path, int]
    read_paths: set[Path]


@dataclass(frozen=True)
class BuildSummary:
    output_paths: tuple[Path, ...]
    grade_counts: dict[str, int]
    review_counts: dict[str, int]
    source_hashes: dict[str, str]
    read_paths: tuple[str, ...]


class AllowlistedSourceReader:
    """Reject every source-data path outside the canonical allowlist."""

    def __init__(self, allowed_paths: Iterable[Path]):
        self._allowed = {path.resolve() for path in allowed_paths}
        self.read_paths: set[Path] = set()

    def _check(self, path: Path) -> Path:
        resolved = path.resolve()
        if resolved not in self._allowed:
            raise ValueError(f"Source path is not in the canonical allowlist: {path}")
        if not resolved.is_file():
            raise FileNotFoundError(f"Canonical source is missing: {path}")
        self.read_paths.add(resolved)
        return resolved

    def read_csv(self, path: Path) -> list[dict[str, str]]:
        resolved = self._check(path)
        with resolved.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    def read_text(self, path: Path) -> str:
        return self._check(path).read_text(encoding="utf-8")

    def sha256(self, path: Path) -> str:
        resolved = self._check(path)
        digest = hashlib.sha256()
        with resolved.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()


def canonical_source_paths(experiment_dir: Path) -> CanonicalSources:
    """Return the exact approved Plan 08 input paths."""

    outputs = experiment_dir / "outputs"
    audit_67 = outputs / "hnmu_dialogue_audit"
    audit_89 = outputs / "hnmu_dialogue_audit_grade8_9"
    return CanonicalSources(
        normalized_67=audit_67 / "normalized_dialogue_rows.csv",
        coverage_67=audit_67 / "coverage_summary.csv",
        missing_67=audit_67 / "missing_field_report.csv",
        duplicates_67=audit_67 / "duplicate_candidates.csv",
        quality_67=audit_67 / "agent_shard_audit/merged/quality_check_suggestions.csv",
        review_67=audit_67 / "agent_shard_audit/merged/hnmu_review_queue_suggestions.csv",
        checklist_67=audit_67 / "agent_shard_audit/merged/raw_dialogue_checklist_results.repaired.csv",
        normalized_89=audit_89 / "normalized_dialogue_rows.csv",
        coverage_89=audit_89 / "coverage_summary.csv",
        missing_89=audit_89 / "missing_field_report.csv",
        duplicates_89=audit_89 / "duplicate_candidates.csv",
        quality_89=audit_89 / "agent_shard_audit/merged/quality_check_suggestions.csv",
        review_89=audit_89 / "agent_shard_audit/merged/hnmu_review_queue_suggestions.csv",
        checklist_89=audit_89 / "agent_shard_audit/merged/raw_dialogue_checklist_results.regex_repaired.csv",
        teacher_report=audit_67
        / "reports/bao-cao-gui-hnmu-ket-qua-ra-soat-du-lieu-hoi-thoai-lop-6-9-20260719.md",
    )


def _display_path(path: Path) -> str:
    resolved = path.resolve()
    try:
        return resolved.relative_to(Path.cwd().resolve()).as_posix()
    except ValueError:
        return resolved.as_posix()


def _require_columns(rows: Sequence[dict[str, str]], required: set[str], label: str) -> None:
    if not rows:
        raise ValueError(f"Canonical source has no data rows: {label}")
    missing = required - set(rows[0])
    if missing:
        raise ValueError(f"Canonical source {label} is missing columns: {sorted(missing)}")


def _index_unique(rows: Sequence[dict[str, str]], label: str) -> dict[str, dict[str, str]]:
    indexed: dict[str, dict[str, str]] = {}
    for row in rows:
        sample_id = row.get("sample_id", "").strip()
        if not sample_id:
            raise ValueError(f"{label} contains an empty sample_id")
        if sample_id in indexed:
            raise ValueError(f"{label} contains duplicate sample_id: {sample_id}")
        indexed[sample_id] = row
    return indexed


def _load_batch(
    reader: AllowlistedSourceReader,
    paths: Sequence[Path],
) -> BatchData:
    normalized, coverage, missing, duplicates, quality, review, checklist = paths
    batch = BatchData(
        normalized=reader.read_csv(normalized),
        coverage=reader.read_csv(coverage),
        missing=reader.read_csv(missing),
        duplicates=reader.read_csv(duplicates),
        quality=reader.read_csv(quality),
        review=reader.read_csv(review),
        checklist=reader.read_csv(checklist),
    )
    _require_columns(
        batch.normalized,
        {"sample_id", "source_file", "source_row_number", "grade", "question", "answer_sgv", "dialogue"},
        _display_path(normalized),
    )
    _require_columns(
        batch.coverage,
        {"dimension", "grade", "topic_id", "topic_label", "lesson_id", "lesson_label", "count"},
        _display_path(coverage),
    )
    _require_columns(
        batch.quality,
        {"sample_id", "source_file", "source_row_number", "grade", "quality_decision"},
        _display_path(quality),
    )
    _require_columns(
        batch.review,
        {"sample_id", "review_reason", "priority", "suggested_question_to_hnmu"},
        _display_path(review),
    )
    _require_columns(
        batch.checklist,
        {"sample_id", "criterion_id", "criterion_group", "criterion_name", "result", "confidence_score"},
        _display_path(checklist),
    )
    if batch.missing:
        _require_columns(batch.missing, {"sample_id", "message"}, _display_path(missing))
    if batch.duplicates:
        _require_columns(
            batch.duplicates,
            {"sample_id_a", "sample_id_b", "duplicate_type", "note"},
            _display_path(duplicates),
        )
    return batch


def _validate_batch(
    batch: BatchData,
    batch_label: str,
    *,
    normalized_by_id: dict[str, dict[str, str]],
    quality_by_id: dict[str, dict[str, str]],
    review_by_id: dict[str, dict[str, str]],
    checklist_by_id: dict[str, list[dict[str, str]]],
    alerts_by_id: dict[str, list[str]],
    coverage_by_grade: dict[str, list[dict[str, str]]],
) -> None:
    normalized = _index_unique(batch.normalized, f"{batch_label} normalized")
    quality = _index_unique(batch.quality, f"{batch_label} quality")
    if set(normalized) != set(quality):
        raise ValueError(f"{batch_label} normalized and quality sample sets differ")
    overlap = set(normalized_by_id) & set(normalized)
    if overlap:
        raise ValueError(f"Sample IDs occur in multiple batches: {sorted(overlap)[:5]}")

    for sample_id, normalized_row in normalized.items():
        quality_row = quality[sample_id]
        grade = normalized_row.get("grade", "").strip()
        if grade not in EXPECTED_GRADE_COUNTS:
            raise ValueError(f"Unknown grade for {sample_id}: {grade!r}")
        for field in ("grade", "source_file", "source_row_number"):
            if normalized_row.get(field, "") != quality_row.get(field, ""):
                raise ValueError(f"{sample_id} has conflicting {field}")
        if not normalized_row["source_file"].endswith(f"Lớp {grade}.xlsx"):
            raise ValueError(f"{sample_id} source_file does not match grade {grade}")
        if not sample_id.startswith(f"HNMU-G{grade}-"):
            raise ValueError(f"{sample_id} ID prefix does not match grade {grade}")
        decision = quality_row.get("quality_decision", "").strip()
        if decision not in ALLOWED_DECISIONS:
            raise ValueError(f"{sample_id} has unknown quality_decision: {decision!r}")

    review = _index_unique(batch.review, f"{batch_label} review queue")
    expected_review_ids = {
        sample_id
        for sample_id, row in quality.items()
        if row.get("quality_decision", "").strip() != "pass"
    }
    if set(review) != expected_review_ids:
        raise ValueError(f"{batch_label} review queue is not the exact non-pass set")
    for sample_id, row in review.items():
        if row.get("grade") and row["grade"] != normalized[sample_id]["grade"]:
            raise ValueError(f"Review queue grade mismatch for {sample_id}")

    seen_pairs: set[tuple[str, str]] = set()
    criteria_by_sample: dict[str, set[str]] = defaultdict(set)
    for row in batch.checklist:
        sample_id = row.get("sample_id", "").strip()
        criterion_id = row.get("criterion_id", "").strip()
        pair = (sample_id, criterion_id)
        if sample_id not in normalized:
            raise ValueError(f"Checklist contains unknown sample_id: {sample_id}")
        if not criterion_id or pair in seen_pairs:
            raise ValueError(f"Checklist has an empty or duplicate key: {pair}")
        seen_pairs.add(pair)
        criteria_by_sample[sample_id].add(criterion_id)
        checklist_by_id[sample_id].append(row)
    if set(criteria_by_sample) != set(normalized):
        raise ValueError(f"{batch_label} checklist sample set differs from normalized")
    bad = {
        sample_id: len(criteria)
        for sample_id, criteria in criteria_by_sample.items()
        if len(criteria) != EXPECTED_CRITERIA_PER_SAMPLE
    }
    if bad:
        raise ValueError(f"Checklist does not have 18 criteria/sample: {list(bad.items())[:5]}")

    for row in batch.missing:
        sample_id = row.get("sample_id", "").strip()
        if sample_id not in normalized:
            raise ValueError(f"Missing-field report contains unknown sample_id: {sample_id}")
        alerts_by_id[sample_id].append(row.get("message", "").strip())

    for row in batch.duplicates:
        sample_a = row.get("sample_id_a", "").strip()
        sample_b = row.get("sample_id_b", "").strip()
        if sample_a not in normalized or sample_b not in normalized:
            raise ValueError(f"Duplicate report contains unknown pair: {sample_a}, {sample_b}")
        if normalized[sample_a]["grade"] != normalized[sample_b]["grade"]:
            raise ValueError(f"Duplicate pair crosses grades: {sample_a}, {sample_b}")
        note = row.get("note", "").strip() or "Có ứng viên trùng cần xem lại."
        alerts_by_id[sample_a].append(f"Trùng với {sample_b}: {note}")
        alerts_by_id[sample_b].append(f"Trùng với {sample_a}: {note}")

    for row in batch.coverage:
        grade = row.get("grade", "")
        if row.get("dimension") == "lesson_by_grade" and grade in EXPECTED_GRADE_COUNTS:
            coverage_by_grade[grade].append(row)

    normalized_by_id.update(normalized)
    quality_by_id.update(quality)
    review_by_id.update(review)


def load_canonical_bundle_data(experiment_dir: Path) -> BundleData:
    """Read and validate exactly the fifteen canonical Plan 08 inputs."""

    sources = canonical_source_paths(experiment_dir)
    missing_paths = [path for path in sources.all_paths() if not path.is_file()]
    if missing_paths:
        formatted = "\n".join(f"- {_display_path(path)}" for path in missing_paths)
        raise FileNotFoundError(f"Missing canonical bundle source files:\n{formatted}")

    reader = AllowlistedSourceReader(sources.all_paths())
    paths = sources.all_paths()
    batch_67 = _load_batch(reader, paths[:7])
    batch_89 = _load_batch(reader, paths[7:14])
    report_text = reader.read_text(sources.teacher_report)
    if "Báo cáo rà soát bước đầu dữ liệu hội thoại" not in report_text:
        raise ValueError("Canonical HNMU report does not contain the expected title")

    normalized_by_id: dict[str, dict[str, str]] = {}
    quality_by_id: dict[str, dict[str, str]] = {}
    review_by_id: dict[str, dict[str, str]] = {}
    checklist_by_id: dict[str, list[dict[str, str]]] = defaultdict(list)
    alerts_by_id: dict[str, list[str]] = defaultdict(list)
    coverage_by_grade: dict[str, list[dict[str, str]]] = defaultdict(list)
    for label, batch in (("grade 6-7", batch_67), ("grade 8-9", batch_89)):
        _validate_batch(
            batch,
            label,
            normalized_by_id=normalized_by_id,
            quality_by_id=quality_by_id,
            review_by_id=review_by_id,
            checklist_by_id=checklist_by_id,
            alerts_by_id=alerts_by_id,
            coverage_by_grade=coverage_by_grade,
        )

    grade_counts = Counter(row["grade"] for row in normalized_by_id.values())
    if dict(sorted(grade_counts.items())) != EXPECTED_GRADE_COUNTS:
        raise ValueError(f"Unexpected grade counts: {dict(grade_counts)}")
    review_counts = Counter(normalized_by_id[sample_id]["grade"] for sample_id in review_by_id)
    if dict(sorted(review_counts.items())) != EXPECTED_REVIEW_COUNTS:
        raise ValueError(f"Unexpected review counts: {dict(review_counts)}")

    source_record_counts = {
        sources.normalized_67: len(batch_67.normalized),
        sources.coverage_67: len(batch_67.coverage),
        sources.missing_67: len(batch_67.missing),
        sources.duplicates_67: len(batch_67.duplicates),
        sources.quality_67: len(batch_67.quality),
        sources.review_67: len(batch_67.review),
        sources.checklist_67: len(batch_67.checklist),
        sources.normalized_89: len(batch_89.normalized),
        sources.coverage_89: len(batch_89.coverage),
        sources.missing_89: len(batch_89.missing),
        sources.duplicates_89: len(batch_89.duplicates),
        sources.quality_89: len(batch_89.quality),
        sources.review_89: len(batch_89.review),
        sources.checklist_89: len(batch_89.checklist),
        sources.teacher_report: len(report_text.splitlines()),
    }
    source_hashes = {path: reader.sha256(path) for path in sources.all_paths()}
    expected_read_paths = {path.resolve() for path in sources.all_paths()}
    if reader.read_paths != expected_read_paths:
        raise ValueError("Builder source read set differs from the fifteen-file allowlist")

    return BundleData(
        normalized_by_id=normalized_by_id,
        quality_by_id=quality_by_id,
        review_by_id=review_by_id,
        checklist_by_id=dict(checklist_by_id),
        alerts_by_id=dict(alerts_by_id),
        coverage_by_grade=dict(coverage_by_grade),
        missing_rows=[*batch_67.missing, *batch_89.missing],
        duplicate_rows=[*batch_67.duplicates, *batch_89.duplicates],
        teacher_report_text=report_text,
        source_hashes=source_hashes,
        source_record_counts=source_record_counts,
        read_paths=reader.read_paths,
    )


def _clean(value: object) -> object:
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub(
        lambda match: f"\\x{ord(match.group(0)):02X}",
        value,
    )


def _append_rows(ws: Worksheet, rows: Iterable[Sequence[object]]) -> None:
    for row in rows:
        values = [_clean(value) for value in row]
        ws.append(values)
        for cell, value in zip(ws[ws.max_row], values):
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"


def _style_table(ws: Worksheet) -> None:
    navy = "1F4E78"
    thin = Side(style="thin", color="D9E1F2")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for row_number in range(2, ws.max_row + 1):
        for cell in ws[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row_number % 2 == 0:
            for cell in ws[row_number]:
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.row_dimensions[1].height = 32


def _set_widths(ws: Worksheet, widths: Mapping[int, float]) -> None:
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def _yes_no(value: str) -> str:
    return "Có" if value.strip().lower() == "true" else "Không"


def _priority(value: str) -> str:
    return PRIORITY_LABELS.get(value.strip().lower(), value.strip())


def _sample_values(data: BundleData, sample_id: str) -> list[str]:
    normalized = data.normalized_by_id[sample_id]
    quality = data.quality_by_id[sample_id]
    review = data.review_by_id.get(sample_id, {})
    decision = quality["quality_decision"].strip()
    return [
        sample_id,
        normalized["grade"],
        normalized["source_file"],
        normalized["source_row_number"],
        normalized.get("stt", ""),
        normalized.get("lesson", ""),
        normalized.get("position", ""),
        normalized.get("question", ""),
        normalized.get("bloom_level", ""),
        normalized.get("answer_sgv", ""),
        normalized.get("dialogue", ""),
        DECISION_LABELS[decision],
        decision,
        quality.get("confidence_score", ""),
        review.get("review_reason", "") or quality.get("failure_reasons", ""),
        review.get("related_criterion_ids", "") or quality.get("blocking_criterion_ids", ""),
        quality.get("suggested_reviewer_action", ""),
        _priority(review.get("priority", "")),
        review.get("suggested_question_to_hnmu", ""),
        "\n".join(data.alerts_by_id.get(sample_id, [])),
        _yes_no(quality.get("needs_hnmu_review", "")),
        _yes_no(quality.get("needs_learning_resource_review", "")),
        _yes_no(quality.get("needs_sgv_verification", "")),
        quality.get("evidence_fragment_ids", ""),
    ]


def _sort_ids(data: BundleData, sample_ids: Iterable[str]) -> list[str]:
    def key(sample_id: str) -> tuple[int, str]:
        value = data.normalized_by_id[sample_id].get("source_row_number", "")
        return (int(value) if value.isdigit() else 10**9, sample_id)

    return sorted(sample_ids, key=key)


def _bloom_band(value: str) -> str:
    value = value.strip().casefold()
    for label in ("Nhận biết", "Thông hiểu", "Vận dụng"):
        if value.startswith(label.casefold()):
            return label
    return "Chưa nhận diện rõ"


def _add_instruction_sheet(wb: Workbook, grade: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[0])
    _append_rows(
        ws,
        [
            [f"Kết quả rà soát dữ liệu hội thoại Tin học lớp {grade}"],
            ["Mục đích", "Giúp thầy cô xem nhanh mẫu nào đã đạt theo checklist hiện tại và mẫu nào cần xác nhận thêm."],
            ["Đọc trước", "Mở trang 01_Tong_quan, sau đó mở 02_Can_ra_soat để xem các mẫu cần chú ý."],
            ["Khi rà soát", "Dùng Mã mẫu, Tệp nguồn và Dòng nguồn để đối chiếu với dữ liệu do HNMU cung cấp."],
            ["Lưu ý", "Cần giáo viên xem lại không đồng nghĩa với mẫu sai; nhãn này cho biết còn điểm cần HNMU/UET xác nhận."],
            ["Ví dụ", "Nếu lý do ghi chưa chắc về đáp án SGV, hãy đối chiếu câu hỏi và đáp án rồi ghi nhận ý kiến chuyên môn."],
            ["Không nên", "Không loại mẫu chỉ vì nhãn Cần giáo viên xem lại khi chưa đọc lý do và nội dung hội thoại."],
            ["Quyền quyết định", "Kết quả là gợi ý rà soát bước đầu. HNMU/UET quyết định cuối cùng về giữ, sửa hoặc loại mẫu."],
        ],
    )
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    for row_number in range(2, ws.max_row + 1):
        ws.cell(row_number, 1).font = Font(bold=True, color="1F4E78")
        ws.cell(row_number, 1).fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _set_widths(ws, {1: 22, 2: 110})
    ws.sheet_view.showGridLines = False


def _add_summary_sheet(wb: Workbook, data: BundleData, grade: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[1])
    sample_ids = [sid for sid, row in data.normalized_by_id.items() if row["grade"] == grade]
    decisions = Counter(data.quality_by_id[sid]["quality_decision"] for sid in sample_ids)
    alert_count = sum(len(data.alerts_by_id.get(sid, [])) for sid in sample_ids)
    duplicate_samples = sum(
        any(alert.startswith("Trùng với ") for alert in data.alerts_by_id.get(sid, []))
        for sid in sample_ids
    )
    _append_rows(
        ws,
        [
            ["Chỉ số", "Số lượng", "Diễn giải"],
            ["Tổng số mẫu", len(sample_ids), "Toàn bộ mẫu canonical của lớp trong lượt rà soát này."],
            ["Đạt theo checklist hiện tại", decisions["pass"], "Có thể ưu tiên cho bước xử lý tiếp theo."],
            ["Cần giáo viên xem lại", decisions["need_human_review"], "Còn ít nhất một tiêu chí chưa chắc; không đồng nghĩa với mẫu sai."],
            ["Chưa nên dùng ở lượt hiện tại", decisions["failed"], "Có ít nhất một tiêu chí không đạt rõ."],
            ["Tổng mẫu trong trang cần rà soát", EXPECTED_REVIEW_COUNTS[grade], "Gồm cả hai nhóm cần xem lại và chưa nên dùng."],
            ["Cảnh báo cơ học", alert_count, "Số cảnh báo thiếu trường, định dạng hoặc trùng lặp gắn với mẫu."],
            ["Mẫu nằm trong cặp trùng", duplicate_samples, "Mỗi mẫu trong một cặp trùng được tính một lần."],
        ],
    )
    _style_table(ws)
    _set_widths(ws, {1: 38, 2: 14, 3: 90})


def _add_sample_sheet(wb: Workbook, data: BundleData, grade: str, review: bool) -> None:
    ws = wb.create_sheet(SHEET_NAMES[2] if review else SHEET_NAMES[3])
    _append_rows(ws, [SAMPLE_HEADERS])
    wanted = [
        sid
        for sid, row in data.normalized_by_id.items()
        if row["grade"] == grade
        and ((data.quality_by_id[sid]["quality_decision"] != "pass") == review)
    ]
    if review:
        priority_order = {"Cao": 0, "Trung bình": 1, "": 2}
        wanted.sort(
            key=lambda sid: (
                priority_order.get(_priority(data.review_by_id[sid].get("priority", "")), 3),
                int(data.normalized_by_id[sid]["source_row_number"]),
                sid,
            )
        )
    else:
        wanted = _sort_ids(data, wanted)
    _append_rows(ws, (_sample_values(data, sid) for sid in wanted))
    _style_table(ws)
    _set_widths(
        ws,
        {
            1: 25, 2: 8, 3: 58, 4: 14, 5: 8, 6: 42, 7: 28, 8: 55,
            9: 28, 10: 65, 11: 90, 12: 30, 13: 24, 14: 14, 15: 75,
            16: 30, 17: 50, 18: 16, 19: 65, 20: 55, 21: 18, 22: 18,
            23: 18, 24: 40,
        },
    )
    for row_number in range(2, ws.max_row + 1):
        decision = ws.cell(row_number, 13).value
        color = {"failed": "F4CCCC", "need_human_review": "FFF2CC", "pass": "D9EAD3"}.get(decision)
        if color:
            for column in (12, 13):
                ws.cell(row_number, column).fill = PatternFill("solid", fgColor=color)


def _add_coverage_sheet(wb: Workbook, data: BundleData, grade: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[4])
    total = EXPECTED_GRADE_COUNTS[grade]
    rows: list[list[object]] = [
        ["Nhóm thống kê", "Mã", "Nội dung", "Số mẫu", "Tỷ lệ trong lớp", "Ghi chú"]
    ]
    lesson_rows = sorted(data.coverage_by_grade.get(grade, []), key=lambda row: row.get("lesson_id", ""))
    topic_counts: Counter[tuple[str, str]] = Counter()
    for row in lesson_rows:
        count = int(row["count"])
        rows.append(["Bài học", row.get("lesson_id", ""), row.get("lesson_label", ""), count, count / total, ""])
        topic_counts[(row.get("topic_id", ""), row.get("topic_label", ""))] += count
    for (topic_id, topic_label), count in sorted(topic_counts.items()):
        rows.append(["Chủ đề", topic_id, topic_label, count, count / total, "Tổng hợp lại riêng cho lớp."])
    bloom_counts = Counter(
        _bloom_band(row.get("bloom_level", ""))
        for row in data.normalized_by_id.values()
        if row["grade"] == grade
    )
    for label in ("Nhận biết", "Thông hiểu", "Vận dụng", "Chưa nhận diện rõ"):
        if bloom_counts[label]:
            rows.append(["Mức nhận thức", "", label, bloom_counts[label], bloom_counts[label] / total, ""])
    _append_rows(ws, rows)
    _style_table(ws)
    _set_widths(ws, {1: 22, 2: 20, 3: 65, 4: 14, 5: 18, 6: 38})
    for row_number in range(2, ws.max_row + 1):
        ws.cell(row_number, 5).number_format = "0.0%"


def _add_checklist_sheet(wb: Workbook, data: BundleData, grade: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[5])
    _append_rows(ws, [CHECKLIST_HEADERS])
    sample_ids = _sort_ids(
        data,
        (sid for sid, row in data.normalized_by_id.items() if row["grade"] == grade),
    )
    rows = []
    for sample_id in sample_ids:
        for row in sorted(data.checklist_by_id[sample_id], key=lambda item: item["criterion_id"]):
            result = row.get("result", "")
            rows.append(
                [
                    sample_id,
                    row.get("criterion_id", ""),
                    row.get("criterion_group", ""),
                    row.get("criterion_name", ""),
                    CRITERION_LABELS.get(result, result),
                    result,
                    row.get("confidence_score", ""),
                    row.get("evidence_fragment_id", ""),
                    row.get("evidence_source", ""),
                    row.get("evidence_match_reason", ""),
                    row.get("reason", ""),
                    row.get("suggested_reviewer_action", ""),
                    row.get("checked_by", ""),
                    row.get("checked_at", ""),
                ]
            )
    _append_rows(ws, rows)
    _style_table(ws)
    _set_widths(
        ws,
        {1: 25, 2: 22, 3: 24, 4: 42, 5: 24, 6: 18, 7: 14, 8: 32,
         9: 55, 10: 70, 11: 70, 12: 60, 13: 40, 14: 30},
    )


def _source_kind(path: Path) -> str:
    name = path.name
    kinds = {
        "normalized_dialogue_rows.csv": "Hội thoại đã chuẩn hóa",
        "coverage_summary.csv": "Thống kê độ phủ",
        "missing_field_report.csv": "Cảnh báo thiếu trường/định dạng",
        "duplicate_candidates.csv": "Ứng viên trùng lặp",
        "quality_check_suggestions.csv": "Kết quả canonical theo mẫu",
        "hnmu_review_queue_suggestions.csv": "Danh sách cần rà soát",
    }
    if name.startswith("raw_dialogue_checklist_results"):
        return "Checklist canonical chi tiết"
    return kinds.get(name, "Báo cáo diễn giải cho HNMU")


def _add_sources_sheet(wb: Workbook, data: BundleData, sources: CanonicalSources) -> None:
    ws = wb.create_sheet(SHEET_NAMES[6])
    rows: list[list[object]] = [
        ["STT", "Đường dẫn canonical", "Loại đầu vào", "SHA-256", "Số bản ghi/dòng", "Builder đã đọc"]
    ]
    for index, path in enumerate(sources.all_paths(), start=1):
        rows.append(
            [
                index,
                _display_path(path),
                _source_kind(path),
                data.source_hashes[path],
                data.source_record_counts[path],
                "Có",
            ]
        )
    _append_rows(ws, rows)
    _style_table(ws)
    _set_widths(ws, {1: 8, 2: 120, 3: 38, 4: 68, 5: 18, 6: 18})


def _create_workbook(data: BundleData, sources: CanonicalSources, grade: str, path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = f"Kết quả rà soát hội thoại HNMU lớp {grade}"
    wb.properties.subject = "Đóng gói output canonical Plan 04 cho giáo viên HNMU"
    wb.properties.creator = "UET — Plan 08"
    wb.properties.description = "Kết quả bước đầu; HNMU/UET giữ quyền quyết định chuyên môn."
    _add_instruction_sheet(wb, grade)
    _add_summary_sheet(wb, data, grade)
    _add_sample_sheet(wb, data, grade, review=True)
    _add_sample_sheet(wb, data, grade, review=False)
    _add_coverage_sheet(wb, data, grade)
    _add_checklist_sheet(wb, data, grade)
    _add_sources_sheet(wb, data, sources)
    if wb.sheetnames != SHEET_NAMES:
        raise AssertionError(f"Unexpected sheet order: {wb.sheetnames}")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _cell_text(value: object) -> str:
    return "" if value is None else str(value)


def _rows_by_sample(ws: Worksheet) -> dict[str, dict[str, str]]:
    headers = [_cell_text(cell.value) for cell in ws[1]]
    if headers != SAMPLE_HEADERS:
        raise ValueError(f"Unexpected sample headers in {ws.title}")
    result: dict[str, dict[str, str]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header: _cell_text(value) for header, value in zip(headers, values)}
        sample_id = row["Mã mẫu (sample_id)"]
        if not sample_id or sample_id in result:
            raise ValueError(f"Empty or duplicate sample ID in {ws.title}: {sample_id!r}")
        result[sample_id] = row
    return result


def validate_phase1_teacher_bundle(experiment_dir: Path, bundle_dir: Path) -> dict[str, object]:
    """Reopen all workbooks and compare them with the canonical sources."""

    data = load_canonical_bundle_data(experiment_dir)
    sources = canonical_source_paths(experiment_dir)
    workbook_paths = {
        grade: bundle_dir / f"lop_{grade}" / f"01_ket_qua_ra_soat_lop_{grade}.xlsx"
        for grade in EXPECTED_GRADE_COUNTS
    }
    source_display = [_display_path(path) for path in sources.all_paths()]

    for grade, workbook_path in workbook_paths.items():
        if not workbook_path.is_file():
            raise FileNotFoundError(f"Missing teacher workbook: {workbook_path}")
        wb = load_workbook(workbook_path, read_only=False, data_only=True)
        if wb.sheetnames != SHEET_NAMES:
            raise ValueError(f"Workbook lớp {grade} has inconsistent sheets: {wb.sheetnames}")
        review_rows = _rows_by_sample(wb[SHEET_NAMES[2]])
        pass_rows = _rows_by_sample(wb[SHEET_NAMES[3]])
        if set(review_rows) & set(pass_rows):
            raise ValueError(f"Workbook lớp {grade} duplicates samples across sheets")
        expected_ids = {
            sid for sid, row in data.normalized_by_id.items() if row["grade"] == grade
        }
        if set(review_rows) | set(pass_rows) != expected_ids:
            raise ValueError(f"Workbook lớp {grade} loses or adds sample IDs")
        if len(review_rows) != EXPECTED_REVIEW_COUNTS[grade]:
            raise ValueError(f"Workbook lớp {grade} has wrong review count")

        for sample_id, row in {**review_rows, **pass_rows}.items():
            normalized = data.normalized_by_id[sample_id]
            quality = data.quality_by_id[sample_id]
            expected = {
                "Lớp": _clean(grade),
                "Tệp nguồn (source_file)": _clean(normalized["source_file"]),
                "Dòng nguồn (source_row_number)": _clean(normalized["source_row_number"]),
                "Câu hỏi": _clean(normalized.get("question", "")),
                "Đáp án SGV": _clean(normalized.get("answer_sgv", "")),
                "Hội thoại": _clean(normalized.get("dialogue", "")),
                "Mã kết quả (quality_decision)": _clean(quality["quality_decision"]),
            }
            for field, expected_value in expected.items():
                if row[field] != expected_value:
                    raise ValueError(f"Workbook lớp {grade} changed {field} for {sample_id}")

        checklist_ws = wb[SHEET_NAMES[5]]
        headers = [_cell_text(cell.value) for cell in checklist_ws[1]]
        if headers != CHECKLIST_HEADERS:
            raise ValueError(f"Workbook lớp {grade} checklist headers differ")
        pairs: set[tuple[str, str]] = set()
        counts: Counter[str] = Counter()
        for values in checklist_ws.iter_rows(min_row=2, values_only=True):
            pair = (_cell_text(values[0]), _cell_text(values[1]))
            if pair in pairs:
                raise ValueError(f"Workbook lớp {grade} duplicates checklist key: {pair}")
            pairs.add(pair)
            counts[pair[0]] += 1
        if set(counts) != expected_ids or set(counts.values()) != {18}:
            raise ValueError(f"Workbook lớp {grade} checklist is not 18 criteria/sample")

        provenance_rows = list(wb[SHEET_NAMES[6]].iter_rows(min_row=2, values_only=True))
        if len(provenance_rows) != 15:
            raise ValueError(f"Workbook lớp {grade} provenance must contain 15 inputs")
        if [_cell_text(row[1]) for row in provenance_rows] != source_display:
            raise ValueError(f"Workbook lớp {grade} provenance differs from allowlist")
        for row, path in zip(provenance_rows, sources.all_paths()):
            if _cell_text(row[3]) != data.source_hashes[path]:
                raise ValueError(f"Workbook lớp {grade} source hash differs: {_display_path(path)}")
        wb.close()

    return {
        "status": "ok",
        "workbooks": [path.as_posix() for path in workbook_paths.values()],
        "grade_counts": EXPECTED_GRADE_COUNTS,
        "review_counts": EXPECTED_REVIEW_COUNTS,
        "source_count": 15,
        "source_hashes": {_display_path(path): value for path, value in data.source_hashes.items()},
        "read_paths": sorted(_display_path(path) for path in data.read_paths),
    }


def build_phase1_teacher_bundle(
    experiment_dir: Path,
    bundle_dir: Path,
    *,
    overwrite: bool = False,
) -> BuildSummary:
    """Build four workbooks atomically and validate them before handoff."""

    data = load_canonical_bundle_data(experiment_dir)
    sources = canonical_source_paths(experiment_dir)
    final_paths = tuple(
        bundle_dir / f"lop_{grade}" / f"01_ket_qua_ra_soat_lop_{grade}.xlsx"
        for grade in EXPECTED_GRADE_COUNTS
    )
    existing = [path for path in final_paths if path.exists()]
    if existing and not overwrite:
        raise FileExistsError(
            "Teacher workbook already exists; use --overwrite:\n"
            + "\n".join(f"- {path}" for path in existing)
        )

    bundle_dir.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix=".hnmu-teacher-bundle-", dir=bundle_dir.parent) as temp:
        staged_dir = Path(temp)
        for grade in EXPECTED_GRADE_COUNTS:
            staged_path = staged_dir / f"lop_{grade}" / f"01_ket_qua_ra_soat_lop_{grade}.xlsx"
            _create_workbook(data, sources, grade, staged_path)
        validate_phase1_teacher_bundle(experiment_dir, staged_dir)
        for grade, final_path in zip(EXPECTED_GRADE_COUNTS, final_paths):
            staged_path = staged_dir / f"lop_{grade}" / final_path.name
            final_path.parent.mkdir(parents=True, exist_ok=True)
            staged_path.replace(final_path)

    validation = validate_phase1_teacher_bundle(experiment_dir, bundle_dir)
    return BuildSummary(
        output_paths=final_paths,
        grade_counts=dict(EXPECTED_GRADE_COUNTS),
        review_counts=dict(EXPECTED_REVIEW_COUNTS),
        source_hashes=dict(validation["source_hashes"]),
        read_paths=tuple(validation["read_paths"]),
    )
