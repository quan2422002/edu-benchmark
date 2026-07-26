"""Render HNMU-friendly fragment summaries and traceable technical appendices."""

from __future__ import annotations

import math
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from edu_benchmark.dialogue_audit.fragment_score_analysis import (
    ANALYSIS_HEADERS,
    FRAGMENT_METRICS,
    _summary_row,
)
from edu_benchmark.dialogue_audit.fragment_score_analysis_repaired import (
    HEADER_LABELS,
    _display_value,
    _guide_rows,
    main_conclusion,
)

ROOT_ANALYSIS_NAME = "05_report_fragment_va_ty_le_dat.md"
ROOT_APPENDIX_NAME = "05_phu_luc_ky_thuat_phan_tich_fragment.xlsx"
GRADE_ANALYSIS_NAME = "07_phan_tich_fragment_va_ket_qua_cham.xlsx"
GRADE_APPENDIX_NAME = "08_phu_luc_ky_thuat_phan_tich_fragment.xlsx"
ROOT_SUMMARY_SHEET = "tom_tat_giua_cac_khoi"
GRADE_SUMMARY_SHEET = "tom_tat_lop"
ROOT_APPENDIX_SHEET = "phu_luc_ky_thuat_fragment"
GRADE_APPENDIX_SHEET = "phu_luc_ky_thuat_fragment"
SUMMARY_DATA_HEADER_ROW = 12
TECHNICAL_DATA_HEADER_ROW = 17

SUMMARY_FIELDS = (
    "outcome_label",
    "metric_label",
    "crude_sample_count",
    "adjusted_sample_count",
    "crude_association",
    "crude_significance",
    "adjusted_association",
    "adjusted_significance",
    "controlled_factors",
    "plain_conclusion",
    "note",
    "analysis_key",
)
SUMMARY_HEADERS = (
    "Kết quả chấm",
    "Cách đo fragment",
    "Số mẫu phân tích",
    "Số mẫu sau điều chỉnh",
    "Liên hệ chưa điều chỉnh",
    "Ý nghĩa thống kê chưa điều chỉnh",
    "Liên hệ sau điều chỉnh",
    "Ý nghĩa thống kê sau điều chỉnh",
    "Yếu tố đã kiểm soát",
    "Kết luận dễ hiểu",
    "Lưu ý",
    "Mã đối chiếu",
)
TECHNICAL_FIELDS = (*ANALYSIS_HEADERS, "analysis_key")
TECHNICAL_HEADERS = (*tuple(HEADER_LABELS[field] for field in ANALYSIS_HEADERS), "Mã đối chiếu")

METRIC_LABELS = {
    "fragment_row_count": "Số tiêu chí có dẫn fragment",
    "fragment_reference_count": "Tổng lượt dẫn fragment",
    "unique_fragment_count": "Số fragment khác nhau",
    "fragment_criterion_coverage": "Tỷ lệ tiêu chí có dẫn fragment",
}
OUTCOME_LABELS = {
    "official_pass": "Trạng thái đạt chính thức",
    "checklist_pass_rate": "Tỷ lệ tiêu chí đạt",
}
ANALYSIS_KEYS = {
    ("official_pass", metric): f"FRG-OP-{index:02d}"
    for index, metric in enumerate(FRAGMENT_METRICS, start=1)
} | {
    ("checklist_pass_rate", metric): f"FRG-CR-{index:02d}"
    for index, metric in enumerate(FRAGMENT_METRICS, start=1)
}


def analysis_key(outcome: object, metric: object) -> str:
    try:
        return ANALYSIS_KEYS[(str(outcome), str(metric))]
    except KeyError as error:
        raise ValueError(f"Unknown fragment analysis pair: {outcome!r}/{metric!r}") from error


def _is_estimable(row: dict[str, object] | None) -> bool:
    return bool(
        row
        and row.get("estimable")
        and isinstance(row.get("statistic_value"), (int, float))
    )


def _has_evidence(row: dict[str, object] | None) -> bool:
    return bool(
        _is_estimable(row)
        and isinstance(row.get("p_value"), (int, float))
        and float(row["p_value"]) < 0.05
    )


def _association_text(row: dict[str, object] | None) -> str:
    if not _is_estimable(row):
        return "Không thể ước lượng do dữ liệu không đủ biến thiên"
    value = float(row["statistic_value"])
    absolute = abs(value)
    if absolute < 0.05:
        description = "Hầu như không có liên hệ"
    elif absolute < 0.20:
        description = "Có xu hướng tăng nhẹ" if value > 0 else "Có xu hướng giảm nhẹ"
    elif absolute < 0.40:
        description = "Có xu hướng tăng ở mức vừa" if value > 0 else "Có xu hướng giảm ở mức vừa"
    else:
        description = "Có xu hướng tăng rõ" if value > 0 else "Có xu hướng giảm rõ"
    return f"{description} (hệ số = {value:.2f})"


def _significance_text(row: dict[str, object] | None) -> str:
    if not _is_estimable(row):
        return "Không thể ước lượng"
    return "Có bằng chứng thống kê" if _has_evidence(row) else "Chưa có bằng chứng thống kê"


def _controlled_factors(*, pooled: bool, grade: str) -> str:
    return "Khối lớp và nhóm chấm" if pooled else f"Nhóm chấm trong lớp {grade}"


def _plain_conclusion(
    crude: dict[str, object],
    adjusted: dict[str, object],
    *,
    pooled: bool,
    grade: str,
) -> str:
    controls = "khối lớp và nhóm chấm" if pooled else "nhóm chấm trong lớp"
    crude_evidence = _has_evidence(crude)
    adjusted_evidence = _has_evidence(adjusted)
    if not _is_estimable(adjusted):
        return (
            f"Không thể ước lượng mối liên hệ sau khi kiểm soát {controls} vì các nhóm không đủ biến thiên. "
            "Kết quả chưa điều chỉnh không được xem là bằng chứng về mối liên hệ độc lập."
        )
    crude_value = float(crude["statistic_value"])
    adjusted_value = float(adjusted["statistic_value"])
    reversed_direction = crude_value * adjusted_value < 0
    disappeared = crude_evidence and not adjusted_evidence
    if reversed_direction:
        return (
            f"Mối liên hệ đổi chiều sau khi kiểm soát {controls}; kết quả chưa ổn định và không nên diễn giải độc lập."
        )
    if disappeared:
        return (
            f"Có bằng chứng trong phân tích chưa điều chỉnh, nhưng chưa còn bằng chứng sau khi kiểm soát {controls}. "
            "Kết quả có thể phản ánh khác biệt giữa các nhóm dữ liệu."
        )
    if crude_evidence and adjusted_evidence:
        return (
            f"Quan sát thấy mối liên hệ cả trước và sau khi kiểm soát {controls}. "
            "Đây vẫn là mối liên hệ quan sát được, không chứng minh quan hệ nhân quả."
        )
    if not crude_evidence and adjusted_evidence:
        return (
            f"Chỉ quan sát thấy bằng chứng sau khi kiểm soát {controls}; cần thận trọng vì kết quả chưa nhất quán giữa hai cách phân tích."
        )
    return (
        f"Chưa có bằng chứng thống kê về mối liên hệ, kể cả sau khi kiểm soát {controls}."
    )


def _plain_note(
    crude: dict[str, object],
    adjusted: dict[str, object],
    *,
    pooled: bool,
) -> str:
    crude_count = int(crude["sample_count"])
    adjusted_count = int(adjusted["sample_count"])
    if not _is_estimable(adjusted):
        scope = "khối lớp và nhóm chấm" if pooled else "các nhóm chấm trong lớp"
        return f"Không có nhóm {scope} đủ biến thiên đồng thời về kết quả chấm và cách đo fragment."
    if adjusted_count < crude_count:
        return (
            f"Phân tích sau điều chỉnh dùng {adjusted_count}/{crude_count} mẫu thuộc các nhóm có đủ biến thiên."
        )
    return "Đọc cùng phụ lục kỹ thuật khi cần kiểm tra hệ số, p-value và phương pháp."


def build_hnmu_summary_rows(
    rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None = None,
) -> list[dict[str, object]]:
    adjustment = (
        "adjusted_for_grade_and_auditor_group"
        if pooled
        else "adjusted_for_auditor_group"
    )
    grade_value = "all" if pooled else str(grade or "")
    output: list[dict[str, object]] = []
    for family, outcome in (
        ("fragment_vs_official_pass", "official_pass"),
        ("fragment_vs_checklist_pass_rate", "checklist_pass_rate"),
    ):
        for metric in FRAGMENT_METRICS:
            crude = _summary_row(rows, family, metric, "crude")
            adjusted = _summary_row(rows, family, metric, adjustment)
            if crude is None or adjusted is None:
                raise ValueError(f"Missing summary pair for {grade_value}/{outcome}/{metric}")
            output.append(
                {
                    "outcome_label": OUTCOME_LABELS[outcome],
                    "metric_label": METRIC_LABELS[metric],
                    "crude_sample_count": int(crude["sample_count"]),
                    "adjusted_sample_count": int(adjusted["sample_count"]),
                    "crude_association": _association_text(crude),
                    "crude_significance": _significance_text(crude),
                    "adjusted_association": _association_text(adjusted),
                    "adjusted_significance": _significance_text(adjusted),
                    "controlled_factors": _controlled_factors(
                        pooled=pooled,
                        grade=grade_value,
                    ),
                    "plain_conclusion": _plain_conclusion(
                        crude,
                        adjusted,
                        pooled=pooled,
                        grade=grade_value,
                    ),
                    "note": _plain_note(crude, adjusted, pooled=pooled),
                    "analysis_key": analysis_key(outcome, metric),
                }
            )
    return output


def hnmu_main_conclusion(
    summary_rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None = None,
) -> str:
    if len(summary_rows) != 8:
        raise ValueError(f"Expected 8 HNMU summary rows, found {len(summary_rows)}")
    adjusted_unavailable = sum(
        row["adjusted_significance"] == "Không thể ước lượng" for row in summary_rows
    )
    crude_evidence = sum(
        row["crude_significance"] == "Có bằng chứng thống kê" for row in summary_rows
    )
    adjusted_evidence = sum(
        row["adjusted_significance"] == "Có bằng chứng thống kê" for row in summary_rows
    )
    changed = sum(
        "đổi chiều" in str(row["plain_conclusion"]).casefold()
        or "chưa còn bằng chứng" in str(row["plain_conclusion"]).casefold()
        or "chưa nhất quán" in str(row["plain_conclusion"]).casefold()
        for row in summary_rows
    )
    if adjusted_unavailable == 8:
        scope = "dữ liệu gộp" if pooled else f"lớp {grade}"
        controls = "khối lớp và nhóm chấm" if pooled else "nhóm chấm"
        return (
            f"Không thể ước lượng cả 8 mối liên hệ sau điều chỉnh trong {scope} vì không có {controls} nào đủ biến thiên. "
            "Các kết quả chưa điều chỉnh không được xem là bằng chứng về mối liên hệ độc lập."
        )
    if crude_evidence and (changed or adjusted_evidence < crude_evidence):
        controls = "khối lớp và nhóm chấm" if pooled else "nhóm chấm trong lớp"
        return (
            "Có một số mối liên hệ trong phân tích chưa điều chỉnh, nhưng kết quả không nhất quán hoặc thay đổi "
            f"sau khi kiểm soát {controls}. Vì vậy, chưa có bằng chứng cho thấy mức độ sử dụng fragment có mối liên hệ "
            "độc lập và ổn định với kết quả chấm."
        )
    if adjusted_evidence:
        controls = "khối lớp và nhóm chấm" if pooled else "nhóm chấm trong lớp"
        return (
            f"Một số mối liên hệ vẫn có bằng chứng thống kê sau khi kiểm soát {controls}, nhưng chiều và độ mạnh "
            "cần được đọc theo từng dòng. Kết quả không chứng minh quan hệ nhân quả."
        )
    return (
        "Chưa có bằng chứng thống kê cho thấy mức độ sử dụng fragment có mối liên hệ độc lập và ổn định với kết quả chấm."
    )


def _summary_guide_rows(conclusion: str, *, pooled: bool) -> list[tuple[str, str]]:
    adjusted_text = (
        "So sánh sau khi đã tính đến khác biệt giữa khối lớp và nhóm chấm."
        if pooled
        else "So sánh sau khi đã tính đến khác biệt giữa các nhóm chấm trong lớp."
    )
    return [
        ("BẢN TÓM TẮT PHÂN TÍCH DÀNH CHO HNMU", ""),
        ("Fragment", "Mã dẫn tới phần học liệu được người chấm dùng làm bằng chứng cho một tiêu chí."),
        ("Bốn cách đo", "Số tiêu chí có dẫn; tổng lượt dẫn; số fragment khác nhau; tỷ lệ tiêu chí có dẫn."),
        ("Hai kết quả chấm", "Trạng thái đạt chính thức và tỷ lệ tiêu chí đạt là hai kết quả khác nhau."),
        ("Chưa điều chỉnh", "Mối liên hệ quan sát trực tiếp, chưa tính đến khác biệt giữa các nhóm dữ liệu."),
        ("Sau điều chỉnh", adjusted_text),
        ("Ba mức kết luận", "Có bằng chứng thống kê; chưa có bằng chứng thống kê; hoặc không thể ước lượng vì dữ liệu thiếu biến thiên."),
        ("Mã đối chiếu", "Dùng mã ở cột cuối để tìm đúng các dòng liên quan trong phụ lục kỹ thuật."),
        ("Giới hạn", "Phân tích mô tả mối liên hệ quan sát được và không chứng minh quan hệ nhân quả."),
        ("KẾT LUẬN TỔNG THỂ", conclusion),
        ("", ""),
    ]


def write_hnmu_summary_workbook(
    path: Path,
    sheet_name: str,
    summary_rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None = None,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    conclusion = hnmu_main_conclusion(summary_rows, pooled=pooled, grade=grade)
    for label, explanation in _summary_guide_rows(conclusion, pooled=pooled):
        worksheet.append([label, explanation])
    worksheet.append(SUMMARY_HEADERS)
    for row in summary_rows:
        worksheet.append([row[field] for field in SUMMARY_FIELDS])

    navy = "1F4E78"
    green = "548235"
    pale_green = "E2F0D9"
    pale_blue = "DDEBF7"
    for cell in worksheet[SUMMARY_DATA_HEADER_ROW]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_number in range(1, SUMMARY_DATA_HEADER_ROW):
        worksheet.cell(row_number, 1).font = Font(bold=True)
        worksheet.cell(row_number, 1).alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
    worksheet["A1"].fill = PatternFill("solid", fgColor=navy)
    worksheet["A1"].font = Font(color="FFFFFF", bold=True)
    worksheet["A10"].fill = PatternFill("solid", fgColor=green)
    worksheet["A10"].font = Font(color="FFFFFF", bold=True)
    worksheet["B10"].fill = PatternFill("solid", fgColor=pale_green)
    for row_number in range(SUMMARY_DATA_HEADER_ROW + 1, worksheet.max_row + 1):
        fill = pale_blue if row_number % 2 else "FFFFFF"
        for cell in worksheet[row_number]:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.cell(row_number, 10).fill = PatternFill("solid", fgColor=pale_green)
        worksheet.cell(row_number, 10).font = Font(bold=True)
    widths = {
        1: 24,
        2: 31,
        3: 16,
        4: 20,
        5: 30,
        6: 24,
        7: 34,
        8: 24,
        9: 25,
        10: 62,
        11: 52,
        12: 17,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.column_dimensions["B"].width = 105
    worksheet.freeze_panes = f"A{SUMMARY_DATA_HEADER_ROW + 1}"
    worksheet.auto_filter.ref = (
        f"A{SUMMARY_DATA_HEADER_ROW}:L{worksheet.max_row}"
    )
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[10].height = 72
    for row_number in range(SUMMARY_DATA_HEADER_ROW + 1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 92
    workbook.save(path)
    workbook.close()


def _technical_value(field: str, row: dict[str, object]) -> object:
    if field == "analysis_key":
        return analysis_key(row["outcome"], row["fragment_metric"])
    return _display_value(field, row.get(field, ""), row)


def write_technical_appendix_workbook(
    path: Path,
    sheet_name: str,
    rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    conclusion = main_conclusion(rows, pooled=pooled)
    for label, explanation in _guide_rows(conclusion, pooled=pooled):
        worksheet.append([label, explanation])
    worksheet.append(TECHNICAL_HEADERS)
    for row in rows:
        worksheet.append([_technical_value(field, row) for field in TECHNICAL_FIELDS])

    navy = "1F4E78"
    for cell in worksheet[TECHNICAL_DATA_HEADER_ROW]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet["A1"].font = Font(bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor=navy)
    worksheet["A14"].font = Font(bold=True, color="FFFFFF")
    worksheet["A14"].fill = PatternFill("solid", fgColor="548235")
    for row_number in range(1, TECHNICAL_DATA_HEADER_ROW):
        worksheet.cell(row_number, 1).font = Font(bold=True)
        worksheet.cell(row_number, 1).alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
    for row_number in range(TECHNICAL_DATA_HEADER_ROW + 1, worksheet.max_row + 1):
        adjustment = str(worksheet.cell(row_number, 5).value or "")
        color = (
            "DDEBF7"
            if adjustment == "crude"
            else "E2F0D9"
            if adjustment.startswith("adjusted")
            else "FFF2CC"
        )
        for cell in worksheet[row_number]:
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(
                wrap_text=cell.column in {6, 18, 19, 28},
                vertical="top",
            )
        for column in (10, 11):
            if isinstance(worksheet.cell(row_number, column).value, (int, float)):
                worksheet.cell(row_number, column).number_format = "0.00%"
        for column in (13, 16, 20, 21, 22, 23, 24, 25):
            if isinstance(worksheet.cell(row_number, column).value, (int, float)):
                worksheet.cell(row_number, column).number_format = "0.0000"
        if isinstance(worksheet.cell(row_number, 14).value, (int, float)):
            worksheet.cell(row_number, 14).number_format = "0.000E+00"
    widths = {
        1: 22,
        2: 34,
        3: 28,
        4: 31,
        5: 39,
        6: 49,
        7: 13,
        8: 12,
        9: 15,
        10: 13,
        11: 18,
        12: 27,
        13: 17,
        14: 15,
        15: 20,
        16: 17,
        17: 15,
        18: 78,
        19: 72,
        20: 12,
        21: 14,
        22: 13,
        23: 14,
        24: 12,
        25: 13,
        26: 20,
        27: 15,
        28: 72,
        29: 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.column_dimensions["B"].width = 105
    worksheet.freeze_panes = f"A{TECHNICAL_DATA_HEADER_ROW + 1}"
    worksheet.auto_filter.ref = (
        f"A{TECHNICAL_DATA_HEADER_ROW}:AC{worksheet.max_row}"
    )
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[15].height = 54
    workbook.save(path)
    workbook.close()


def _read_workbook_rows(
    path: Path,
    sheet_name: str,
    *,
    header_row: int,
    expected_headers: Sequence[str],
) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, data_only=True)
    if workbook.sheetnames != [sheet_name]:
        raise ValueError(f"{path.name} must have exactly one sheet named {sheet_name}")
    worksheet = workbook[sheet_name]
    headers = tuple(cell.value for cell in worksheet[header_row])
    if headers != tuple(expected_headers):
        raise ValueError(f"Unexpected headers in {path.name}")
    rows = list(worksheet.iter_rows(min_row=header_row + 1, values_only=True))
    if worksheet.freeze_panes != f"A{header_row + 1}" or not worksheet.auto_filter.ref:
        raise ValueError(f"{path.name} misses freeze pane or filter")
    workbook.close()
    return rows


def read_hnmu_summary_workbook(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    return _read_workbook_rows(
        path,
        sheet_name,
        header_row=SUMMARY_DATA_HEADER_ROW,
        expected_headers=SUMMARY_HEADERS,
    )


def read_technical_appendix_workbook(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    return _read_workbook_rows(
        path,
        sheet_name,
        header_row=TECHNICAL_DATA_HEADER_ROW,
        expected_headers=TECHNICAL_HEADERS,
    )


def expected_summary_display_rows(
    rows: Iterable[dict[str, object]],
) -> list[tuple[object, ...]]:
    return [tuple(row[field] for field in SUMMARY_FIELDS) for row in rows]


def expected_technical_display_rows(
    rows: Iterable[dict[str, object]],
) -> list[tuple[object, ...]]:
    return [tuple(_technical_value(field, row) for field in TECHNICAL_FIELDS) for row in rows]


def validate_summary_keys(
    summary_rows: Sequence[dict[str, object]],
    technical_rows: Sequence[dict[str, object]],
) -> None:
    summary_keys = {str(row["analysis_key"]) for row in summary_rows}
    technical_keys = {
        analysis_key(row["outcome"], row["fragment_metric"])
        for row in technical_rows
    }
    if len(summary_rows) != 8 or len(summary_keys) != 8:
        raise ValueError("HNMU summary must contain 8 unique analysis keys")
    if summary_keys != technical_keys:
        raise ValueError("HNMU summary keys do not reconcile with the technical appendix")


def assert_no_nan(rows: Iterable[Sequence[object]]) -> None:
    for row in rows:
        for value in row:
            if isinstance(value, float) and math.isnan(value):
                raise ValueError("NaN found in fragment analysis workbook")
