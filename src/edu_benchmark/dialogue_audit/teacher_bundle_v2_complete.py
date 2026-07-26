"""Build and validate the complete repaired HNMU Phase 1 teacher bundle v2."""

from __future__ import annotations

import csv
import difflib
import hashlib
import math
import re
import shutil
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

from edu_benchmark.dialogue_audit.fragment_score_analysis import (
    FRAGMENT_METRICS,
    _summary_row,
)
from edu_benchmark.dialogue_audit.fragment_analysis_hnmu_compact import (
    GRADE_ANALYSIS_NAME,
    GRADE_APPENDIX_NAME,
    GRADE_APPENDIX_SHEET,
    GRADE_SUMMARY_SHEET,
    ROOT_ANALYSIS_NAME,
    ROOT_APPENDIX_NAME,
    ROOT_APPENDIX_SHEET,
    ROOT_SUMMARY_SHEET,
    SUMMARY_DATA_HEADER_ROW,
    SUMMARY_TITLE,
    TECHNICAL_DATA_HEADER_ROW,
    build_hnmu_summary_rows,
    expected_summary_display_rows,
    expected_technical_display_rows,
    hnmu_main_conclusion,
    read_hnmu_summary_workbook,
    read_technical_appendix_workbook,
    summary_intro_blocks,
    validate_summary_keys,
    write_hnmu_summary_workbook,
    write_technical_appendix_workbook,
)
from edu_benchmark.dialogue_audit.fragment_analysis_root_deliverables import (
    BUCKET_HEADERS,
    DICTIONARY_HEADERS,
    NON_ESTIMABLE_HEADERS,
    READABLE_HEADERS,
    ROOT_BUCKET_SHEET,
    ROOT_DICTIONARY_SHEET,
    ROOT_NON_ESTIMABLE_SHEET,
    ROOT_RAW_SHEET,
    ROOT_READABLE_SHEET,
    ROOT_STATISTICS_SHEET,
    ROOT_TECHNICAL_SHEETS,
    STATISTICS_HEADERS,
    expected_root_bucket_rows,
    expected_root_dictionary_rows,
    expected_root_non_estimable_rows,
    expected_root_readable_rows,
    expected_root_statistics_rows,
    read_root_raw_technical_rows,
    root_report_markdown,
    write_readable_root_technical_workbook,
    write_root_report,
)
from edu_benchmark.dialogue_audit.fragment_score_analysis_repaired import (
    build_repaired_fragment_data,
    prepare_analysis_rows,
)
from edu_benchmark.dialogue_audit.teacher_bundle_v2_hnmu_docs import (
    file_manifest_text_hnmu,
    fragment_report_section_hnmu,
    grade_readme_text_hnmu,
    root_readme_text_hnmu,
)
from edu_benchmark.dialogue_audit.hnmu_audit import (
    RawDialogueRow,
    lesson_code,
    normalize_text,
)
from edu_benchmark.dialogue_audit.teacher_bundle import (
    EXPECTED_CRITERIA_PER_SAMPLE,
    EXPECTED_GRADE_COUNTS,
    BundleData,
    load_canonical_bundle_data,
)
from edu_benchmark.dialogue_audit.teacher_bundle_v2 import (
    CHECKLIST_FIELDS,
    CRITERIA_HEADERS,
    MISSING_FIELDS,
    NORMALIZED_FIELDS,
    QUALITY_FIELDS,
    STATUS_HEADERS,
    _assert_workbook_rows,
    _checklist_rows,
    _chi_square_result,
    _criteria_rows,
    _missing_rows,
    _normalized_rows,
    _quality_rows,
    _status_counts,
    _status_rows,
    _validate_csv_exact,
    _write_csv,
    _write_single_sheet_workbook,
    format_percentage_half_up,
    round_percentage_fraction_half_up,
)

LESSON_CATALOG = Path("shared/learning_resources/registries/sgk_thcs_topic_lesson_map_v0.csv")
ROOT_OUTPUT_NAMES = (
    "README.md",
    "01_bao_cao_tong_quan.md",
    "02_checklist_tieu_chi.xlsx",
    "03_thong_ke_pass_reject_giua_cac_khoi.xlsx",
    "04_thong_ke_do_phu_mau_pass_giua_cac_khoi.xlsx",
    ROOT_ANALYSIS_NAME,
    "06_ket_qua_cham_tong_the_tung_mau.csv",
    "07_mau_thieu_sai_truong_du_lieu.csv",
    "08_ung_vien_trung_lap.csv",
    "09_du_lieu_tho_sau_chuan_hoa.csv",
    ROOT_APPENDIX_NAME,
    "DANH_MUC_FILE.md",
)
GRADE_OUTPUT_NAMES = (
    "README.md",
    "01_du_lieu_tho_sau_chuan_hoa.csv",
    "02_thong_ke_do_phu_mau_pass.xlsx",
    "03_ket_qua_cham_tong_the_tung_mau.csv",
    "04_ket_qua_cham_chi_tiet_tung_tieu_chi.csv",
    "05_mau_thieu_sai_truong_du_lieu.csv",
    "06_ung_vien_trung_lap.csv",
    GRADE_ANALYSIS_NAME,
    GRADE_APPENDIX_NAME,
)
ROOT_QUALITY_NAME = ROOT_OUTPUT_NAMES[6]
ROOT_MISSING_NAME = ROOT_OUTPUT_NAMES[7]
ROOT_DUPLICATE_NAME = ROOT_OUTPUT_NAMES[8]
ROOT_NORMALIZED_NAME = ROOT_OUTPUT_NAMES[9]

COVERAGE_HEADERS = (
    "grade",
    "grade_label",
    "lesson_id",
    "lesson_name",
    "total_sample_count",
    "pass_sample_count",
    "non_pass_sample_count",
    "pass_rate",
    "has_any_pass",
    "coverage_status",
)
DUPLICATE_FIELDS = (
    "grade",
    "sample_id_a",
    "source_file_a",
    "source_row_number_a",
    "grade_b",
    "sample_id_b",
    "source_file_b",
    "source_row_number_b",
    "duplicate_scope",
    "duplicate_type",
    "similarity",
    "note",
)

PATH_LEAK_PATTERNS = (
    ("experiments/", re.compile(r"experiments/", re.IGNORECASE)),
    ("outputs/", re.compile(r"outputs/", re.IGNORECASE)),
    ("shared/", re.compile(r"shared/", re.IGNORECASE)),
    ("absolute Linux path", re.compile(r"(?:^|[\s`\"'])/(?:home|mnt|tmp|var|opt|srv)/", re.IGNORECASE)),
    ("absolute Windows path", re.compile(r"\b[A-Za-z]:[\\/]")),
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _portable_file_name(value: object) -> str:
    text = str(value or "").strip().replace("\\", "/")
    return text.rsplit("/", 1)[-1] if text else ""


def _portable_reference(value: object) -> str:
    text = str(value or "")
    parts = []
    for part in text.split(";"):
        stripped = part.strip()
        folded = stripped.replace("\\", "/")
        if (
            folded.startswith(("shared/", "experiments/", "outputs/", "/"))
            or re.match(r"^[A-Za-z]:/", folded)
        ):
            stripped = _portable_file_name(folded)
        parts.append(stripped)
    return "; ".join(part for part in parts if part)


_EMBEDDED_WINDOWS_PATH = re.compile(r"\b[A-Za-z]:[\\/][^\s\"\x27<>|]+")
_EMBEDDED_LINUX_PATH = re.compile(r"(?<!\w)/(?:home|mnt|tmp|var|opt|srv)/[^\s\"\x27<>|]+")
_EMBEDDED_CODEBASE_PATH = re.compile(r"(?<!\w)(?:experiments|outputs|shared)/[^\s\"\x27<>|]+", re.IGNORECASE)


def _portable_embedded_paths(value: str) -> str:
    text = value
    for pattern in (_EMBEDDED_WINDOWS_PATH, _EMBEDDED_LINUX_PATH, _EMBEDDED_CODEBASE_PATH):
        text = pattern.sub(lambda match: _portable_file_name(match.group(0)), text)
    return text


def _sanitize_rows(rows: Iterable[dict[str, object]]) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for source in rows:
        row = {key: str(value) if value is not None else "" for key, value in source.items()}
        for field in ("source_file", "source_file_a", "source_file_b"):
            if field in row:
                row[field] = _portable_file_name(row[field])
        for field in ("evidence_fragment_id", "evidence_source"):
            if field in row:
                row[field] = _portable_reference(row[field])
        row = {key: _portable_embedded_paths(value) for key, value in row.items()}
        output.append(row)
    return output


def _sample_sort_key(row: dict[str, str]) -> tuple[int, str]:
    return int(row["grade"]), row.get("sample_id", "")


def _normalized_rows_portable(data: BundleData) -> list[dict[str, str]]:
    return sorted(_sanitize_rows(_normalized_rows(data)), key=_sample_sort_key)


def _quality_rows_portable(data: BundleData) -> list[dict[str, str]]:
    return sorted(_sanitize_rows(_quality_rows(data)), key=_sample_sort_key)


def _checklist_rows_portable(data: BundleData) -> list[dict[str, str]]:
    return sorted(
        _sanitize_rows(_checklist_rows(data)),
        key=lambda row: (int(row["grade"]), row["sample_id"], row["criterion_id"]),
    )


def _missing_rows_portable(data: BundleData) -> list[dict[str, str]]:
    return sorted(_sanitize_rows(_missing_rows(data)), key=_sample_sort_key)


def _raw_dialogue_rows(data: BundleData) -> list[RawDialogueRow]:
    rows = []
    for sample_id, source in data.normalized_by_id.items():
        rows.append(
            RawDialogueRow(
                sample_id=sample_id,
                source_file=source["source_file"],
                source_row_number=int(source["source_row_number"]),
                grade=source["grade"],
                grade_label=source.get("grade_label", f"Lớp {source['grade']}"),
                stt=source.get("stt", ""),
                lesson=source.get("lesson", ""),
                position=source.get("position", ""),
                question=source.get("question", ""),
                bloom_level=source.get("bloom_level", ""),
                answer_sgv=source.get("answer_sgv", ""),
                dialogue=source.get("dialogue", ""),
            )
        )
    return sorted(rows, key=lambda row: (int(row.grade), row.sample_id))


def _all_grade_duplicate_candidates(rows: Sequence[RawDialogueRow], threshold: float = 0.96) -> list[dict[str, object]]:
    """Apply the canonical exact and near-duplicate rules across all grades."""

    output: list[dict[str, object]] = []
    by_question: dict[str, list[RawDialogueRow]] = defaultdict(list)
    by_dialogue: dict[str, list[RawDialogueRow]] = defaultdict(list)
    for row in rows:
        by_question[normalize_text(row.question)].append(row)
        by_dialogue[normalize_text(row.dialogue)].append(row)
    for duplicate_type, groups in (("exact_question", by_question), ("exact_dialogue", by_dialogue)):
        for key, group in groups.items():
            if not key or len(group) < 2:
                continue
            base = group[0]
            for other in group[1:]:
                output.append({
                    "duplicate_type": duplicate_type,
                    "sample_id_a": base.sample_id,
                    "sample_id_b": other.sample_id,
                    "similarity": 1.0,
                    "note": "Trùng chính xác sau chuẩn hóa khoảng trắng/chữ thường.",
                })

    compact = [
        (row, normalize_text(" ".join((row.lesson, row.question, row.answer_sgv, row.dialogue))))
        for row in rows
    ]
    for index, (left, left_text) in enumerate(compact):
        if len(left_text) < 120:
            continue
        for right, right_text in compact[index + 1 :]:
            if len(right_text) < 120:
                continue
            if min(len(left_text), len(right_text)) / max(len(left_text), len(right_text)) < threshold:
                continue
            matcher = difflib.SequenceMatcher(None, left_text, right_text)
            if matcher.real_quick_ratio() < threshold or matcher.quick_ratio() < threshold:
                continue
            ratio = matcher.ratio()
            if threshold <= ratio < 1.0:
                scope = "trong cùng lớp" if left.grade == right.grade else "giữa các lớp"
                output.append({
                    "duplicate_type": "near_duplicate_combined_text",
                    "sample_id_a": left.sample_id,
                    "sample_id_b": right.sample_id,
                    "similarity": round(ratio, 4),
                    "note": f"Gần trùng theo bài + câu hỏi + đáp án + hội thoại {scope}.",
                })
    return output


def _duplicate_rows_all_grades(data: BundleData) -> list[dict[str, str]]:
    raw_rows = _raw_dialogue_rows(data)
    candidates = _all_grade_duplicate_candidates(raw_rows, threshold=0.96)
    output: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for candidate in candidates:
        sample_a = str(candidate["sample_id_a"])
        sample_b = str(candidate["sample_id_b"])
        if sample_b < sample_a:
            sample_a, sample_b = sample_b, sample_a
        key = (str(candidate["duplicate_type"]), sample_a, sample_b)
        if key in seen:
            continue
        seen.add(key)
        left = data.normalized_by_id[sample_a]
        right = data.normalized_by_id[sample_b]
        same_grade = left["grade"] == right["grade"]
        output.append(
            {
                "grade": left["grade"],
                "sample_id_a": sample_a,
                "source_file_a": _portable_file_name(left["source_file"]),
                "source_row_number_a": left["source_row_number"],
                "grade_b": right["grade"],
                "sample_id_b": sample_b,
                "source_file_b": _portable_file_name(right["source_file"]),
                "source_row_number_b": right["source_row_number"],
                "duplicate_scope": "Trong cùng lớp" if same_grade else "Giữa các lớp",
                "duplicate_type": str(candidate["duplicate_type"]),
                "similarity": str(candidate.get("similarity", "")),
                "note": str(candidate.get("note", "")),
            }
        )
    output.sort(key=lambda row: (int(row["grade"]), int(row["grade_b"]), row["sample_id_a"], row["sample_id_b"], row["duplicate_type"]))
    canonical = {
        (row["sample_id_a"], row["sample_id_b"], row["duplicate_type"])
        for row in data.duplicate_rows
    }
    generated = {(row["sample_id_a"], row["sample_id_b"], row["duplicate_type"]) for row in output}
    if not canonical <= generated:
        raise ValueError(f"Global duplicate run loses canonical candidates: {sorted(canonical - generated)}")
    return output


def _read_lesson_catalog() -> tuple[list[dict[str, str]], str]:
    if not LESSON_CATALOG.is_file():
        raise FileNotFoundError(f"Missing canonical lesson catalog: {LESSON_CATALOG}")
    with LESSON_CATALOG.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = [row for row in csv.DictReader(handle) if row.get("item_type") == "bai_hoc" and row.get("grade") in EXPECTED_GRADE_COUNTS]
    expected = {"6": 17, "7": 16, "8": 20, "9": 22}
    counts = Counter(row["grade"] for row in rows)
    if dict(counts) != expected:
        raise ValueError(f"Canonical lesson catalog counts differ: {dict(counts)}")
    ids = [row["item_id"] for row in rows]
    if len(ids) != len(set(ids)):
        raise ValueError("Canonical lesson catalog has duplicate lesson IDs")
    return rows, _sha256(LESSON_CATALOG)


def _coverage_rows(data: BundleData, catalog: Sequence[dict[str, str]]) -> list[list[object]]:
    total: Counter[tuple[str, str]] = Counter()
    passed: Counter[tuple[str, str]] = Counter()
    for sample_id, row in data.normalized_by_id.items():
        key = (row["grade"], lesson_code(row["lesson"]))
        total[key] += 1
        if data.quality_by_id[sample_id]["quality_decision"] == "pass":
            passed[key] += 1
    output: list[list[object]] = []
    catalog_keys: set[tuple[str, str]] = set()
    for lesson in catalog:
        grade = lesson["grade"]
        lesson_id = lesson["item_id"]
        match = re.search(r"-B0*(\d+)([AB]?)$", lesson_id, re.IGNORECASE)
        if not match:
            raise ValueError(f"Cannot derive lesson code from catalog ID: {lesson_id}")
        code = f"{int(match.group(1))}{match.group(2).upper()}"
        key = (grade, code)
        catalog_keys.add(key)
        total_count = total[key]
        pass_count = passed[key]
        if pass_count > total_count:
            raise ValueError(f"Pass count exceeds total for {lesson_id}")
        if total_count == 0:
            pass_rate = None
            status = "không có mẫu dữ liệu"
        elif pass_count == 0:
            pass_rate = 0.0
            status = "không có mẫu pass"
        else:
            pass_rate = round_percentage_fraction_half_up(pass_count, total_count)
            status = "có mẫu pass"
        output.append(
            [
                grade,
                f"Lớp {grade}",
                lesson_id,
                lesson["source_label"],
                total_count,
                pass_count,
                total_count - pass_count,
                pass_rate,
                pass_count > 0,
                status,
            ]
        )
    if set(total) - catalog_keys:
        raise ValueError(f"Samples map outside lesson catalog: {sorted(set(total) - catalog_keys)}")
    for grade, expected in EXPECTED_GRADE_COUNTS.items():
        if sum(row[4] for row in output if row[0] == grade) != expected:
            raise ValueError(f"Coverage total does not reconcile for grade {grade}")
        source_ids = {row["lesson_id"] for row in data.coverage_by_grade[grade]}
        output_ids = {row[2] for row in output if row[0] == grade}
        if source_ids != output_ids:
            raise ValueError(f"Coverage summary and lesson catalog differ for grade {grade}")
    return output


def _filter_grade(rows: Sequence[dict[str, str]], grade: str) -> list[dict[str, str]]:
    return [row for row in rows if row["grade"] == grade]


def _fragment_stat_text(row: dict[str, object] | None) -> str:
    if row is None or not isinstance(row.get("statistic_value"), (int, float)):
        return "không thể ước lượng"
    p_value = row.get("p_value")
    p_text = f", p={float(p_value):.4g}" if isinstance(p_value, (int, float)) else ""
    return f"{float(row['statistic_value']):.3f}{p_text}"


def _fragment_report_section(root_rows: Sequence[dict[str, object]]) -> str:
    lines = [
        "## Fragment đầy đủ hơn có đi kèm tỷ lệ đạt cao hơn không",
        "",
        "Phân tích đặt câu hỏi liệu mức độ ghi tham chiếu fragment có đi cùng kết quả chấm ở cấp mẫu hay không. "
        "Bốn cách đo gồm số tiêu chí có fragment, tổng lượt tham chiếu, số fragment phân biệt và tỷ lệ tiêu chí có fragment. "
        "Hai kết quả được xét riêng: trạng thái pass chính thức và tỷ lệ tiêu chí đạt.",
        "",
        "Phân tích thô dùng tương quan point-biserial cho trạng thái pass chính thức và Spearman rho cho tỷ lệ tiêu chí đạt. "
        "Phân tích sau điều chỉnh dùng demeaning hoặc rank residualization trong strata: file lớp kiểm soát nhóm auditor/shard; "
        "file gộp kiểm soát đồng thời khối lớp và nhóm auditor/shard.",
        "",
        "| Kết quả | Cách đo fragment | Thô | Sau kiểm soát khối lớp và nhóm auditor/shard |",
        "|---|---|---:|---:|",
    ]
    for family, outcome in (
        ("fragment_vs_official_pass", "Trạng thái pass chính thức"),
        ("fragment_vs_checklist_pass_rate", "Tỷ lệ tiêu chí đạt"),
    ):
        for metric in FRAGMENT_METRICS:
            crude = _summary_row(root_rows, family, metric, "crude")
            adjusted = _summary_row(root_rows, family, metric, "adjusted_for_grade_and_auditor_group")
            lines.append(f"| {outcome} | {metric} | {_fragment_stat_text(crude)} | {_fragment_stat_text(adjusted)} |")
    lines.extend(
        [
            "",
            "Kết quả thô phải được đọc cùng kết quả sau điều chỉnh vì cách ghi fragment khác nhau giữa khối, shard và auditor. "
            "Một số nhóm không có đủ biến thiên để ước lượng; các dòng này được ghi rõ trong workbook và không có p-value giả.",
            "",
            main_conclusion(root_rows, pooled=True),
            "",
            "Phân tích này là quan sát hậu kiểm, không chứng minh rằng fragment tạo ra hoặc cải thiện kết quả chấm.",
        ]
    )
    return "\n".join(lines)


def _root_readme_text(row_counts: dict[str, int], duplicate_count: int) -> str:
    return f"""# Bộ kết quả rà soát Phase 1 gửi HNMU

Bộ hồ sơ gồm phần tổng hợp ở thư mục này và bốn thư mục riêng `lop_6/`, `lop_7/`, `lop_8/`, `lop_9/`.

## Nên đọc theo thứ tự

1. Đọc `01_bao_cao_tong_quan.md` để xem kết quả chung.
2. Mở `02_checklist_tieu_chi.xlsx` để xem định nghĩa chung của các tiêu chí.
3. Mở `03_thong_ke_pass_reject_giua_cac_khoi.xlsx` và `04_thong_ke_do_phu_mau_pass_giua_cac_khoi.xlsx` để so sánh bốn khối.
4. Mở `05_phan_tich_fragment_va_ket_qua_cham_giua_cac_khoi.xlsx`; đọc khối “Hướng dẫn đọc bảng” và “Kết luận chính” ở đầu sheet trước khi xem từng dòng.
5. Dùng bốn CSV tổng hợp ở root để tra cứu toàn bộ lớp 6–9.
6. Khi cần xem từng tiêu chí, mở file 04 trong đúng thư mục lớp; không có checklist chi tiết gộp lớn ở root.

## Các file tổng hợp mới

- `06_ket_qua_cham_tong_the_tung_mau.csv`: {row_counts['quality']} mẫu, mỗi mẫu một dòng.
- `07_mau_thieu_sai_truong_du_lieu.csv`: {row_counts['missing']} cảnh báo.
- `08_ung_vien_trung_lap.csv`: {duplicate_count} ứng viên, gồm cả trùng trong lớp và giữa lớp.
- `09_du_lieu_tho_sau_chuan_hoa.csv`: {row_counts['normalized']} mẫu chuẩn hóa, mỗi mẫu một dòng.

File duplicate được tính lại trên toàn bộ 1.050 mẫu bằng ba quy tắc: trùng câu hỏi sau chuẩn hóa, trùng hội thoại sau chuẩn hóa, và gần trùng nội dung kết hợp ở ngưỡng 0,96. Cột `duplicate_scope` phân biệt “Trong cùng lớp” với “Giữa các lớp”.

## Nguồn checklist và đường dẫn

Kết quả chi tiết lớp 6–7 dùng `raw_dialogue_checklist_results.repaired.csv`; lớp 8–9 dùng `raw_dialogue_checklist_results.regex_repaired.csv`. Mỗi mẫu có đủ 18 tiêu chí.

Trong bundle, trường nguồn chỉ giữ tên file, ví dụ `Lớp 6.xlsx`; các liên kết giữa file dùng tên hoặc đường dẫn tương đối như `lop_6/03_ket_qua_cham_tong_the_tung_mau.csv`. Không dùng đường dẫn máy chủ nội bộ.

Ví dụ đúng: tìm một `sample_id` trong file 06, sau đó mở cùng mã ở file 09 hoặc trong thư mục lớp tương ứng.

Ví dụ không đúng: coi mối liên hệ thống kê trong file 05 là bằng chứng rằng nhiều fragment làm mẫu tốt hơn. Phân tích chỉ mô tả mối liên hệ quan sát được.

Trong file 03, `pass` là trạng thái tổng thể chính thức của mẫu. Đây không phải `checklist_pass_rate`, tức tỷ lệ tiêu chí đạt trong phân tích fragment.
"""


def _grade_readme_text(
    grade: str,
    rows: dict[str, Sequence[dict[str, str]]],
    coverage_count: int,
    analysis_count: int,
) -> str:
    status = Counter(row["quality_decision"] for row in rows["quality"])
    return f"""# Kết quả rà soát lớp {grade}

Thư mục này chỉ chứa dữ liệu lớp {grade}. Dùng `sample_id` để đối chiếu giữa các file.

## Số bản ghi

- `01_du_lieu_tho_sau_chuan_hoa.csv`: {len(rows['normalized'])}.
- `02_thong_ke_do_phu_mau_pass.xlsx`: {coverage_count} bài học, kể cả bài không có mẫu pass.
- `03_ket_qua_cham_tong_the_tung_mau.csv`: {len(rows['quality'])}.
- `04_ket_qua_cham_chi_tiet_tung_tieu_chi.csv`: {len(rows['checklist'])}.
- `05_mau_thieu_sai_truong_du_lieu.csv`: {len(rows['missing'])}.
- `06_ung_vien_trung_lap.csv`: {len(rows['duplicates'])} ứng viên trong lớp.
- `07_phan_tich_fragment_va_ket_qua_cham.xlsx`: {analysis_count} dòng kết quả thống kê.

## Trạng thái

- pass: {status['pass']}.
- need_human_review: {status['need_human_review']}.
- failed: {status['failed']}.

File 07 là phân tích riêng lớp {grade}. Đọc “Hướng dẫn đọc bảng” và “Kết luận chính” ở đầu sheet; kết quả adjusted đã kiểm soát nhóm auditor/shard trong lớp. `estimable = Không` nghĩa là nhóm đó không đủ biến thiên để ước lượng, không phải dữ liệu bị mất.

Trong file 07, `official_pass` lấy từ trạng thái tổng thể chính thức, còn `checklist_pass_rate` là tỷ lệ tiêu chí đạt của từng mẫu. Hai kết quả này không đồng nghĩa với nhau.

Ví dụ đúng: dùng kết quả adjusted cùng số mẫu, p-value và cảnh báo để mô tả mức liên hệ.

Ví dụ không đúng: kết luận fragment là nguyên nhân làm mẫu pass. Giáo viên HNMU/UET vẫn giữ quyền đánh giá chuyên môn.
"""


def _overview_text(
    data: BundleData,
    coverage_rows: Sequence[Sequence[object]],
    root_analysis_rows: Sequence[dict[str, object]],
    root_duplicate_count: int,
) -> str:
    counts = _status_counts(data)
    test = _chi_square_result(data)
    lines = [
        "# Báo cáo tổng quan kết quả rà soát Phase 1",
        "",
        "## Quy mô và trạng thái",
        "",
        "| Khối | Tổng | Pass chính thức | Cần xem lại | Failed | Non-pass |",
        "|---:|---:|---:|---:|---:|---:|",
    ]
    for grade in EXPECTED_GRADE_COUNTS:
        total = sum(counts[grade].values())
        non_pass = counts[grade]["need_human_review"] + counts[grade]["failed"]
        lines.append(
            f"| {grade} | {total} | {counts[grade]['pass']} ({format_percentage_half_up(counts[grade]['pass'], total)}) | "
            f"{counts[grade]['need_human_review']} ({format_percentage_half_up(counts[grade]['need_human_review'], total)}) | "
            f"{counts[grade]['failed']} ({format_percentage_half_up(counts[grade]['failed'], total)}) | {non_pass} ({format_percentage_half_up(non_pass, total)}) |"
        )
    lines.extend(
        [
            "",
            "Kiểm định bảng khối × ba trạng thái loại trừ nhau:",
            "",
            f"- Chi-square: {test.statistic:.6f}",
            f"- Bậc tự do: {test.degrees_of_freedom}",
            f"- p-value: {test.p_value:.8g}",
            f"- Cramér’s V: {test.cramers_v:.6f}",
            "",
            "## Độ phủ bài học của mẫu pass",
            "",
        ]
    )
    zero_counts = Counter(str(row[0]) for row in coverage_rows if int(row[5]) == 0)
    for grade in EXPECTED_GRADE_COUNTS:
        lesson_count = sum(str(row[0]) == grade for row in coverage_rows)
        lines.append(f"- Lớp {grade}: {lesson_count} bài trong danh mục; {zero_counts[grade]} bài không có mẫu pass.")
    lines.extend(
        [
            "",
            "Bảng độ phủ bắt đầu từ danh mục bài học đầy đủ rồi mới nối số mẫu và trạng thái. Vì vậy bài có 0 mẫu pass vẫn xuất hiện và được ghi “không có mẫu pass”.",
            "",
            "## Dữ liệu tổng hợp và ứng viên trùng lặp",
            "",
            f"Bốn CSV root cho phép tra toàn bộ lớp 6–9. File ứng viên trùng lặp có {root_duplicate_count} dòng và được chạy lại trên toàn bộ dữ liệu, không phải chỉ nối file lớp.",
            "",
            fragment_report_section_hnmu(root_analysis_rows),
            "",
            "## Giới hạn và quyền quyết định",
            "",
            "Các trạng thái và phân tích là kết quả rà soát hiện tại. Chúng hỗ trợ giáo viên tìm mẫu cần xem, không thay thế phán quyết chuyên môn của HNMU/UET.",
        ]
    )
    return "\n".join(lines) + "\n"


def _grade_rows(
    normalized: Sequence[dict[str, str]],
    quality: Sequence[dict[str, str]],
    checklist: Sequence[dict[str, str]],
    missing: Sequence[dict[str, str]],
    duplicates: Sequence[dict[str, str]],
    grade: str,
) -> dict[str, list[dict[str, str]]]:
    return {
        "normalized": _filter_grade(normalized, grade),
        "quality": _filter_grade(quality, grade),
        "checklist": _filter_grade(checklist, grade),
        "missing": _filter_grade(missing, grade),
        "duplicates": [row for row in duplicates if row["grade"] == grade and row["grade_b"] == grade],
    }


def _build_files(data: BundleData, catalog: Sequence[dict[str, str]], directory: Path) -> None:
    directory.mkdir(parents=True, exist_ok=False)
    normalized = _normalized_rows_portable(data)
    quality = _quality_rows_portable(data)
    checklist = _checklist_rows_portable(data)
    missing = _missing_rows_portable(data)
    duplicates = _duplicate_rows_all_grades(data)
    coverage = _coverage_rows(data, catalog)
    fragment_data = build_repaired_fragment_data(data)
    grade_analysis, root_analysis = prepare_analysis_rows(fragment_data)
    _validate_analysis_scope_rows(grade_analysis, root_analysis)
    root_summary = build_hnmu_summary_rows(root_analysis, pooled=True)
    grade_summaries = {
        grade: build_hnmu_summary_rows(grade_analysis[grade], pooled=False, grade=grade)
        for grade in EXPECTED_GRADE_COUNTS
    }
    validate_summary_keys(root_summary, root_analysis)
    for grade in EXPECTED_GRADE_COUNTS:
        validate_summary_keys(grade_summaries[grade], grade_analysis[grade])

    grade_payloads: dict[str, dict[str, list[dict[str, str]]]] = {}
    for grade in EXPECTED_GRADE_COUNTS:
        grade_dir = directory / f"lop_{grade}"
        grade_dir.mkdir()
        rows = _grade_rows(normalized, quality, checklist, missing, duplicates, grade)
        grade_payloads[grade] = rows
        _write_csv(grade_dir / GRADE_OUTPUT_NAMES[1], NORMALIZED_FIELDS, rows["normalized"])
        _write_csv(grade_dir / GRADE_OUTPUT_NAMES[3], QUALITY_FIELDS, rows["quality"])
        _write_csv(grade_dir / GRADE_OUTPUT_NAMES[4], CHECKLIST_FIELDS, rows["checklist"])
        _write_csv(grade_dir / GRADE_OUTPUT_NAMES[5], MISSING_FIELDS, rows["missing"])
        _write_csv(grade_dir / GRADE_OUTPUT_NAMES[6], DUPLICATE_FIELDS, rows["duplicates"])

    _write_single_sheet_workbook(
        directory / ROOT_OUTPUT_NAMES[2],
        "checklist_tieu_chi",
        CRITERIA_HEADERS,
        _criteria_rows(data),
        {1: 22, 2: 24, 3: 55, 4: 22, 5: 16},
    )
    _write_single_sheet_workbook(
        directory / ROOT_OUTPUT_NAMES[3],
        "pass_reject_giua_cac_khoi",
        STATUS_HEADERS,
        _status_rows(data),
        {1: 10, 2: 16, 3: 24, 4: 34, 5: 14, 6: 16, 7: 18},
        percentage_columns=(6,),
    )
    _write_csv(directory / ROOT_QUALITY_NAME, QUALITY_FIELDS, quality)
    _write_csv(directory / ROOT_MISSING_NAME, MISSING_FIELDS, missing)
    _write_csv(directory / ROOT_DUPLICATE_NAME, DUPLICATE_FIELDS, duplicates)
    _write_csv(directory / ROOT_NORMALIZED_NAME, NORMALIZED_FIELDS, normalized)

    _write_single_sheet_workbook(
        directory / ROOT_OUTPUT_NAMES[4],
        "do_phu_pass_giua_cac_khoi",
        COVERAGE_HEADERS,
        coverage,
        {1: 10, 2: 14, 3: 18, 4: 58, 5: 20, 6: 20, 7: 22, 8: 14, 9: 16, 10: 24},
        percentage_columns=(8,),
    )
    for grade in EXPECTED_GRADE_COUNTS:
        grade_coverage = [row for row in coverage if row[0] == grade]
        _write_single_sheet_workbook(
            directory / f"lop_{grade}" / GRADE_OUTPUT_NAMES[2],
            "do_phu_mau_pass",
            COVERAGE_HEADERS,
            grade_coverage,
            {1: 10, 2: 14, 3: 18, 4: 58, 5: 20, 6: 20, 7: 22, 8: 14, 9: 16, 10: 24},
            percentage_columns=(8,),
        )

    write_root_report(
        directory / ROOT_ANALYSIS_NAME,
        root_summary,
    )
    write_readable_root_technical_workbook(
        directory / ROOT_APPENDIX_NAME,
        root_analysis,
    )
    for grade in EXPECTED_GRADE_COUNTS:
        write_hnmu_summary_workbook(
            directory / f"lop_{grade}" / GRADE_ANALYSIS_NAME,
            GRADE_SUMMARY_SHEET,
            grade_summaries[grade],
            pooled=False,
            grade=grade,
        )
        write_technical_appendix_workbook(
            directory / f"lop_{grade}" / GRADE_APPENDIX_NAME,
            GRADE_APPENDIX_SHEET,
            grade_analysis[grade],
            pooled=False,
        )

    row_counts = {"normalized": len(normalized), "quality": len(quality), "missing": len(missing)}
    (directory / "README.md").write_text(root_readme_text_hnmu(row_counts, duplicates), encoding="utf-8")
    (directory / "01_bao_cao_tong_quan.md").write_text(
        _overview_text(data, coverage, root_analysis, len(duplicates)), encoding="utf-8"
    )
    for grade in EXPECTED_GRADE_COUNTS:
        rows = grade_payloads[grade]
        (directory / f"lop_{grade}/README.md").write_text(
            grade_readme_text_hnmu(
                grade,
                {name: len(values) for name, values in rows.items()},
                sum(row[0] == grade for row in coverage),
                Counter(row["quality_decision"] for row in rows["quality"]),
                len(grade_analysis[grade]),
            ),
            encoding="utf-8",
        )
    (directory / "DANH_MUC_FILE.md").write_text(
        file_manifest_text_hnmu(
            {grade: EXPECTED_GRADE_COUNTS[grade] for grade in EXPECTED_GRADE_COUNTS},
            {
                "all": len(root_analysis),
                **{grade: len(grade_analysis[grade]) for grade in EXPECTED_GRADE_COUNTS},
            },
        ),
        encoding="utf-8",
    )


def _read_csv(path: Path, fields: Sequence[str]) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if tuple(reader.fieldnames or ()) != tuple(fields):
            raise ValueError(f"Unexpected schema in {path.name}: {reader.fieldnames}")
        return list(reader)


def _assert_display_rows(
    actual: Sequence[Sequence[object]],
    expected: Sequence[Sequence[object]],
    *,
    path: Path,
    start_row: int,
) -> None:
    if len(actual) != len(expected):
        raise ValueError(f"{path.name} has {len(actual)} rows; expected {len(expected)}")
    for row_number, (actual_row, expected_row) in enumerate(zip(actual, expected), start=start_row):
        for column, (actual_value, expected_value) in enumerate(zip(actual_row, expected_row), start=1):
            if expected_value == "" and actual_value is None:
                continue
            if isinstance(expected_value, float):
                if not isinstance(actual_value, (int, float)) or not math.isclose(
                    float(actual_value), expected_value, rel_tol=1e-12, abs_tol=1e-12
                ):
                    raise ValueError(f"Numeric mismatch in {path.name} at {row_number},{column}")
            elif actual_value != expected_value:
                raise ValueError(f"Value mismatch in {path.name} at {row_number},{column}")


def _assert_summary_rows(
    path: Path,
    sheet: str,
    expected_rows: Sequence[dict[str, object]],
) -> None:
    actual = read_hnmu_summary_workbook(path, sheet)
    expected = expected_summary_display_rows(expected_rows)
    _assert_display_rows(
        actual,
        expected,
        path=path,
        start_row=SUMMARY_DATA_HEADER_ROW + 1,
    )


def _assert_technical_rows(
    path: Path,
    sheet: str,
    expected_rows: Sequence[dict[str, object]],
) -> None:
    actual = read_technical_appendix_workbook(path, sheet)
    expected = expected_technical_display_rows(expected_rows)
    _assert_display_rows(
        actual,
        expected,
        path=path,
        start_row=TECHNICAL_DATA_HEADER_ROW + 1,
    )


def _assert_summary_guidance(
    path: Path,
    sheet: str,
    summary_rows: Sequence[dict[str, object]],
    *,
    pooled: bool,
    grade: str | None = None,
) -> None:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet]
    if worksheet["A1"].value != SUMMARY_TITLE:
        raise ValueError(f"HNMU summary title is incorrect in {path.name}")
    expected_blocks = summary_intro_blocks(summary_rows, pooled=pooled, grade=grade)
    actual_blocks = tuple(
        (str(worksheet.cell(row, 1).value or ""), str(worksheet.cell(row, 2).value or ""))
        for row in range(2, 5)
    )
    if actual_blocks != expected_blocks:
        raise ValueError(f"HNMU intro blocks differ from current data in {path.name}")
    if any(not worksheet.cell(row, 2).alignment.wrap_text for row in range(2, 5)):
        raise ValueError(f"HNMU intro blocks do not wrap in {path.name}")
    visible_rows = read_hnmu_summary_workbook(path, sheet)
    visible_text = " ".join(
        str(cell.value or "")
        for row in worksheet.iter_rows()
        for cell in row
    ).casefold()
    forbidden = (
        "point-biserial",
        "spearman",
        "effect size",
        "informative strata",
        "non-estimable strata",
        "adjusted_for_",
        "fragment_row_count",
        "fragment_reference_count",
        "unique_fragment_count",
        "fragment_criterion_coverage",
        "official_pass",
        "checklist_pass_rate",
        "p-value",
        "r =",
        "crude",
        "adjusted",
        "estimable",
        "trước điều chỉnh",
        "sau điều chỉnh",
        "có bằng chứng thống kê",
        "mối liên hệ độc lập",
        "sau khi kiểm soát",
        "cùng khối lớp",
        "mã đối chiếu",
        "số mẫu phân tích",
        "yếu tố đã kiểm soát",
    )
    if any(token in visible_text for token in forbidden):
        raise ValueError(f"Technical wording leaks into grade HNMU summary: {path.name}")
    if worksheet.max_column != 5:
        raise ValueError(f"Grade HNMU summary must use five readable columns: {path.name}")
    if any(len(str(row[4]).split()) > 25 for row in visible_rows):
        raise ValueError(f"HNMU interpretation exceeds 25 words in {path.name}")
    if pooled:
        if any(row["controlled_factors"] != "Khối lớp và nhóm chấm" for row in summary_rows):
            raise ValueError("Pooled summary does not control grade and reviewer group")
        adjusted_counts = {
            row["analysis_key"]: row["adjusted_sample_count"] for row in summary_rows
        }
        if any(
            count != (308 if key in {"FRG-OP-03", "FRG-CR-03"} else 350)
            for key, count in adjusted_counts.items()
        ):
            raise ValueError("Pooled adjusted sample counts changed from 350/308")
        interpretations = {
            row["analysis_key"]: row["plain_interpretation"] for row in summary_rows
        }
        expected_interpretations = {
            "FRG-CR-01": "Không thấy bằng chứng về mối liên hệ ở cả hai phân tích.",
            "FRG-OP-01": "Mối liên hệ quan sát được không còn sau điều chỉnh.",
            "FRG-OP-04": "Mối liên hệ quan sát được không còn sau điều chỉnh.",
            "FRG-CR-04": "Mối liên hệ chỉ xuất hiện sau điều chỉnh và cần được diễn giải thận trọng.",
            "FRG-OP-03": "Mối liên hệ đổi chiều sau điều chỉnh nên kết quả chưa ổn định.",
            "FRG-CR-03": "Mối liên hệ đổi chiều sau điều chỉnh nên kết quả chưa ổn định.",
        }
        if any(interpretations[key] != text for key, text in expected_interpretations.items()):
            raise ValueError("Pooled key interpretation no longer matches evidence rules")
    else:
        if any("Khối lớp" in str(row["controlled_factors"]) for row in summary_rows):
            raise ValueError(f"Grade summary incorrectly mentions grade control: {path.name}")
        if grade == "8" and not all(
            row["adjusted_sample_count"] == 0
            and row["adjusted_estimable"] is False
            for row in summary_rows
        ):
            raise ValueError("Grade 8 summary no longer shows 8/8 adjusted analyses as non-estimable")
    workbook.close()




def _assert_root_report(
    path: Path,
    summary_rows: Sequence[dict[str, object]],
) -> None:
    expected = root_report_markdown(summary_rows)
    actual = path.read_text(encoding="utf-8")
    if actual != expected:
        raise ValueError("Root HNMU fragment report differs from current data")
    required_headings = (
        "# Kết quả phân tích fragment và tỷ lệ đạt",
        "## Câu hỏi cần trả lời",
        "## Trả lời ngắn",
        "## Kết quả được hiểu như thế nào?",
        "## Kết luận",
        "## Lưu ý khi diễn giải",
    )
    if tuple(line for line in actual.splitlines() if line.startswith("#")) != required_headings:
        raise ValueError("Root HNMU report headings changed")
    if any(line.lstrip().startswith("|") for line in actual.splitlines()):
        raise ValueError("Root HNMU report must not contain a Markdown table")
    visible_text = actual.casefold()
    forbidden = (
        "hệ số tương quan",
        "p-value",
        "ý nghĩa thống kê",
        "liên hệ dương",
        "liên hệ âm",
        "trước điều chỉnh",
        "sau điều chỉnh",
        "kiểm soát biến",
        "effect size",
        "strata",
        "estimable",
        "fragment_row_count",
        "fragment_reference_count",
        "unique_fragment_count",
        "fragment_criterion_coverage",
        "checklist_pass_rate",
        "official_pass",
        "các mẫu có nhiều tiêu chí được dẫn fragment hơn",
    )
    if any(token in visible_text for token in forbidden):
        raise ValueError("Technical wording leaks into root HNMU report")
    focused = next(
        row
        for row in summary_rows
        if row["analysis_key"] == "FRG-OP-04"
    )
    total_label = f"{int(focused['crude_sample_count']):,}".replace(",", ".")
    comparable_label = f"{int(focused['adjusted_sample_count']):,}".replace(",", ".")
    required_text = (
        "Tỷ lệ tiêu chí có dẫn fragment",
        f"{comparable_label} trong tổng số {total_label} mẫu",
    )
    if any(text not in actual for text in required_text):
        raise ValueError("Root HNMU report misses focused metric or dynamic sample limit")
    if len(actual.split()) > 500:
        raise ValueError("Root HNMU report exceeds one-page reading length")


def _sheet_data_rows(
    worksheet,
    headers: Sequence[str],
) -> list[tuple[object, ...]]:
    actual_headers = tuple(
        worksheet.cell(4, column).value
        for column in range(1, len(headers) + 1)
    )
    if actual_headers != tuple(headers):
        raise ValueError(f"Unexpected headers in sheet {worksheet.title}")
    return list(
        worksheet.iter_rows(
            min_row=5,
            max_col=len(headers),
            values_only=True,
        )
    )


def _assert_root_technical_workbook(
    path: Path,
    root_rows: Sequence[dict[str, object]],
    summary_rows: Sequence[dict[str, object]],
) -> None:
    actual_raw = read_root_raw_technical_rows(path)
    expected_raw = expected_technical_display_rows(root_rows)
    _assert_display_rows(
        actual_raw,
        expected_raw,
        path=path,
        start_row=TECHNICAL_DATA_HEADER_ROW + 1,
    )
    workbook = load_workbook(path, data_only=True)
    if tuple(workbook.sheetnames) != ROOT_TECHNICAL_SHEETS:
        raise ValueError("Root technical workbook sheet order changed")
    if workbook.active.title != ROOT_READABLE_SHEET:
        raise ValueError("Root technical workbook must open on readable results")
    expected_by_sheet = {
        ROOT_READABLE_SHEET: (READABLE_HEADERS, expected_root_readable_rows(root_rows)),
        ROOT_BUCKET_SHEET: (BUCKET_HEADERS, expected_root_bucket_rows(root_rows)),
        ROOT_STATISTICS_SHEET: (STATISTICS_HEADERS, expected_root_statistics_rows(root_rows)),
        ROOT_NON_ESTIMABLE_SHEET: (
            NON_ESTIMABLE_HEADERS,
            expected_root_non_estimable_rows(root_rows),
        ),
        ROOT_DICTIONARY_SHEET: (DICTIONARY_HEADERS, expected_root_dictionary_rows()),
    }
    for sheet_name, (headers, expected_rows) in expected_by_sheet.items():
        worksheet = workbook[sheet_name]
        actual_rows = _sheet_data_rows(worksheet, headers)
        _assert_display_rows(
            actual_rows,
            expected_rows,
            path=path,
            start_row=5,
        )
        if worksheet.freeze_panes != "A5" or len(worksheet.tables) != 1:
            raise ValueError(f"Readable technical sheet misses freeze/filter: {sheet_name}")
    readable = workbook[ROOT_READABLE_SHEET]
    if readable.max_column != 6 or len(expected_by_sheet[ROOT_READABLE_SHEET][1]) != 8:
        raise ValueError("Readable technical sheet must contain 8 results × 6 columns")
    if any(
        len(str(row[4]).split()) > 25
        for row in expected_by_sheet[ROOT_READABLE_SHEET][1]
    ):
        raise ValueError("Readable technical conclusion exceeds 25 words")
    focus_rows = [
        row
        for row in expected_by_sheet[ROOT_READABLE_SHEET][1]
        if row[5] == "FRG-OP-04"
    ]
    if len(focus_rows) != 1:
        raise ValueError("Readable technical sheet misses FRG-OP-04")
    bucket_rows = expected_by_sheet[ROOT_BUCKET_SHEET][1]
    reference_bucket_labels = {
        str(row[1])
        for row in bucket_rows
        if row[0] == "Tổng lượt dẫn fragment"
    }
    required_reference_labels = {
        "Không quá 5 lượt dẫn fragment",
        "Từ 6 đến 7 lượt dẫn fragment",
        "Trên 7 lượt dẫn fragment",
    }
    if not required_reference_labels <= reference_bucket_labels:
        raise ValueError("Fragment reference bucket labels overlap or miss a range")
    non_estimable_rows = expected_by_sheet[ROOT_NON_ESTIMABLE_SHEET][1]
    if any(
        not re.fullmatch(r"Nhóm chấm \d{2}", str(row[1]))
        for row in non_estimable_rows
    ):
        raise ValueError("Non-estimable sheet leaks technical reviewer group labels")
    report = root_report_markdown(summary_rows)
    if "Chưa thể khẳng định" not in str(focus_rows[0][4]) or "**Chưa thể khẳng định.**" not in report:
        raise ValueError("FRG-OP-04 and HNMU report conclusions disagree")
    raw = workbook[ROOT_RAW_SHEET]
    if raw.max_row != 396 or raw.max_column != 29:
        raise ValueError("Raw technical sheet no longer preserves 396 × 29 cells")
    if raw.freeze_panes != f"A{TECHNICAL_DATA_HEADER_ROW + 1}" or not raw.auto_filter.ref:
        raise ValueError("Raw technical sheet loses freeze pane or filter")
    workbook.close()

def _validate_analysis_scope_rows(
    grade_rows: dict[str, list[dict[str, object]]],
    root_rows: Sequence[dict[str, object]],
) -> None:
    for grade, rows in grade_rows.items():
        adjustments = {str(row["adjustment"]) for row in rows}
        if adjustments - {"crude", "adjusted_for_auditor_group"}:
            raise ValueError(f"Grade {grade} analysis contains pooled adjustment labels: {sorted(adjustments)}")
    pooled_adjusted = [
        row
        for row in root_rows
        if row["grade"] == "all" and str(row["adjustment"]).startswith("adjusted_")
    ]
    if not pooled_adjusted or any(
        row["adjustment"] != "adjusted_for_grade_and_auditor_group" for row in pooled_adjusted
    ):
        raise ValueError("Pooled analysis does not consistently use grade + auditor_group adjustment")

    grade8 = grade_rows["8"]
    adjusted_summaries = [
        _summary_row(grade8, family, metric, "adjusted_for_auditor_group")
        for family in ("fragment_vs_official_pass", "fragment_vs_checklist_pass_rate")
        for metric in FRAGMENT_METRICS
    ]
    if any(row is None for row in adjusted_summaries):
        raise ValueError("Grade 8 misses an adjusted summary row")
    if not all(
        row["estimable"] is False
        and row["sample_count"] == 0
        and row["strata_with_variation"] == 0
        and row["strata_total"] == 6
        and row["p_value"] == ""
        for row in adjusted_summaries
        if row is not None
    ):
        raise ValueError("Grade 8 adjusted results no longer match the 0/6 non-estimable invariant")


def _scan_path_leaks(bundle_dir: Path) -> list[str]:
    findings: list[str] = []
    for path in sorted(bundle_dir.rglob("*")):
        if not path.is_file():
            continue
        relative = path.relative_to(bundle_dir).as_posix()
        values: list[str] = []
        if path.suffix.lower() in {".md", ".csv"}:
            values = [path.read_text(encoding="utf-8-sig")]
        elif path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            for worksheet in workbook.worksheets:
                values.extend(str(value) for row in worksheet.iter_rows(values_only=True) for value in row if value is not None)
            workbook.close()
        for value in values:
            for label, pattern in PATH_LEAK_PATTERNS:
                if pattern.search(value):
                    findings.append(f"{relative}: {label}")
                    break
    return findings


def _validate_strict_status(data: BundleData) -> None:
    for sample_id, checklist in data.checklist_by_id.items():
        results = {row["result"] for row in checklist}
        expected = "failed" if "fail" in results else "need_human_review" if "uncertain" in results else "pass"
        actual = data.quality_by_id[sample_id]["quality_decision"]
        if actual != expected:
            raise ValueError(f"Official status does not reconcile for {sample_id}: {actual} != {expected}")


def validate_complete_phase1_teacher_bundle_v2(
    experiment_dir: Path,
    bundle_dir: Path,
    *,
    expected_source_hashes: dict[str, str] | None = None,
) -> dict[str, object]:
    data = load_canonical_bundle_data(experiment_dir)
    catalog, catalog_hash = _read_lesson_catalog()
    current_hashes = {path.as_posix(): digest for path, digest in data.source_hashes.items()}
    current_hashes[LESSON_CATALOG.as_posix()] = catalog_hash
    if expected_source_hashes is not None and current_hashes != expected_source_hashes:
        raise ValueError("Canonical sources changed during complete bundle build")
    _validate_strict_status(data)
    expected_status_counts = {"pass": 665, "need_human_review": 382, "failed": 3}
    if _status_counts(data)["all"] != expected_status_counts:
        raise ValueError("Canonical overall status counts changed")
    if not bundle_dir.is_dir():
        raise FileNotFoundError(f"Missing bundle: {bundle_dir}")
    actual_root = {path.name for path in bundle_dir.iterdir()}
    expected_root = set(ROOT_OUTPUT_NAMES) | {f"lop_{grade}" for grade in EXPECTED_GRADE_COUNTS}
    if actual_root != expected_root:
        raise ValueError(f"Root file set differs: missing={sorted(expected_root-actual_root)}, extra={sorted(actual_root-expected_root)}")

    normalized = _normalized_rows_portable(data)
    quality = _quality_rows_portable(data)
    checklist = _checklist_rows_portable(data)
    missing = _missing_rows_portable(data)
    duplicates = _duplicate_rows_all_grades(data)
    coverage = _coverage_rows(data, catalog)
    fragment_data = build_repaired_fragment_data(data)
    grade_analysis, root_analysis = prepare_analysis_rows(fragment_data)
    _validate_analysis_scope_rows(grade_analysis, root_analysis)
    root_summary = build_hnmu_summary_rows(root_analysis, pooled=True)
    grade_summaries = {
        grade: build_hnmu_summary_rows(grade_analysis[grade], pooled=False, grade=grade)
        for grade in EXPECTED_GRADE_COUNTS
    }
    validate_summary_keys(root_summary, root_analysis)
    for grade in EXPECTED_GRADE_COUNTS:
        validate_summary_keys(grade_summaries[grade], grade_analysis[grade])

    root_normalized = _validate_csv_exact(bundle_dir / ROOT_NORMALIZED_NAME, NORMALIZED_FIELDS, normalized)
    root_quality = _validate_csv_exact(bundle_dir / ROOT_QUALITY_NAME, QUALITY_FIELDS, quality)
    root_missing = _validate_csv_exact(bundle_dir / ROOT_MISSING_NAME, MISSING_FIELDS, missing)
    root_duplicates = _validate_csv_exact(bundle_dir / ROOT_DUPLICATE_NAME, DUPLICATE_FIELDS, duplicates)
    if len({row["sample_id"] for row in root_normalized}) != len(root_normalized) or len(root_normalized) != 1050:
        raise ValueError("Root normalized CSV loses or duplicates sample_id")
    if len({row["sample_id"] for row in root_quality}) != len(root_quality) or len(root_quality) != 1050:
        raise ValueError("Root quality CSV loses or duplicates sample_id")
    if root_normalized != sorted(root_normalized, key=_sample_sort_key) or root_quality != sorted(root_quality, key=_sample_sort_key):
        raise ValueError("Root sample CSVs are not stably sorted by grade and sample_id")
    if any(row["duplicate_scope"] == "Giữa các lớp" and row["grade"] == row["grade_b"] for row in root_duplicates):
        raise ValueError("Cross-grade duplicate label disagrees with grade columns")

    _assert_workbook_rows(bundle_dir / ROOT_OUTPUT_NAMES[2], "checklist_tieu_chi", CRITERIA_HEADERS, _criteria_rows(data))
    _assert_workbook_rows(bundle_dir / ROOT_OUTPUT_NAMES[3], "pass_reject_giua_cac_khoi", STATUS_HEADERS, _status_rows(data))
    _assert_workbook_rows(bundle_dir / ROOT_OUTPUT_NAMES[4], "do_phu_pass_giua_cac_khoi", COVERAGE_HEADERS, coverage)
    _assert_root_report(
        bundle_dir / ROOT_ANALYSIS_NAME,
        root_summary,
    )
    _assert_root_technical_workbook(
        bundle_dir / ROOT_APPENDIX_NAME,
        root_analysis,
        root_summary,
    )

    all_ids: set[str] = set()
    all_pairs: set[tuple[str, str]] = set()
    file_row_counts: dict[str, int] = {
        ROOT_NORMALIZED_NAME: len(root_normalized), ROOT_QUALITY_NAME: len(root_quality),
        ROOT_MISSING_NAME: len(root_missing), ROOT_DUPLICATE_NAME: len(root_duplicates),
        ROOT_OUTPUT_NAMES[2]: len(_criteria_rows(data)), ROOT_OUTPUT_NAMES[3]: len(_status_rows(data)),
        ROOT_OUTPUT_NAMES[4]: len(coverage), ROOT_ANALYSIS_NAME: 1,
        ROOT_APPENDIX_NAME: len(root_analysis),
    }
    for grade, expected_count in EXPECTED_GRADE_COUNTS.items():
        grade_dir = bundle_dir / f"lop_{grade}"
        if {path.name for path in grade_dir.iterdir()} != set(GRADE_OUTPUT_NAMES):
            raise ValueError(f"lop_{grade} file set differs")
        expected_rows = _grade_rows(normalized, quality, checklist, missing, duplicates, grade)
        actual_normalized = _validate_csv_exact(grade_dir / GRADE_OUTPUT_NAMES[1], NORMALIZED_FIELDS, expected_rows["normalized"])
        actual_quality = _validate_csv_exact(grade_dir / GRADE_OUTPUT_NAMES[3], QUALITY_FIELDS, expected_rows["quality"])
        actual_checklist = _validate_csv_exact(grade_dir / GRADE_OUTPUT_NAMES[4], CHECKLIST_FIELDS, expected_rows["checklist"])
        _validate_csv_exact(grade_dir / GRADE_OUTPUT_NAMES[5], MISSING_FIELDS, expected_rows["missing"])
        _validate_csv_exact(grade_dir / GRADE_OUTPUT_NAMES[6], DUPLICATE_FIELDS, expected_rows["duplicates"])
        grade_coverage = [row for row in coverage if row[0] == grade]
        _assert_workbook_rows(grade_dir / GRADE_OUTPUT_NAMES[2], "do_phu_mau_pass", COVERAGE_HEADERS, grade_coverage)
        _assert_summary_rows(
            grade_dir / GRADE_ANALYSIS_NAME,
            GRADE_SUMMARY_SHEET,
            grade_summaries[grade],
        )
        _assert_summary_guidance(
            grade_dir / GRADE_ANALYSIS_NAME,
            GRADE_SUMMARY_SHEET,
            grade_summaries[grade],
            pooled=False,
            grade=grade,
        )
        _assert_technical_rows(
            grade_dir / GRADE_APPENDIX_NAME,
            GRADE_APPENDIX_SHEET,
            grade_analysis[grade],
        )
        expected_grade_readme = grade_readme_text_hnmu(
            grade,
            {name: len(values) for name, values in expected_rows.items()},
            len(grade_coverage),
            Counter(row["quality_decision"] for row in expected_rows["quality"]),
            len(grade_analysis[grade]),
        )
        actual_grade_readme = (grade_dir / "README.md").read_text(encoding="utf-8")
        if actual_grade_readme != expected_grade_readme:
            raise ValueError(f"lop_{grade} README differs from generated file counts")
        ids = {row["sample_id"] for row in actual_normalized}
        pairs = {(row["sample_id"], row["criterion_id"]) for row in actual_checklist}
        if len(ids) != expected_count or ids != {row["sample_id"] for row in actual_quality} or all_ids & ids:
            raise ValueError(f"lop_{grade} sample partition is invalid")
        if len(pairs) != expected_count * EXPECTED_CRITERIA_PER_SAMPLE or all_pairs & pairs:
            raise ValueError(f"lop_{grade} criterion partition is invalid")
        if any(row["grade"] != grade for row in actual_normalized + actual_quality + actual_checklist):
            raise ValueError(f"lop_{grade} contains another grade")
        all_ids.update(ids)
        all_pairs.update(pairs)
        prefix = f"lop_{grade}/"
        file_row_counts.update({
            prefix + GRADE_OUTPUT_NAMES[1]: len(actual_normalized),
            prefix + GRADE_OUTPUT_NAMES[2]: len(grade_coverage),
            prefix + GRADE_OUTPUT_NAMES[3]: len(actual_quality),
            prefix + GRADE_OUTPUT_NAMES[4]: len(actual_checklist),
            prefix + GRADE_OUTPUT_NAMES[5]: len(expected_rows["missing"]),
            prefix + GRADE_OUTPUT_NAMES[6]: len(expected_rows["duplicates"]),
            prefix + GRADE_ANALYSIS_NAME: len(grade_summaries[grade]),
            prefix + GRADE_APPENDIX_NAME: len(grade_analysis[grade]),
        })
    if all_ids != set(data.normalized_by_id) or len(all_pairs) != 18900:
        raise ValueError("Four grade directories do not partition 1,050 samples × 18 criteria")
    if Counter(len(rows) for rows in data.checklist_by_id.values()) != Counter({18: 1050}):
        raise ValueError("Criterion distribution after repair is not 18 for every sample")

    zero_pass_lessons = {
        grade: sum(row[0] == grade and row[5] == 0 for row in coverage)
        for grade in EXPECTED_GRADE_COUNTS
    }
    for path in bundle_dir.rglob("*.xlsx"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if path.name == ROOT_APPENDIX_NAME:
            if tuple(workbook.sheetnames) != ROOT_TECHNICAL_SHEETS:
                raise ValueError("Root technical workbook has unexpected sheets")
        elif len(workbook.sheetnames) != 1:
            raise ValueError(f"Workbook has more than one sheet: {path.name}")
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, float) and math.isnan(value):
                        raise ValueError(f"NaN in workbook: {path.name}")
                    if isinstance(value, str) and value.strip().casefold() == "nan":
                        raise ValueError(f"NaN text in workbook: {path.name}")
        workbook.close()
    leaks = _scan_path_leaks(bundle_dir)
    if leaks:
        raise ValueError(f"Codebase path leaks found: {leaks[:10]}")
    if any(path.name == "04_ket_qua_cham_chi_tiet_tung_tieu_chi.csv" for path in bundle_dir.iterdir()):
        raise ValueError("Root must not contain the large detailed checklist")
    root_readme = (bundle_dir / "README.md").read_text(encoding="utf-8")
    report = (bundle_dir / "01_bao_cao_tong_quan.md").read_text(encoding="utf-8")
    expected_root_readme = root_readme_text_hnmu(
        {"normalized": len(normalized), "quality": len(quality), "missing": len(missing)},
        duplicates,
    )
    if root_readme != expected_root_readme:
        raise ValueError("Root README differs from generated file counts")
    for name in ROOT_OUTPUT_NAMES[1:]:
        if name not in root_readme and name not in {ROOT_OUTPUT_NAMES[1]}:
            raise ValueError(f"Root README misses deliverable: {name}")
    if "Fragment đầy đủ hơn có đi kèm tỷ lệ đạt cao hơn không" not in report:
        raise ValueError("Overview misses repaired fragment section")
    if "308 mẫu" in report and "16 tiêu chí" in report:
        raise ValueError("Overview retains stale missing-criterion statement")
    expected_report = _overview_text(data, coverage, root_analysis, len(duplicates))
    if report != expected_report:
        raise ValueError("Overview report differs from canonical counts or generated analysis")
    focused = next(row for row in root_summary if row["analysis_key"] == "FRG-OP-04")
    total_label = f"{int(focused['crude_sample_count']):,}".replace(",", ".")
    comparable_label = f"{int(focused['adjusted_sample_count']):,}".replace(",", ".")
    if (
        "Tỷ lệ tiêu chí có dẫn fragment" not in report
        or f"{comparable_label} trong tổng số {total_label} mẫu" not in report
    ):
        raise ValueError("Overview misses focused metric or dynamic sample limit")
    if (
        len(duplicates) == 1
        and duplicates[0]["duplicate_scope"] == "Trong cùng lớp"
        and duplicates[0]["grade"] == "9"
        and (
            "Có 1 ứng viên trùng trong lớp 9" not in root_readme
            or "không phát hiện trường hợp trùng giữa các lớp" not in root_readme
        )
    ):
        raise ValueError("Root README misstates duplicate scope")
    expected_manifest = file_manifest_text_hnmu(
        dict(EXPECTED_GRADE_COUNTS),
        {
            "all": len(root_analysis),
            **{grade: len(grade_analysis[grade]) for grade in EXPECTED_GRADE_COUNTS},
        },
    )
    actual_manifest = (bundle_dir / "DANH_MUC_FILE.md").read_text(encoding="utf-8")
    if actual_manifest != expected_manifest:
        raise ValueError("File manifest differs from the actual bundle structure or row counts")

    return {
        "status": "ok",
        "grade_counts": dict(Counter(row["grade"] for row in normalized)),
        "sample_count": len(normalized),
        "criterion_pair_count": len(all_pairs),
        "criterion_count_distribution": {18: 1050},
        "status_counts": _status_counts(data),
        "root_csv_row_counts": {
            ROOT_QUALITY_NAME: len(root_quality), ROOT_MISSING_NAME: len(root_missing),
            ROOT_DUPLICATE_NAME: len(root_duplicates), ROOT_NORMALIZED_NAME: len(root_normalized),
        },
        "cross_grade_duplicate_count": sum(row["duplicate_scope"] == "Giữa các lớp" for row in root_duplicates),
        "zero_pass_lessons": zero_pass_lessons,
        "lesson_counts": dict(Counter(str(row[0]) for row in coverage)),
        "fragment_summary_rows": {
            ROOT_ANALYSIS_NAME: 1,
            **{f"lop_{grade}/{GRADE_ANALYSIS_NAME}": len(grade_summaries[grade]) for grade in EXPECTED_GRADE_COUNTS},
        },
        "fragment_appendix_rows": {
            ROOT_APPENDIX_NAME: len(root_analysis),
            **{f"lop_{grade}/{GRADE_APPENDIX_NAME}": len(grade_analysis[grade]) for grade in EXPECTED_GRADE_COUNTS},
        },
        "workbook_sheets": {
            path.relative_to(bundle_dir).as_posix(): _workbook_sheet_name(path)
            for path in sorted(bundle_dir.rglob("*.xlsx"))
        },
        "source_hashes": current_hashes,
        "path_leak_count": 0,
        "file_row_counts": file_row_counts,
    }


def _workbook_sheet_name(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames[0]
    finally:
        workbook.close()


def rebuild_complete_phase1_teacher_bundle_v2(experiment_dir: Path, bundle_dir: Path) -> dict[str, object]:
    data = load_canonical_bundle_data(experiment_dir)
    catalog, catalog_hash = _read_lesson_catalog()
    initial_hashes = {path.as_posix(): digest for path, digest in data.source_hashes.items()}
    initial_hashes[LESSON_CATALOG.as_posix()] = catalog_hash
    bundle_dir.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix=".hnmu-bundle-v2-complete-", dir=bundle_dir.parent) as temporary:
        temporary_dir = Path(temporary)
        staged = temporary_dir / "bundle"
        _build_files(data, catalog, staged)
        validate_complete_phase1_teacher_bundle_v2(
            experiment_dir, staged, expected_source_hashes=initial_hashes
        )
        if bundle_dir.exists():
            previous = temporary_dir / "previous_bundle"
            bundle_dir.replace(previous)
            try:
                staged.replace(bundle_dir)
            except Exception:
                previous.replace(bundle_dir)
                raise
            shutil.rmtree(previous)
        else:
            staged.replace(bundle_dir)
    return validate_complete_phase1_teacher_bundle_v2(
        experiment_dir, bundle_dir, expected_source_hashes=initial_hashes
    )
