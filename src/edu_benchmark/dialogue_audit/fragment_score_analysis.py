"""Analyze fragment references against canonical HNMU audit outcomes.

This module reads only the two merged raw checklist files and the two merged
canonical quality-decision files.  ``source_file`` values remain opaque: they
are never opened or dereferenced.
"""

from __future__ import annotations

import csv
import hashlib
import math
import statistics
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, Sequence

from openpyxl import load_workbook

from edu_benchmark.dialogue_audit.teacher_bundle import EXPECTED_GRADE_COUNTS
from edu_benchmark.dialogue_audit.teacher_bundle_v2 import _write_single_sheet_workbook

EXPECTED_PASS_COUNTS = {"6": 106, "7": 132, "8": 209, "9": 218}
EXPECTED_CRITERION_DISTRIBUTION_67 = {16: 308, 18: 154}
ROOT_ANALYSIS_NAME = "05_phan_tich_fragment_va_ket_qua_cham_giua_cac_khoi.xlsx"
GRADE_ANALYSIS_NAME = "07_phan_tich_fragment_va_ket_qua_cham.xlsx"
ROOT_SHEET_NAME = "fragment_giua_cac_khoi"
GRADE_SHEET_NAME = "fragment_va_ket_qua_cham"
REPORT_HEADING = "## Mối liên hệ giữa fragment và kết quả chấm"
README_HEADING = "## Phân tích fragment và kết quả chấm"

FRAGMENT_METRICS = (
    "fragment_row_count",
    "fragment_reference_count",
    "unique_fragment_count",
    "fragment_criterion_coverage",
)

ANALYSIS_HEADERS = (
    "grade",
    "analysis_family",
    "outcome",
    "fragment_metric",
    "adjustment",
    "grouping_or_bucket",
    "sample_count",
    "pass_count",
    "non_pass_count",
    "pass_rate",
    "pass_rate_difference_from_lowest_bucket",
    "statistic_name",
    "statistic_value",
    "p_value",
    "effect_size_name",
    "effect_size_value",
    "estimable",
    "interpretation",
    "warning",
    "metric_min",
    "metric_q1",
    "metric_median",
    "metric_q3",
    "metric_max",
    "metric_mean",
    "strata_with_variation",
    "strata_total",
    "method",
)

ANALYSIS_WIDTHS = {
    1: 10,
    2: 28,
    3: 24,
    4: 30,
    5: 38,
    6: 48,
    7: 14,
    8: 12,
    9: 16,
    10: 14,
    11: 20,
    12: 28,
    13: 18,
    14: 16,
    15: 22,
    16: 18,
    17: 12,
    18: 72,
    19: 72,
    20: 14,
    21: 14,
    22: 14,
    23: 14,
    24: 14,
    25: 14,
    26: 22,
    27: 14,
    28: 72,
}


@dataclass(frozen=True)
class FragmentAnalysisSources:
    checklist_67: Path
    checklist_89: Path
    quality_67: Path
    quality_89: Path

    @classmethod
    def from_experiment(cls, experiment_dir: Path) -> "FragmentAnalysisSources":
        return cls(
            checklist_67=experiment_dir
            / "outputs/hnmu_dialogue_audit/agent_shard_audit/merged/raw_dialogue_checklist_results.csv",
            checklist_89=experiment_dir
            / "outputs/hnmu_dialogue_audit_grade8_9/agent_shard_audit/merged/raw_dialogue_checklist_results.csv",
            quality_67=experiment_dir
            / "outputs/hnmu_dialogue_audit/agent_shard_audit/merged/quality_check_suggestions.csv",
            quality_89=experiment_dir
            / "outputs/hnmu_dialogue_audit_grade8_9/agent_shard_audit/merged/quality_check_suggestions.csv",
        )

    def paths(self) -> tuple[Path, ...]:
        return (self.checklist_67, self.checklist_89, self.quality_67, self.quality_89)


@dataclass(frozen=True)
class SampleAnalysisRecord:
    sample_id: str
    grade: str
    checked_by: str
    unified_shard: str
    auditor_group: str
    official_overall_status: str
    official_pass: int
    official_non_pass: int
    observed_criterion_count: int
    applicable_criterion_count: int
    criterion_pass_count: int
    checklist_pass_rate: float
    fragment_row_count: int
    fragment_reference_count: int
    unique_fragment_count: int
    has_any_fragment: int
    fragment_criterion_coverage: float


@dataclass(frozen=True)
class FragmentAnalysisData:
    records: tuple[SampleAnalysisRecord, ...]
    source_hashes: dict[Path, str]
    criterion_pair_count: int
    criterion_count_distribution_67: dict[int, int]
    join_success_count: int
    join_failure_count: int


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.is_file():
        raise FileNotFoundError(f"Missing canonical fragment-analysis source: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = list(reader.fieldnames or [])
        return fields, list(reader)


def _require_fields(path: Path, fields: Sequence[str], required: set[str]) -> None:
    missing = required - set(fields)
    if missing:
        raise ValueError(f"{path} misses required columns: {sorted(missing)}")


def load_fragment_analysis_data(experiment_dir: Path) -> FragmentAnalysisData:
    """Load and join the four explicit canonical sources at sample level."""

    sources = FragmentAnalysisSources.from_experiment(experiment_dir)
    missing_paths = [path for path in sources.paths() if not path.is_file()]
    if missing_paths:
        raise FileNotFoundError(
            "Missing canonical fragment-analysis sources: "
            + ", ".join(path.as_posix() for path in missing_paths)
        )
    initial_hashes = {path: _sha256(path) for path in sources.paths()}

    quality_by_id: dict[str, dict[str, str]] = {}
    for path in (sources.quality_67, sources.quality_89):
        fields, rows = _read_csv(path)
        _require_fields(path, fields, {"sample_id", "grade", "quality_decision"})
        for row in rows:
            sample_id = row["sample_id"].strip()
            if not sample_id or sample_id in quality_by_id:
                raise ValueError(f"Missing or duplicate sample_id in canonical quality source: {sample_id!r}")
            quality_by_id[sample_id] = row

    checklist_by_id: dict[str, list[dict[str, str]]] = defaultdict(list)
    shard_field_by_id: dict[str, str] = {}
    criterion_pairs: set[tuple[str, str]] = set()
    for path, shard_field in (
        (sources.checklist_67, "source_shard"),
        (sources.checklist_89, "shard_id"),
    ):
        fields, rows = _read_csv(path)
        _require_fields(
            path,
            fields,
            {
                "sample_id",
                "criterion_id",
                "result",
                "evidence_fragment_id",
                "checked_by",
                shard_field,
            },
        )
        for row in rows:
            sample_id = row["sample_id"].strip()
            pair = (sample_id, row["criterion_id"].strip())
            if not sample_id or not pair[1] or pair in criterion_pairs:
                raise ValueError(f"Missing or duplicate checklist key: {pair}")
            criterion_pairs.add(pair)
            checklist_by_id[sample_id].append(row)
            previous = shard_field_by_id.setdefault(sample_id, shard_field)
            if previous != shard_field:
                raise ValueError(f"Sample appears in both checklist schemas: {sample_id}")

    quality_ids = set(quality_by_id)
    checklist_ids = set(checklist_by_id)
    failed_ids = sorted(quality_ids ^ checklist_ids)
    if failed_ids:
        raise ValueError(
            f"Checklist/official-status join failed for {len(failed_ids)} sample_id: {failed_ids[:10]}"
        )
    if len(quality_ids) != 1050:
        raise ValueError(f"Expected 1,050 joined sample_id values, found {len(quality_ids)}")

    records: list[SampleAnalysisRecord] = []
    seen_grades: dict[str, set[str]] = defaultdict(set)
    criterion_count_distribution_67: Counter[int] = Counter()
    allowed_results = {"pass", "uncertain", "fail", "not_applicable"}
    allowed_statuses = {"pass", "need_human_review", "failed", "fail", "reject"}
    for sample_id in sorted(quality_ids):
        quality = quality_by_id[sample_id]
        grade = quality["grade"].strip()
        if grade not in EXPECTED_GRADE_COUNTS:
            raise ValueError(f"Unexpected grade for {sample_id}: {grade!r}")
        if sample_id in seen_grades[grade]:
            raise ValueError(f"Duplicate sample in grade {grade}: {sample_id}")
        seen_grades[grade].add(sample_id)
        checklist = checklist_by_id[sample_id]
        results = [row["result"].strip().casefold() for row in checklist]
        unexpected_results = set(results) - allowed_results
        if unexpected_results:
            raise ValueError(f"Unexpected criterion result for {sample_id}: {unexpected_results}")
        checked_by_values = {row["checked_by"].strip() for row in checklist}
        shard_field = shard_field_by_id[sample_id]
        shard_values = {row[shard_field].strip() for row in checklist}
        if len(checked_by_values) != 1 or len(shard_values) != 1 or "" in checked_by_values | shard_values:
            raise ValueError(f"Inconsistent auditor/shard within sample {sample_id}")
        checked_by = next(iter(checked_by_values))
        unified_shard = next(iter(shard_values))
        status = quality["quality_decision"].strip().casefold()
        if status not in allowed_statuses:
            raise ValueError(f"Unexpected official status for {sample_id}: {status!r}")
        official_pass = int(status == "pass")
        applicable = sum(result in {"pass", "uncertain", "fail"} for result in results)
        if not applicable:
            raise ValueError(f"No applicable criterion for {sample_id}")
        fragment_rows = sum(bool(row["evidence_fragment_id"].strip()) for row in checklist)
        fragment_ids = [
            value.strip()
            for row in checklist
            for value in row["evidence_fragment_id"].split(";")
            if value.strip()
        ]
        record = SampleAnalysisRecord(
            sample_id=sample_id,
            grade=grade,
            checked_by=checked_by,
            unified_shard=unified_shard,
            auditor_group=f"{checked_by} | {unified_shard}",
            official_overall_status=status,
            official_pass=official_pass,
            official_non_pass=1 - official_pass,
            observed_criterion_count=len(checklist),
            applicable_criterion_count=applicable,
            criterion_pass_count=sum(result == "pass" for result in results),
            checklist_pass_rate=sum(result == "pass" for result in results) / applicable,
            fragment_row_count=fragment_rows,
            fragment_reference_count=len(fragment_ids),
            unique_fragment_count=len(set(fragment_ids)),
            has_any_fragment=int(bool(fragment_ids)),
            fragment_criterion_coverage=fragment_rows / applicable,
        )
        records.append(record)
        if grade in {"6", "7"}:
            criterion_count_distribution_67[len(checklist)] += 1

    actual_grade_counts = Counter(record.grade for record in records)
    if dict(actual_grade_counts) != dict(EXPECTED_GRADE_COUNTS):
        raise ValueError(f"Wrong grade partition: {dict(actual_grade_counts)}")
    actual_pass = Counter(record.grade for record in records if record.official_pass)
    if dict(actual_pass) != EXPECTED_PASS_COUNTS:
        raise ValueError(f"Official pass counts do not match the approved totals: {dict(actual_pass)}")
    if dict(criterion_count_distribution_67) != EXPECTED_CRITERION_DISTRIBUTION_67:
        raise ValueError(
            "Grade 6–7 criterion-count distribution differs: "
            f"{dict(criterion_count_distribution_67)}"
        )
    final_hashes = {path: _sha256(path) for path in sources.paths()}
    if final_hashes != initial_hashes:
        raise ValueError("Canonical fragment-analysis sources changed while being read")
    return FragmentAnalysisData(
        records=tuple(sorted(records, key=lambda row: (int(row.grade), row.sample_id))),
        source_hashes=initial_hashes,
        criterion_pair_count=len(criterion_pairs),
        criterion_count_distribution_67=dict(criterion_count_distribution_67),
        join_success_count=len(records),
        join_failure_count=0,
    )


def _mean(values: Sequence[float]) -> float:
    return sum(values) / len(values)


def _rank(values: Sequence[float]) -> list[float]:
    order = sorted(range(len(values)), key=lambda index: values[index])
    ranks = [0.0] * len(values)
    start = 0
    while start < len(order):
        end = start + 1
        while end < len(order) and values[order[end]] == values[order[start]]:
            end += 1
        average_rank = (start + 1 + end) / 2.0
        for position in range(start, end):
            ranks[order[position]] = average_rank
        start = end
    return ranks


def _pearson(x_values: Sequence[float], y_values: Sequence[float]) -> float | None:
    if len(x_values) != len(y_values) or len(x_values) < 3:
        return None
    mean_x = _mean(x_values)
    mean_y = _mean(y_values)
    centered_x = [value - mean_x for value in x_values]
    centered_y = [value - mean_y for value in y_values]
    denominator = math.sqrt(
        sum(value * value for value in centered_x) * sum(value * value for value in centered_y)
    )
    if denominator == 0:
        return None
    return max(-1.0, min(1.0, sum(x * y for x, y in zip(centered_x, centered_y)) / denominator))


def _beta_continued_fraction(a: float, b: float, x: float) -> float:
    maximum_iterations = 300
    epsilon = 3e-14
    tiny = 1e-300
    qab = a + b
    qap = a + 1.0
    qam = a - 1.0
    c = 1.0
    d = 1.0 - qab * x / qap
    if abs(d) < tiny:
        d = tiny
    d = 1.0 / d
    result = d
    for iteration in range(1, maximum_iterations + 1):
        twice = 2 * iteration
        coefficient = iteration * (b - iteration) * x / ((qam + twice) * (a + twice))
        d = 1.0 + coefficient * d
        if abs(d) < tiny:
            d = tiny
        c = 1.0 + coefficient / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        result *= d * c
        coefficient = -(a + iteration) * (qab + iteration) * x / (
            (a + twice) * (qap + twice)
        )
        d = 1.0 + coefficient * d
        if abs(d) < tiny:
            d = tiny
        c = 1.0 + coefficient / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        delta = d * c
        result *= delta
        if abs(delta - 1.0) < epsilon:
            return result
    raise ArithmeticError("Incomplete-beta continued fraction did not converge")


def _regularized_beta(x: float, a: float, b: float) -> float:
    if x <= 0:
        return 0.0
    if x >= 1:
        return 1.0
    front = math.exp(
        math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b)
        + a * math.log(x)
        + b * math.log1p(-x)
    )
    if x < (a + 1.0) / (a + b + 2.0):
        return front * _beta_continued_fraction(a, b, x) / a
    return 1.0 - front * _beta_continued_fraction(b, a, 1.0 - x) / b


def _correlation_p_value(correlation: float | None, degrees_of_freedom: int) -> float | None:
    if correlation is None or degrees_of_freedom <= 0:
        return None
    if abs(correlation) >= 1.0:
        return 0.0
    t_squared = correlation * correlation * degrees_of_freedom / (1.0 - correlation * correlation)
    return min(1.0, max(0.0, _regularized_beta(degrees_of_freedom / (degrees_of_freedom + t_squared), degrees_of_freedom / 2.0, 0.5)))


def _regularized_gamma_q(a: float, x: float) -> float:
    if x < 0 or a <= 0:
        raise ValueError("Invalid regularized-gamma arguments")
    if x == 0:
        return 1.0
    epsilon = 3e-14
    tiny = 1e-300
    if x < a + 1.0:
        term = 1.0 / a
        total = term
        ap = a
        for _ in range(1, 10000):
            ap += 1.0
            term *= x / ap
            total += term
            if abs(term) < abs(total) * epsilon:
                p_value = total * math.exp(-x + a * math.log(x) - math.lgamma(a))
                return max(0.0, min(1.0, 1.0 - p_value))
        raise ArithmeticError("Incomplete-gamma series did not converge")
    b = x + 1.0 - a
    c = 1.0 / tiny
    d = 1.0 / b
    result = d
    for iteration in range(1, 10000):
        coefficient = -iteration * (iteration - a)
        b += 2.0
        d = coefficient * d + b
        if abs(d) < tiny:
            d = tiny
        c = b + coefficient / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        delta = d * c
        result *= delta
        if abs(delta - 1.0) < epsilon:
            return max(
                0.0,
                min(1.0, result * math.exp(-x + a * math.log(x) - math.lgamma(a))),
            )
    raise ArithmeticError("Incomplete-gamma continued fraction did not converge")


def _quantile(values: Sequence[float], probability: float) -> float:
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * probability
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (position - lower)


def _distribution(values: Sequence[float]) -> dict[str, float]:
    return {
        "min": min(values),
        "q1": _quantile(values, 0.25),
        "median": statistics.median(values),
        "q3": _quantile(values, 0.75),
        "max": max(values),
        "mean": _mean(values),
    }


def _format_number(value: float) -> str:
    return f"{value:.6f}".rstrip("0").rstrip(".")


def _metric_buckets(records: Sequence[SampleAnalysisRecord], metric: str) -> list[tuple[str, list[SampleAnalysisRecord]]]:
    values = [float(getattr(record, metric)) for record in records]
    unique = sorted(set(values))
    if len(unique) <= 8:
        return [
            (f"{metric} = {_format_number(value)}", [record for record in records if float(getattr(record, metric)) == value])
            for value in unique
        ]
    cut_points = sorted(set(_quantile(values, probability) for probability in (0.25, 0.5, 0.75)))
    groups: list[list[SampleAnalysisRecord]] = [[] for _ in range(len(cut_points) + 1)]
    for record in records:
        value = float(getattr(record, metric))
        bucket_index = sum(value > cut for cut in cut_points)
        groups[bucket_index].append(record)
    output: list[tuple[str, list[SampleAnalysisRecord]]] = []
    for index, group in enumerate(groups):
        if not group:
            continue
        if index == 0:
            label = f"{metric} ≤ {_format_number(cut_points[0])}"
        elif index == len(cut_points):
            label = f"{metric} > {_format_number(cut_points[-1])}"
        else:
            label = (
                f"{_format_number(cut_points[index - 1])} < {metric} "
                f"≤ {_format_number(cut_points[index])}"
            )
        output.append((label, group))
    return output


def _chi_square_for_buckets(buckets: Sequence[tuple[str, Sequence[SampleAnalysisRecord]]]) -> tuple[float | None, float | None, float | None, str]:
    if len(buckets) < 2:
        return None, None, None, "Không đủ hai bucket có dữ liệu."
    rows = [
        [sum(record.official_pass for record in group), sum(record.official_non_pass for record in group)]
        for _, group in buckets
    ]
    column_totals = [sum(row[index] for row in rows) for index in range(2)]
    total = sum(column_totals)
    statistic = 0.0
    minimum_expected = math.inf
    for row in rows:
        row_total = sum(row)
        for index, observed in enumerate(row):
            expected = row_total * column_totals[index] / total
            if expected <= 0:
                return None, None, None, "Outcome hoặc bucket không có biến thiên."
            minimum_expected = min(minimum_expected, expected)
            statistic += (observed - expected) ** 2 / expected
    degrees = len(rows) - 1
    p_value = _regularized_gamma_q(degrees / 2.0, statistic / 2.0)
    cramers_v = math.sqrt(statistic / total)
    warning = ""
    if minimum_expected < 5:
        warning = "Có ô tần suất kỳ vọng dưới 5; p-value chi-square cần được diễn giải thận trọng."
    return statistic, p_value, cramers_v, warning


def _association_interpretation(value: float | None, adjusted: bool, significant: bool) -> str:
    if value is None:
        return "Không thể ước lượng mối liên hệ."
    absolute = abs(value)
    strength = "rất yếu" if absolute < 0.1 else "yếu" if absolute < 0.3 else "vừa" if absolute < 0.5 else "mạnh"
    direction = "dương" if value > 0 else "âm" if value < 0 else "gần 0"
    prefix = "Sau khi kiểm soát strata, quan sát thấy mối liên hệ" if adjusted else "Quan sát thấy mối liên hệ thô"
    evidence = "có ý nghĩa thống kê" if significant else "chưa có ý nghĩa thống kê ở ngưỡng 0,05"
    return f"{prefix} {direction}, mức {strength}; {evidence}. Kết quả không mang ý nghĩa nhân quả."


def _base_row(
    grade: str,
    family: str,
    outcome: str,
    metric: str,
    adjustment: str,
    grouping: str,
    records: Sequence[SampleAnalysisRecord],
) -> dict[str, object]:
    values = [float(getattr(record, metric)) for record in records]
    distribution = _distribution(values)
    pass_count = sum(record.official_pass for record in records)
    return {
        "grade": grade,
        "analysis_family": family,
        "outcome": outcome,
        "fragment_metric": metric,
        "adjustment": adjustment,
        "grouping_or_bucket": grouping,
        "sample_count": len(records),
        "pass_count": pass_count,
        "non_pass_count": len(records) - pass_count,
        "pass_rate": pass_count / len(records),
        "pass_rate_difference_from_lowest_bucket": "",
        "statistic_name": "",
        "statistic_value": "",
        "p_value": "",
        "effect_size_name": "",
        "effect_size_value": "",
        "estimable": False,
        "interpretation": "",
        "warning": "",
        "metric_min": distribution["min"],
        "metric_q1": distribution["q1"],
        "metric_median": distribution["median"],
        "metric_q3": distribution["q3"],
        "metric_max": distribution["max"],
        "metric_mean": distribution["mean"],
        "strata_with_variation": "",
        "strata_total": "",
        "method": "",
    }


def _residual_association(
    records: Sequence[SampleAnalysisRecord],
    metric: str,
    outcome: str,
    stratum_key: Callable[[SampleAnalysisRecord], str],
    *,
    rank_transform: bool,
) -> tuple[float | None, float | None, int, int, int, list[tuple[str, list[SampleAnalysisRecord], str]]]:
    groups: dict[str, list[SampleAnalysisRecord]] = defaultdict(list)
    for record in records:
        groups[stratum_key(record)].append(record)
    informative: list[tuple[str, list[SampleAnalysisRecord]]] = []
    diagnostics: list[tuple[str, list[SampleAnalysisRecord], str]] = []
    for name, group in sorted(groups.items()):
        x_unique = len({float(getattr(record, metric)) for record in group})
        y_unique = len({float(getattr(record, outcome)) for record in group})
        if x_unique < 2 or y_unique < 2:
            missing_parts = []
            if x_unique < 2:
                missing_parts.append("fragment metric")
            if y_unique < 2:
                missing_parts.append("outcome")
            diagnostics.append((name, group, " và ".join(missing_parts)))
        else:
            informative.append((name, group))
    if not informative:
        return None, None, 0, len(groups), 0, diagnostics, 0
    selected = [record for _, group in informative for record in group]
    raw_x = [float(getattr(record, metric)) for record in selected]
    raw_y = [float(getattr(record, outcome)) for record in selected]
    transformed_x = _rank(raw_x) if rank_transform else raw_x
    transformed_y = _rank(raw_y) if rank_transform else raw_y
    positions = {id(record): index for index, record in enumerate(selected)}
    residual_x: list[float] = []
    residual_y: list[float] = []
    for _, group in informative:
        indices = [positions[id(record)] for record in group]
        mean_x = _mean([transformed_x[index] for index in indices])
        mean_y = _mean([transformed_y[index] for index in indices])
        residual_x.extend(transformed_x[index] - mean_x for index in indices)
        residual_y.extend(transformed_y[index] - mean_y for index in indices)
    statistic = _pearson(residual_x, residual_y)
    degrees = len(selected) - len(informative) - 1
    p_value = _correlation_p_value(statistic, degrees)
    return statistic, p_value, len(informative), len(groups), len(selected), diagnostics, sum(record.official_pass for record in selected)


def _analysis_rows_for_scope(
    records: Sequence[SampleAnalysisRecord],
    grade: str,
    *,
    pooled: bool,
) -> list[dict[str, object]]:
    if not records:
        raise ValueError(f"No records for analysis scope {grade}")
    rows: list[dict[str, object]] = []
    adjusted_label = (
        "adjusted_for_grade_and_auditor_group" if pooled else "adjusted_for_auditor_group"
    )
    stratum_key = (
        (lambda record: f"grade={record.grade} | {record.auditor_group}")
        if pooled
        else (lambda record: record.auditor_group)
    )
    for metric in FRAGMENT_METRICS:
        x_values = [float(getattr(record, metric)) for record in records]
        pass_values = [float(record.official_pass) for record in records]
        crude = _pearson(x_values, pass_values)
        crude_p = _correlation_p_value(crude, len(records) - 2)
        row = _base_row(
            grade,
            "fragment_vs_official_pass",
            "official_pass",
            metric,
            "crude",
            "all_samples",
            records,
        )
        row.update(
            statistic_name="point_biserial_r",
            statistic_value=crude if crude is not None else "",
            p_value=crude_p if crude_p is not None else "",
            effect_size_name="r_pb",
            effect_size_value=crude if crude is not None else "",
            estimable=crude is not None,
            interpretation=_association_interpretation(
                crude, False, crude_p is not None and crude_p < 0.05
            ),
            warning=(
                ""
                if crude is not None
                else "Không thể ước lượng mối liên hệ vì không tồn tại nhóm đối chứng hoặc biến không có biến thiên."
            ),
            method="Point-biserial correlation ở cấp sample_id; p-value hai phía theo phân phối t.",
        )
        rows.append(row)

        buckets = _metric_buckets(records, metric)
        lowest_rate = min(
            sum(record.official_pass for record in group) / len(group) for _, group in buckets
        )
        chi_square, chi_p, cramers_v, chi_warning = _chi_square_for_buckets(buckets)
        bucket_test = _base_row(
            grade,
            "fragment_vs_official_pass",
            "official_pass",
            metric,
            "crude",
            "bucket_association_test",
            records,
        )
        bucket_test.update(
            statistic_name="chi_square",
            statistic_value=chi_square if chi_square is not None else "",
            p_value=chi_p if chi_p is not None else "",
            effect_size_name="Cramer's V",
            effect_size_value=cramers_v if cramers_v is not None else "",
            estimable=chi_square is not None,
            interpretation=(
                "So sánh mô tả tỷ lệ pass giữa các bucket; bảng kiểm định không hàm ý quan hệ nhân quả."
                if chi_square is not None
                else "Không thể ước lượng quan hệ giữa các bucket."
            ),
            warning=chi_warning,
            method="Chi-square trên bảng bucket × pass/non-pass; effect size Cramér’s V.",
        )
        rows.append(bucket_test)
        for bucket_label, group in buckets:
            bucket_row = _base_row(
                grade,
                "fragment_vs_official_pass",
                "official_pass",
                metric,
                "crude",
                bucket_label,
                group,
            )
            bucket_pass = sum(record.official_pass for record in group)
            bucket_rate = bucket_pass / len(group)
            bucket_row.update(
                pass_rate_difference_from_lowest_bucket=bucket_rate - lowest_rate,
                statistic_name="bucket_pass_rate",
                statistic_value=bucket_rate,
                estimable=True,
                interpretation="Tỷ lệ pass quan sát trong nhóm giá trị; chỉ dùng để mô tả.",
                method="Nhóm theo giá trị nếu có ≤8 giá trị; nếu nhiều hơn thì dùng ranh giới tứ phân vị.",
            )
            rows.append(bucket_row)

        adjusted, adjusted_p, varying, total_strata, used_count, diagnostics, used_pass_count = _residual_association(
            records,
            metric,
            "official_pass",
            stratum_key,
            rank_transform=False,
        )
        adjusted_row = _base_row(
            grade,
            "fragment_vs_official_pass",
            "official_pass",
            metric,
            adjusted_label,
            "informative_strata_combined",
            records,
        )
        adjusted_warning = (
            f"Loại khỏi ước lượng {total_strata - varying}/{total_strata} strata không có biến thiên đồng thời; "
            f"dùng {used_count}/{len(records)} mẫu."
            if total_strata != varying
            else ""
        )
        if adjusted is None:
            adjusted_warning = (
                "Không thể ước lượng mối liên hệ vì không tồn tại nhóm đối chứng hoặc biến không có biến thiên."
            )
        adjusted_row.update(
            sample_count=used_count,
            pass_count=used_pass_count,
            non_pass_count=used_count - used_pass_count,
            pass_rate=used_pass_count / used_count if used_count else "",
            statistic_name="within_stratum_residual_r",
            statistic_value=adjusted if adjusted is not None else "",
            p_value=adjusted_p if adjusted_p is not None else "",
            effect_size_name="partial_r",
            effect_size_value=adjusted if adjusted is not None else "",
            estimable=adjusted is not None,
            interpretation=_association_interpretation(
                adjusted, True, adjusted_p is not None and adjusted_p < 0.05
            ),
            warning=adjusted_warning,
            strata_with_variation=varying,
            strata_total=total_strata,
            method=(
                "Demean outcome nhị phân và fragment metric trong strata grade × auditor_group."
                if pooled
                else "Demean outcome nhị phân và fragment metric trong từng auditor_group."
            ) + " Tương quan Pearson của residual; p-value hai phía theo phân phối t.",
        )
        rows.append(adjusted_row)
        for name, group, missing_part in diagnostics:
            diagnostic = _base_row(
                grade,
                "fragment_vs_official_pass",
                "official_pass",
                metric,
                adjusted_label,
                f"non_estimable_stratum: {name}",
                group,
            )
            diagnostic.update(
                estimable=False,
                interpretation="Không ước lượng riêng strata này.",
                warning=(
                    "Không thể ước lượng mối liên hệ vì không tồn tại nhóm đối chứng hoặc biến không có biến thiên. "
                    f"Thành phần không biến thiên: {missing_part}."
                ),
                strata_with_variation=0,
                strata_total=1,
                method="Kiểm tra biến thiên trong strata trước khi residualization.",
            )
            rows.append(diagnostic)

    for metric in FRAGMENT_METRICS:
        x_values = [float(getattr(record, metric)) for record in records]
        y_values = [record.checklist_pass_rate for record in records]
        crude = _pearson(_rank(x_values), _rank(y_values))
        crude_p = _correlation_p_value(crude, len(records) - 2)
        row = _base_row(
            grade,
            "fragment_vs_checklist_pass_rate",
            "checklist_pass_rate",
            metric,
            "crude",
            "all_samples",
            records,
        )
        row.update(
            statistic_name="Spearman_rho",
            statistic_value=crude if crude is not None else "",
            p_value=crude_p if crude_p is not None else "",
            effect_size_name="Spearman_rho",
            effect_size_value=crude if crude is not None else "",
            estimable=crude is not None,
            interpretation=_association_interpretation(
                crude, False, crude_p is not None and crude_p < 0.05
            ),
            warning=(
                ""
                if crude is not None
                else "Không thể ước lượng mối liên hệ vì không tồn tại nhóm đối chứng hoặc biến không có biến thiên."
            ),
            method="Spearman rho ở cấp sample_id; p-value hai phía theo phân phối t xấp xỉ.",
        )
        rows.append(row)
        adjusted, adjusted_p, varying, total_strata, used_count, diagnostics, used_pass_count = _residual_association(
            records,
            metric,
            "checklist_pass_rate",
            stratum_key,
            rank_transform=True,
        )
        adjusted_row = _base_row(
            grade,
            "fragment_vs_checklist_pass_rate",
            "checklist_pass_rate",
            metric,
            adjusted_label,
            "informative_strata_combined",
            records,
        )
        adjusted_warning = (
            f"Loại khỏi ước lượng {total_strata - varying}/{total_strata} strata không có biến thiên đồng thời; "
            f"dùng {used_count}/{len(records)} mẫu."
            if total_strata != varying
            else ""
        )
        if adjusted is None:
            adjusted_warning = (
                "Không thể ước lượng mối liên hệ vì không tồn tại nhóm đối chứng hoặc biến không có biến thiên."
            )
        adjusted_row.update(
            sample_count=used_count,
            pass_count=used_pass_count,
            non_pass_count=used_count - used_pass_count,
            pass_rate=used_pass_count / used_count if used_count else "",
            statistic_name="within_stratum_rank_residual_r",
            statistic_value=adjusted if adjusted is not None else "",
            p_value=adjusted_p if adjusted_p is not None else "",
            effect_size_name="partial_rank_r",
            effect_size_value=adjusted if adjusted is not None else "",
            estimable=adjusted is not None,
            interpretation=_association_interpretation(
                adjusted, True, adjusted_p is not None and adjusted_p < 0.05
            ),
            warning=adjusted_warning,
            strata_with_variation=varying,
            strata_total=total_strata,
            method=(
                "Rank residualization trong strata grade × auditor_group."
                if pooled
                else "Rank residualization trong từng auditor_group."
            ) + " Tương quan Pearson của rank residual; p-value hai phía theo phân phối t.",
        )
        rows.append(adjusted_row)
        for name, group, missing_part in diagnostics:
            diagnostic = _base_row(
                grade,
                "fragment_vs_checklist_pass_rate",
                "checklist_pass_rate",
                metric,
                adjusted_label,
                f"non_estimable_stratum: {name}",
                group,
            )
            diagnostic.update(
                estimable=False,
                interpretation="Không ước lượng riêng strata này.",
                warning=(
                    "Không thể ước lượng mối liên hệ vì không tồn tại nhóm đối chứng hoặc biến không có biến thiên. "
                    f"Thành phần không biến thiên: {missing_part}."
                ),
                strata_with_variation=0,
                strata_total=1,
                method="Kiểm tra biến thiên trong strata trước khi rank residualization.",
            )
            rows.append(diagnostic)
    return rows


def _summary_row(
    rows: Sequence[dict[str, object]], family: str, metric: str, adjustment: str
) -> dict[str, object] | None:
    return next(
        (
            row
            for row in rows
            if row["analysis_family"] == family
            and row["fragment_metric"] == metric
            and row["adjustment"] == adjustment
            and row["grouping_or_bucket"] in {"all_samples", "informative_strata_combined"}
        ),
        None,
    )


def build_analysis_rows(data: FragmentAnalysisData) -> tuple[dict[str, list[dict[str, object]]], list[dict[str, object]]]:
    """Create four grade row sets plus a pooled/comparative root row set."""

    grade_rows: dict[str, list[dict[str, object]]] = {}
    for grade in EXPECTED_GRADE_COUNTS:
        records = [record for record in data.records if record.grade == grade]
        grade_rows[grade] = _analysis_rows_for_scope(records, grade, pooled=False)
    pooled_rows = _analysis_rows_for_scope(data.records, "all", pooled=True)
    root_rows = [row.copy() for row in pooled_rows]
    root_rows.extend(row.copy() for grade in EXPECTED_GRADE_COUNTS for row in grade_rows[grade])

    for family, outcome in (
        ("fragment_vs_official_pass", "official_pass"),
        ("fragment_vs_checklist_pass_rate", "checklist_pass_rate"),
    ):
        for metric in FRAGMENT_METRICS:
            grade_stats = [
                _summary_row(grade_rows[grade], family, metric, "crude")
                for grade in EXPECTED_GRADE_COUNTS
            ]
            pooled_crude = _summary_row(pooled_rows, family, metric, "crude")
            pooled_adjusted = _summary_row(
                pooled_rows, family, metric, "adjusted_for_grade_and_auditor_group"
            )
            values = [
                float(row["statistic_value"])
                for row in grade_stats
                if row is not None and isinstance(row["statistic_value"], (int, float))
            ]
            pooled_value = (
                float(pooled_crude["statistic_value"])
                if pooled_crude is not None and isinstance(pooled_crude["statistic_value"], (int, float))
                else None
            )
            adjusted_value = (
                float(pooled_adjusted["statistic_value"])
                if pooled_adjusted is not None and isinstance(pooled_adjusted["statistic_value"], (int, float))
                else None
            )
            signs = {0 if abs(value) < 1e-12 else (1 if value > 0 else -1) for value in values}
            consistent = len(signs - {0}) <= 1
            reversal = (
                pooled_value is not None
                and adjusted_value is not None
                and pooled_value * adjusted_value < 0
            )
            attenuation = (
                pooled_value is not None
                and adjusted_value is not None
                and abs(pooled_value) >= 0.1
                and abs(adjusted_value) < abs(pooled_value) * 0.25
            )
            warning_parts = []
            if not consistent:
                warning_parts.append("Chiều liên hệ không nhất quán giữa bốn lớp.")
            if reversal:
                warning_parts.append("Kết quả pooled đổi chiều sau điều chỉnh; có dấu hiệu confounding/Simpson’s paradox.")
            elif attenuation:
                warning_parts.append("Mối liên hệ pooled giảm mạnh sau điều chỉnh; có dấu hiệu confounding theo khối hoặc auditor/shard.")
            base = _base_row(
                "all",
                "cross_grade_consistency",
                outcome,
                metric,
                "comparison_across_grades",
                "grade_6_vs_7_vs_8_vs_9_and_pooled",
                data.records,
            )
            base.update(
                statistic_name="grade_statistic_range",
                statistic_value=(max(values) - min(values)) if values else "",
                effect_size_name="pooled_adjusted_statistic",
                effect_size_value=adjusted_value if adjusted_value is not None else "",
                estimable=bool(values) and pooled_value is not None and adjusted_value is not None,
                interpretation=(
                    "Chiều và độ mạnh tương đối nhất quán giữa các lớp."
                    if consistent and not reversal and not attenuation
                    else "Kết quả thô và kết quả trong lớp/đã điều chỉnh không hoàn toàn nhất quán."
                ) + " Không diễn giải theo hướng nhân quả.",
                warning=" ".join(warning_parts),
                method="So sánh dấu và biên độ statistic thô theo lớp với statistic pooled trước/sau điều chỉnh.",
            )
            root_rows.append(base)
    return grade_rows, root_rows


def _rows_as_lists(rows: Iterable[dict[str, object]]) -> list[list[object]]:
    return [[row.get(header, "") for header in ANALYSIS_HEADERS] for row in rows]


def _write_analysis_workbook(path: Path, sheet_name: str, rows: Sequence[dict[str, object]]) -> None:
    _write_single_sheet_workbook(
        path,
        sheet_name,
        ANALYSIS_HEADERS,
        _rows_as_lists(rows),
        ANALYSIS_WIDTHS,
        percentage_columns=(10, 11),
    )


def _format_stat(row: dict[str, object] | None) -> str:
    if row is None or not isinstance(row.get("statistic_value"), (int, float)):
        return "không ước lượng được"
    value = float(row["statistic_value"])
    p_value = row.get("p_value")
    p_text = f", p={float(p_value):.4g}" if isinstance(p_value, (int, float)) else ""
    return f"{row['statistic_name']}={value:.3f}{p_text}"


def _main_conclusion(root_rows: Sequence[dict[str, object]]) -> str:
    crude = [
        _summary_row(root_rows, "fragment_vs_checklist_pass_rate", metric, "crude")
        for metric in FRAGMENT_METRICS
    ]
    adjusted = [
        _summary_row(
            root_rows,
            "fragment_vs_checklist_pass_rate",
            metric,
            "adjusted_for_grade_and_auditor_group",
        )
        for metric in FRAGMENT_METRICS
    ]
    crude_values = [float(row["statistic_value"]) for row in crude if row and isinstance(row["statistic_value"], (int, float))]
    adjusted_values = [float(row["statistic_value"]) for row in adjusted if row and isinstance(row["statistic_value"], (int, float))]
    inconsistent = any(
        bool(row.get("warning"))
        for row in root_rows
        if row["analysis_family"] == "cross_grade_consistency"
    )
    attenuated = (
        crude_values
        and adjusted_values
        and _mean([abs(value) for value in adjusted_values])
        < _mean([abs(value) for value in crude_values]) * 0.5
    )
    if inconsistent or attenuated:
        return (
            "Chưa có bằng chứng cho thấy số lượng hoặc độ phủ fragment có mối liên hệ độc lập và ổn định với khả năng pass. "
            "Mối liên hệ thô chủ yếu bị chi phối bởi khác biệt giữa khối và quy trình chấm của auditor/shard."
        )
    return (
        "Có quan sát thấy một số mối liên hệ ở dữ liệu hiện tại, nhưng độ mạnh và tính nhất quán cần được đọc cùng kết quả "
        "đã kiểm soát grade và auditor/shard; phân tích này không cung cấp bằng chứng nhân quả."
    )


def fragment_report_section(data: FragmentAnalysisData, root_rows: Sequence[dict[str, object]]) -> str:
    lines = [
        REPORT_HEADING,
        "",
        "Phân tích đặt câu hỏi liệu mức độ ghi tham chiếu fragment có liên hệ với kết quả chấm ở cấp mẫu hay không. "
        "Bốn metric gồm `fragment_row_count`, `fragment_reference_count`, `unique_fragment_count` và "
        "`fragment_criterion_coverage`. Hai outcome được giữ tách biệt: `official_pass` lấy nguyên trạng thái tổng thể "
        "canonical, còn `checklist_pass_rate` là tỷ lệ criterion `pass` trên các criterion áp dụng.",
        "",
        "Đơn vị quan sát là `sample_id`. Phân tích thô dùng point-biserial cho `official_pass` và Spearman rho cho "
        "`checklist_pass_rate`. Phân tích điều chỉnh dùng demeaning hoặc rank residualization trong strata "
        "`grade × auditor_group`; `auditor_group` là tổ hợp `checked_by` và shard đã chuẩn hóa. Strata không có biến thiên "
        "đồng thời ở outcome và fragment metric bị loại khỏi ước lượng và được ghi cảnh báo riêng.",
        "",
        "### Kết quả pooled trước và sau điều chỉnh",
        "",
        "| Outcome | Fragment metric | Thô | Kiểm soát grade + auditor_group |",
        "|---|---|---:|---:|",
    ]
    for family, outcome in (
        ("fragment_vs_official_pass", "official_pass"),
        ("fragment_vs_checklist_pass_rate", "checklist_pass_rate"),
    ):
        for metric in FRAGMENT_METRICS:
            crude = _summary_row(root_rows, family, metric, "crude")
            adjusted = _summary_row(
                root_rows, family, metric, "adjusted_for_grade_and_auditor_group"
            )
            lines.append(
                f"| `{outcome}` | `{metric}` | {_format_stat(crude)} | {_format_stat(adjusted)} |"
            )
    has_any_values = {record.has_any_fragment for record in data.records}
    lines.extend(
        [
            "",
            "### Ảnh hưởng của khối và quy trình chấm",
            "",
            "Dữ liệu lớp 6–7 gồm 154 mẫu được chấm 18 tiêu chí và 308 mẫu chỉ được chấm 16 tiêu chí. "
            "Khác biệt này trùng với shard/auditor; phân tích không tự điền `RAW-CON-06` hoặc `RAW-CON-07` cho các mẫu không được chấm. "
            "Do cách ghi fragment và số criterion thay đổi theo quy trình, kết quả pooled thô có nguy cơ phản ánh cấu trúc shard hơn là một mối liên hệ độc lập.",
            "",
            (
                "`has_any_fragment` có biến thiên trong toàn bộ dữ liệu nhưng không thuộc tám cặp phân tích đã định trước; "
                "biến này chỉ được dùng như kiểm tra mô tả."
                if len(has_any_values) > 1
                else "Không phân tích `has_any_fragment`: Không thể ước lượng mối liên hệ vì không tồn tại nhóm đối chứng hoặc biến không có biến thiên."
            ),
            "",
            "### Giới hạn và kết luận",
            "",
            "Đây là phân tích quan sát trên output audit đã có. Auditor và shard không được phân công ngẫu nhiên; fragment có thể được ghi theo quy trình khác nhau; "
            "một số strata không có nhóm đối chứng hoặc không có biến thiên. p-value không thay thế đánh giá chuyên môn và không chứng minh quan hệ nhân quả.",
            "",
            _main_conclusion(root_rows),
        ]
    )
    return "\n".join(lines)


def _insert_before_appendix(text: str, section: str) -> str:
    if REPORT_HEADING in text:
        raise FileExistsError("The overview already contains the fragment-analysis section")
    marker = "## Phụ lục —"
    if marker in text:
        return text.replace(marker, section.rstrip() + "\n\n" + marker, 1)
    return text.rstrip() + "\n\n" + section.rstrip() + "\n"


def _append_readme_section(text: str, section: str) -> str:
    if README_HEADING in text:
        raise FileExistsError("README already contains the fragment-analysis section")
    return text.rstrip() + "\n\n" + section.rstrip() + "\n"


def _root_readme_section(row_count: int) -> str:
    return f"""{README_HEADING}

- `05_phan_tich_fragment_va_ket_qua_cham_giua_cac_khoi.xlsx`: {row_count} dòng kết quả tổng hợp, gồm kết quả riêng từng lớp, kết quả pooled và kết quả kiểm soát đồng thời `grade` với `auditor_group`.

Mỗi dòng là một kết quả thống kê ở cấp `sample_id`; file này không coi từng dòng criterion là một quan sát độc lập. Đọc các dòng `crude` cùng các dòng `adjusted_for_grade_and_auditor_group` và cảnh báo strata trước khi kết luận. Các mối liên hệ được mô tả không mang ý nghĩa nhân quả."""


def _grade_readme_section(grade: str, row_count: int) -> str:
    return f"""{README_HEADING}

- `07_phan_tich_fragment_va_ket_qua_cham.xlsx`: {row_count} dòng kết quả, chỉ dùng {EXPECTED_GRADE_COUNTS[grade]} mẫu lớp {grade}.

File 07 phân tích riêng lớp {grade} ở cấp `sample_id`, gồm tám cặp giữa bốn metric fragment với `official_pass` hoặc `checklist_pass_rate`. Các dòng `adjusted_for_auditor_group` kiểm soát khác biệt quy trình chấm trong lớp; strata không có biến thiên được đánh dấu `estimable = false` và không có p-value giả. Kết quả không mang ý nghĩa nhân quả."""


def _workbook_rows(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if workbook.sheetnames != [sheet_name]:
        raise ValueError(f"{path} must have exactly one sheet named {sheet_name}")
    worksheet = workbook[sheet_name]
    headers = tuple(cell.value for cell in worksheet[1])
    if headers != ANALYSIS_HEADERS:
        raise ValueError(f"Unexpected analysis headers in {path}: {headers}")
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    workbook.close()
    return rows


def _assert_rows_equal(
    path: Path,
    sheet_name: str,
    expected_rows: Sequence[dict[str, object]],
) -> None:
    actual = _workbook_rows(path, sheet_name)
    expected = [tuple(row.get(header, "") for header in ANALYSIS_HEADERS) for row in expected_rows]
    if len(actual) != len(expected):
        raise ValueError(f"{path} has {len(actual)} rows; expected {len(expected)}")
    for row_number, (actual_row, expected_row) in enumerate(zip(actual, expected), start=2):
        for column_number, (actual_value, expected_value) in enumerate(zip(actual_row, expected_row), start=1):
            if expected_value == "" and actual_value is None:
                continue
            if isinstance(expected_value, float):
                if not isinstance(actual_value, (int, float)) or not math.isclose(
                    float(actual_value), expected_value, rel_tol=1e-12, abs_tol=1e-12
                ):
                    raise ValueError(f"Numeric mismatch in {path} at {row_number},{column_number}")
            elif actual_value != expected_value:
                raise ValueError(f"Value mismatch in {path} at {row_number},{column_number}")


def validate_fragment_analysis_outputs(
    experiment_dir: Path,
    bundle_dir: Path,
    *,
    expected_source_hashes: dict[Path, str] | None = None,
) -> dict[str, object]:
    data = load_fragment_analysis_data(experiment_dir)
    if expected_source_hashes is not None and data.source_hashes != expected_source_hashes:
        raise ValueError("Canonical fragment-analysis source hashes changed")
    grade_rows, root_rows = build_analysis_rows(data)
    _assert_rows_equal(bundle_dir / ROOT_ANALYSIS_NAME, ROOT_SHEET_NAME, root_rows)
    workbook_rows = {ROOT_ANALYSIS_NAME: len(root_rows)}
    workbook_sheets = {ROOT_ANALYSIS_NAME: ROOT_SHEET_NAME}
    for grade in EXPECTED_GRADE_COUNTS:
        path = bundle_dir / f"lop_{grade}" / GRADE_ANALYSIS_NAME
        _assert_rows_equal(path, GRADE_SHEET_NAME, grade_rows[grade])
        if any(row["grade"] != grade for row in grade_rows[grade]):
            raise ValueError(f"Grade analysis workbook contains another grade: {grade}")
        workbook_rows[f"lop_{grade}/{GRADE_ANALYSIS_NAME}"] = len(grade_rows[grade])
        workbook_sheets[f"lop_{grade}/{GRADE_ANALYSIS_NAME}"] = GRADE_SHEET_NAME
    root_readme = (bundle_dir / "README.md").read_text(encoding="utf-8")
    overview = (bundle_dir / "01_bao_cao_tong_quan.md").read_text(encoding="utf-8")
    if ROOT_ANALYSIS_NAME not in root_readme or REPORT_HEADING not in overview:
        raise ValueError("Root README/report does not document fragment analysis")
    for grade in EXPECTED_GRADE_COUNTS:
        readme = (bundle_dir / f"lop_{grade}/README.md").read_text(encoding="utf-8")
        if GRADE_ANALYSIS_NAME not in readme:
            raise ValueError(f"Grade {grade} README does not document fragment analysis")
    pass_counts = Counter(record.grade for record in data.records if record.official_pass)
    sample_counts = Counter(record.grade for record in data.records)
    return {
        "status": "ok",
        "join_success_count": data.join_success_count,
        "join_failure_count": data.join_failure_count,
        "sample_counts": dict(sample_counts),
        "pass_counts": dict(pass_counts),
        "criterion_pair_count": data.criterion_pair_count,
        "criterion_count_distribution_67": data.criterion_count_distribution_67,
        "source_hashes": {path.as_posix(): digest for path, digest in data.source_hashes.items()},
        "workbook_row_counts": workbook_rows,
        "workbook_sheets": workbook_sheets,
    }


def _render_updated_documents(
    bundle_dir: Path,
    data: FragmentAnalysisData,
    grade_rows: dict[str, list[dict[str, object]]],
    root_rows: list[dict[str, object]],
) -> dict[Path, str]:
    updates = {
        bundle_dir / "README.md": _append_readme_section(
            (bundle_dir / "README.md").read_text(encoding="utf-8"),
            _root_readme_section(len(root_rows)),
        ),
        bundle_dir / "01_bao_cao_tong_quan.md": _insert_before_appendix(
            (bundle_dir / "01_bao_cao_tong_quan.md").read_text(encoding="utf-8"),
            fragment_report_section(data, root_rows),
        ),
    }
    for grade in EXPECTED_GRADE_COUNTS:
        path = bundle_dir / f"lop_{grade}/README.md"
        updates[path] = _append_readme_section(
            path.read_text(encoding="utf-8"),
            _grade_readme_section(grade, len(grade_rows[grade])),
        )
    return updates


def add_fragment_analysis_outputs(
    experiment_dir: Path,
    bundle_dir: Path,
) -> dict[str, object]:
    """Add the five new workbooks and update only the six approved Markdown files."""

    data = load_fragment_analysis_data(experiment_dir)
    grade_rows, root_rows = build_analysis_rows(data)
    targets = [bundle_dir / ROOT_ANALYSIS_NAME] + [
        bundle_dir / f"lop_{grade}" / GRADE_ANALYSIS_NAME for grade in EXPECTED_GRADE_COUNTS
    ]
    existing = [path for path in targets if path.exists()]
    if existing:
        raise FileExistsError(
            "Refusing to overwrite existing fragment-analysis deliverables: "
            + ", ".join(path.as_posix() for path in existing)
        )
    updates = _render_updated_documents(bundle_dir, data, grade_rows, root_rows)
    immutable_existing = [
        path
        for path in bundle_dir.rglob("*")
        if path.is_file() and path not in updates
    ]
    immutable_hashes = {path: _sha256(path) for path in immutable_existing}
    bundle_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix=".fragment-analysis-", dir=bundle_dir) as temporary:
        staging = Path(temporary)
        root_staged = staging / ROOT_ANALYSIS_NAME
        _write_analysis_workbook(root_staged, ROOT_SHEET_NAME, root_rows)
        staged_grade_paths: dict[str, Path] = {}
        for grade in EXPECTED_GRADE_COUNTS:
            staged = staging / f"lop_{grade}_{GRADE_ANALYSIS_NAME}"
            _write_analysis_workbook(staged, GRADE_SHEET_NAME, grade_rows[grade])
            staged_grade_paths[grade] = staged
        root_staged.replace(bundle_dir / ROOT_ANALYSIS_NAME)
        for grade, staged in staged_grade_paths.items():
            staged.replace(bundle_dir / f"lop_{grade}" / GRADE_ANALYSIS_NAME)
    for path, text in updates.items():
        path.write_text(text, encoding="utf-8")
    changed_immutable = [
        path for path, digest in immutable_hashes.items() if _sha256(path) != digest
    ]
    if changed_immutable:
        raise ValueError(
            "Existing immutable deliverables changed unexpectedly: "
            + ", ".join(path.as_posix() for path in changed_immutable)
        )
    return validate_fragment_analysis_outputs(
        experiment_dir,
        bundle_dir,
        expected_source_hashes=data.source_hashes,
    )


def summarized_associations(
    data: FragmentAnalysisData,
) -> dict[str, dict[str, dict[str, object]]]:
    """Return concise crude/adjusted statistics for the implementation report."""

    grade_rows, root_rows = build_analysis_rows(data)
    output: dict[str, dict[str, dict[str, object]]] = {}
    for grade in (*EXPECTED_GRADE_COUNTS, "all"):
        rows = grade_rows[grade] if grade != "all" else root_rows
        adjusted = (
            "adjusted_for_auditor_group"
            if grade != "all"
            else "adjusted_for_grade_and_auditor_group"
        )
        result: dict[str, dict[str, object]] = {}
        for family in ("fragment_vs_official_pass", "fragment_vs_checklist_pass_rate"):
            for metric in FRAGMENT_METRICS:
                key = f"{family}:{metric}"
                result[key] = {
                    "crude": _summary_row(rows, family, metric, "crude"),
                    "adjusted": _summary_row(rows, family, metric, adjusted),
                }
        output[grade] = result
    return output
