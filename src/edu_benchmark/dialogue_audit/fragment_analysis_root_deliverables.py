"""Readable root deliverables for HNMU fragment analysis."""

from __future__ import annotations

import re
import tempfile
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from edu_benchmark.dialogue_audit.fragment_analysis_hnmu import (
    TECHNICAL_DATA_HEADER_ROW,
    TECHNICAL_FIELDS,
    TECHNICAL_HEADERS,
    analysis_key,
    write_technical_appendix_workbook,
)
from edu_benchmark.dialogue_audit.fragment_analysis_hnmu_compact import (
    OUTCOME_LABELS,
    build_hnmu_summary_rows,
)

ROOT_REPORT_NAME = "05_report_fragment_va_ty_le_dat.md"
ROOT_READABLE_SHEET = "01_Ket_qua_de_doc"
ROOT_BUCKET_SHEET = "02_Ty_le_dat_theo_nhom"
ROOT_STATISTICS_SHEET = "03_Ket_qua_thong_ke"
ROOT_NON_ESTIMABLE_SHEET = "04_Nhom_khong_du_dieu_kien"
ROOT_DICTIONARY_SHEET = "05_Tu_dien"
ROOT_RAW_SHEET = "99_Du_lieu_ky_thuat_goc"
FRAGMENT_LABELS = {
    "fragment_row_count": "Số tiêu chí có dẫn fragment",
    "fragment_reference_count": "Tổng lượt dẫn fragment",
    "unique_fragment_count": "Số fragment khác nhau",
    "fragment_criterion_coverage": "Tỷ lệ tiêu chí có dẫn fragment",
}

ROOT_TECHNICAL_SHEETS = (
    ROOT_READABLE_SHEET,
    ROOT_BUCKET_SHEET,
    ROOT_STATISTICS_SHEET,
    ROOT_NON_ESTIMABLE_SHEET,
    ROOT_DICTIONARY_SHEET,
    ROOT_RAW_SHEET,
)

READABLE_HEADERS = (
    "Nội dung về fragment",
    "Kết quả chấm được xem xét",
    "Khi xem tất cả mẫu",
    "Khi so các mẫu cùng khối lớp và nhóm chấm",
    "Kết luận dễ hiểu",
    "Mã đối chiếu",
)
BUCKET_HEADERS = (
    "Cách đo fragment",
    "Nhóm giá trị",
    "Số mẫu",
    "Số đạt",
    "Số không đạt",
    "Tỷ lệ đạt",
)
STATISTICS_HEADERS = (
    "Phạm vi",
    "Cách đo fragment",
    "Kết quả chấm",
    "Cách so sánh",
    "Số mẫu thực sự được dùng",
    "Phép tính thống kê",
    "Kết quả thống kê",
    "Cảnh báo",
    "Mã đối chiếu",
)
NON_ESTIMABLE_HEADERS = (
    "Khối lớp",
    "Nhóm chấm",
    "Cách đo fragment",
    "Kết quả chấm",
    "Số mẫu",
    "Lý do không thể tính",
    "Mã đối chiếu",
)
DICTIONARY_HEADERS = (
    "Tên hiển thị",
    "Ý nghĩa dễ hiểu",
    "Tên kỹ thuật trong dữ liệu",
)


def _focused_row(
    summary_rows: Sequence[dict[str, object]],
) -> dict[str, object]:
    matches = [
        row
        for row in summary_rows
        if row.get("metric") == "fragment_criterion_coverage"
        and row.get("outcome") == "official_pass"
    ]
    if len(matches) != 1 or matches[0].get("analysis_key") != "FRG-OP-04":
        raise ValueError(
            "HNMU report requires official_pass × fragment_criterion_coverage"
        )
    return matches[0]


def _result_is_clear(row: dict[str, object], prefix: str) -> bool:
    return bool(
        row.get(f"{prefix}_evidence")
        and isinstance(row.get(f"{prefix}_statistic"), (int, float))
        and abs(float(row[f"{prefix}_statistic"])) > 0.05
    )


def _trend_label(
    row: dict[str, object],
    prefix: str,
) -> str:
    if prefix == "adjusted" and not row.get("adjusted_estimable"):
        return "Không đủ dữ liệu để kết luận"
    if not _result_is_clear(row, prefix):
        return "Không thấy khác biệt rõ ràng"
    value = float(row[f"{prefix}_statistic"])
    return "Có xu hướng cao hơn" if value > 0 else "Có xu hướng thấp hơn"


def _pair_conclusion(row: dict[str, object]) -> str:
    if not row.get("adjusted_estimable"):
        return "Không đủ dữ liệu phù hợp để kết luận."
    crude_clear = _result_is_clear(row, "crude")
    adjusted_clear = _result_is_clear(row, "adjusted")
    if not crude_clear and not adjusted_clear:
        return "Chưa thấy mối liên hệ rõ ràng."
    if crude_clear and not adjusted_clear:
        if row.get("analysis_key") == "FRG-OP-04":
            return (
                "Chưa thể khẳng định fragment đầy đủ hơn đi kèm tỷ lệ đạt cao hơn; "
                "khác biệt không rõ trong cùng khối và nhóm chấm."
            )
        return (
            "Sự khác biệt không còn rõ ràng khi so sánh các mẫu cùng khối lớp "
            "và nhóm chấm."
        )
    if not crude_clear and adjusted_clear:
        return (
            "Xu hướng chỉ xuất hiện trong phép so sánh theo nhóm và cần được "
            "diễn giải thận trọng."
        )
    crude = float(row["crude_statistic"])
    adjusted = float(row["adjusted_statistic"])
    if crude * adjusted < 0:
        return (
            "Hai cách so sánh cho kết quả trái chiều nên chưa thể đưa ra "
            "kết luận ổn định."
        )
    return (
        "Xu hướng vẫn xuất hiện khi so sánh các mẫu cùng khối lớp và nhóm chấm, "
        "nhưng không chứng minh quan hệ nguyên nhân."
    )


def root_report_content(
    summary_rows: Sequence[dict[str, object]],
) -> dict[str, str]:
    row = _focused_row(summary_rows)
    total_count = int(row["crude_sample_count"])
    comparable_count = int(row["adjusted_sample_count"])
    total_label = f"{total_count:,}".replace(",", ".")
    comparable_label = f"{comparable_count:,}".replace(",", ".")
    crude_clear_positive = bool(
        _result_is_clear(row, "crude") and float(row["crude_statistic"]) > 0
    )
    adjusted_clear_positive = bool(
        row.get("adjusted_estimable")
        and _result_is_clear(row, "adjusted")
        and float(row["adjusted_statistic"]) > 0
    )
    if adjusted_clear_positive:
        short_answer = "**Có dấu hiệu cho thấy có.**"
        explanation = (
            "Khi xem toàn bộ dữ liệu, các mẫu có “Tỷ lệ tiêu chí có dẫn fragment” "
            "cao hơn có xu hướng đạt nhiều hơn. Trong những nhóm có đủ dữ liệu "
            "để so sánh giữa các mẫu cùng khối lớp và cùng nhóm chấm, xu hướng "
            "này vẫn được ghi nhận."
        )
        conclusion = (
            "Có dấu hiệu cho thấy các mẫu có “Tỷ lệ tiêu chí có dẫn fragment” cao "
            "hơn cũng có tỷ lệ đạt chính thức cao hơn. Kết quả chỉ cho biết hai "
            "yếu tố đi cùng nhau, không cho biết yếu tố nào tạo ra yếu tố nào."
        )
    elif not row.get("adjusted_estimable"):
        short_answer = "**Chưa đủ dữ liệu để trả lời.**"
        explanation = (
            "Khi xem toàn bộ dữ liệu, có thể quan sát một xu hướng giữa “Tỷ lệ "
            "tiêu chí có dẫn fragment” và tỷ lệ đạt. Tuy nhiên, các nhóm có thể "
            "so sánh trực tiếp chưa đủ dữ liệu để trả lời câu hỏi."
        )
        conclusion = (
            "Chưa đủ dữ liệu phù hợp để khẳng định các mẫu có “Tỷ lệ tiêu chí có "
            "dẫn fragment” cao hơn cũng có tỷ lệ đạt chính thức cao hơn."
        )
    else:
        short_answer = "**Chưa thể khẳng định.**"
        first_sentence = (
            "Khi xem toàn bộ dữ liệu, các mẫu có “Tỷ lệ tiêu chí có dẫn fragment” "
            "cao hơn có xu hướng đạt nhiều hơn."
            if crude_clear_positive
            else (
                "Khi xem toàn bộ dữ liệu, chưa thấy rõ các mẫu có “Tỷ lệ tiêu chí "
                "có dẫn fragment” cao hơn đạt nhiều hơn."
            )
        )
        explanation = (
            first_sentence
            + " Tuy nhiên, trong những nhóm có đủ dữ liệu để so sánh giữa các "
            "mẫu cùng khối lớp và cùng nhóm chấm, xu hướng này không còn rõ ràng."
        )
        conclusion = (
            "Chưa thể khẳng định các mẫu có “Tỷ lệ tiêu chí có dẫn fragment” cao "
            "hơn cũng có tỷ lệ đạt chính thức cao hơn. Kết quả quan sát trên toàn "
            "bộ dữ liệu có thể chịu ảnh hưởng từ sự khác nhau giữa các khối lớp "
            "và nhóm chấm."
        )
    data_limit = (
        "Phép so sánh trong cùng khối lớp và nhóm chấm chỉ sử dụng được "
        f"{comparable_label} trong tổng số {total_label} mẫu, vì nhiều nhóm không "
        "có đủ sự khác biệt để thực hiện phép tính. Do đó, kết quả cần được diễn "
        "giải thận trọng."
    )
    return {
        "short_answer": short_answer,
        "explanation": explanation,
        "conclusion": conclusion,
        "data_limit": data_limit,
        "sheet_conclusion": _pair_conclusion(row),
    }


def root_report_markdown(
    summary_rows: Sequence[dict[str, object]],
) -> str:
    content = root_report_content(summary_rows)
    return "\n".join(
        [
            "# Kết quả phân tích fragment và tỷ lệ đạt",
            "",
            "## Câu hỏi cần trả lời",
            "",
            (
                "Các mẫu có “Tỷ lệ tiêu chí có dẫn fragment” cao hơn có tỷ lệ "
                "đạt chính thức cao hơn không?"
            ),
            "",
            "## Trả lời ngắn",
            "",
            content["short_answer"],
            "",
            "## Kết quả được hiểu như thế nào?",
            "",
            content["explanation"],
            "",
            content["data_limit"],
            "",
            "## Kết luận",
            "",
            content["conclusion"],
            "",
            "## Lưu ý khi diễn giải",
            "",
            (
                "- “Tỷ lệ tiêu chí có dẫn fragment” là phần trăm tiêu chí trong "
                "mẫu có ghi dẫn fragment."
            ),
            (
                "- Phân tích chỉ cho biết hai yếu tố có đi cùng nhau hay không, "
                "không chứng minh fragment đầy đủ hơn là nguyên nhân làm tăng tỷ lệ đạt."
            ),
            (
                "- Các số liệu và phương pháp đầy đủ được trình bày trong bảng "
                "chi tiết kỹ thuật."
            ),
            "",
        ]
    )


def write_root_report(
    path: Path,
    summary_rows: Sequence[dict[str, object]],
) -> None:
    path.write_text(root_report_markdown(summary_rows), encoding="utf-8")


def _readable_rows(
    summary_rows: Sequence[dict[str, object]],
) -> list[list[object]]:
    rows: list[list[object]] = []
    for row in summary_rows:
        conclusion = _pair_conclusion(row)
        if len(conclusion.split()) > 25:
            raise ValueError(
                f"Readable conclusion exceeds 25 words: {row['analysis_key']}"
            )
        rows.append(
            [
                FRAGMENT_LABELS[str(row["metric"])],
                row["outcome_label"],
                _trend_label(row, "crude"),
                _trend_label(row, "adjusted"),
                conclusion,
                row["analysis_key"],
            ]
        )
    return rows


def _scope_label(value: object) -> str:
    return "Toàn bộ lớp 6–9" if str(value) == "all" else f"Lớp {value}"


def _comparison_label(value: object) -> str:
    labels = {
        "crude": "Xem tất cả mẫu",
        "adjusted_for_grade_and_auditor_group": (
            "So trong cùng khối lớp và nhóm chấm"
        ),
        "adjusted_for_auditor_group": "So trong cùng nhóm chấm",
        "comparison_across_grades": "So sánh kết quả giữa các khối",
    }
    return labels.get(str(value), str(value))


def _bucket_label(metric: str, grouping: str) -> str:
    expression = grouping.split("=", 1)[-1].strip()
    if metric == "fragment_criterion_coverage":
        if "≤" in grouping or ">" in grouping:
            return grouping.replace(metric, "Tỷ lệ tiêu chí có dẫn fragment")
        value = float(expression)
        if abs(value) < 1e-12:
            return "Không có tiêu chí nào được dẫn fragment"
        return f"{value:.0%} tiêu chí có dẫn fragment"
    if metric == "fragment_row_count":
        if expression == "0":
            return "Không có tiêu chí nào được dẫn fragment"
        return f"{expression} tiêu chí có dẫn fragment"
    if metric == "fragment_reference_count":
        if grouping.startswith("5 <") and "≤ 7" in grouping:
            return "Từ 6 đến 7 lượt dẫn fragment"
        if "≤" in grouping:
            match = re.search(r"≤\s*([0-9.]+)", grouping)
            return f"Không quá {match.group(1)} lượt dẫn fragment" if match else grouping
        if ">" in grouping:
            match = re.search(r">\s*([0-9.]+)", grouping)
            return f"Trên {match.group(1)} lượt dẫn fragment" if match else grouping
        if expression == "0":
            return "Không có lượt dẫn fragment"
        return f"{expression} lượt dẫn fragment"
    if metric == "unique_fragment_count":
        if expression == "0":
            return "Không có fragment"
        if expression == "1":
            return "1 fragment"
        return f"{expression} fragment khác nhau"
    return grouping


def _bucket_rows(
    rows: Sequence[dict[str, object]],
) -> list[list[object]]:
    output: list[list[object]] = []
    for row in rows:
        if not (
            str(row.get("grade")) == "all"
            and row.get("outcome") == "official_pass"
            and row.get("statistic_name") == "bucket_pass_rate"
        ):
            continue
        metric = str(row["fragment_metric"])
        output.append(
            [
                FRAGMENT_LABELS[metric],
                _bucket_label(metric, str(row["grouping_or_bucket"])),
                row["sample_count"],
                row["pass_count"],
                row["non_pass_count"],
                row["pass_rate"],
            ]
        )
    return output


def _statistic_method(row: dict[str, object]) -> str:
    name = str(row.get("statistic_name") or "")
    labels = {
        "point_biserial_r": "Tương quan điểm–nhị phân ở cấp mẫu",
        "Spearman_rho": "Tương quan thứ hạng Spearman ở cấp mẫu",
        "within_stratum_residual_r": (
            "Tương quan phần dư trong các nhóm khối lớp và nhóm chấm"
        ),
        "within_stratum_rank_residual_r": (
            "Tương quan thứ hạng trong các nhóm khối lớp và nhóm chấm"
        ),
        "chi_square": "Kiểm định chi-square giữa các nhóm giá trị",
        "grade_statistic_range": "So sánh độ nhất quán giữa các khối",
    }
    return labels.get(name, "Không thể tính" if not name else name)


def _effect_label(value: object) -> str:
    labels = {
        "r_pb": "Hệ số điểm–nhị phân",
        "Spearman_rho": "Hệ số Spearman",
        "partial_r": "Hệ số khi so trong nhóm",
        "partial_rank_r": "Hệ số thứ hạng khi so trong nhóm",
        "Cramer's V": "Cramér’s V",
        "pooled_adjusted_statistic": "Kết quả gộp theo nhóm",
    }
    return labels.get(str(value), str(value))


def _statistic_result(row: dict[str, object]) -> str:
    if not bool(row.get("estimable")):
        return "Không thể tính"
    parts: list[str] = []
    value = row.get("statistic_value")
    if isinstance(value, (int, float)):
        parts.append(f"Kết quả chính = {float(value):.4f}")
    p_value = row.get("p_value")
    if isinstance(p_value, (int, float)):
        parts.append(f"p-value = {float(p_value):.4g}")
    effect = row.get("effect_size_value")
    if isinstance(effect, (int, float)):
        parts.append(
            f"{_effect_label(row.get('effect_size_name'))} = {float(effect):.4f}"
        )
    return "; ".join(parts) if parts else "Không có giá trị số"


def _friendly_warning(row: dict[str, object]) -> str:
    if not bool(row.get("estimable")):
        return (
            "Không có đủ biến thiên đồng thời về kết quả chấm và cách đo fragment."
        )
    warning = str(row.get("warning") or "")
    if not warning:
        return ""
    if "tần suất kỳ vọng dưới 5" in warning:
        return (
            "Một số nhóm có quá ít mẫu; kết quả kiểm định cần được diễn giải "
            "thận trọng."
        )
    if "Loại khỏi ước lượng" in warning:
        numbers = re.findall(r"\d+/\d+", warning)
        samples = re.findall(r"dùng\s+(\d+/\d+)\s+mẫu", warning)
        group_text = numbers[0] if numbers else "một số"
        sample_text = samples[0] if samples else ""
        suffix = f"; phép tính dùng {sample_text} mẫu" if sample_text else ""
        return f"{group_text} nhóm không đủ biến thiên{suffix}."
    if "confounding" in warning or "Simpson" in warning or "pooled" in warning:
        return (
            "Kết quả gộp thay đổi khi so theo nhóm; khác biệt giữa khối lớp hoặc "
            "nhóm chấm có thể ảnh hưởng."
        )
    return warning.replace("strata", "nhóm").replace("pooled", "gộp")


def _statistics_rows(
    rows: Sequence[dict[str, object]],
) -> list[list[object]]:
    main_groups = {
        "all_samples",
        "informative_strata_combined",
        "bucket_association_test",
        "grade_6_vs_7_vs_8_vs_9_and_pooled",
    }
    output: list[list[object]] = []
    for row in rows:
        if str(row.get("grouping_or_bucket")) not in main_groups:
            continue
        output.append(
            [
                _scope_label(row["grade"]),
                FRAGMENT_LABELS[str(row["fragment_metric"])],
                OUTCOME_LABELS[str(row["outcome"])],
                _comparison_label(row["adjustment"]),
                row["sample_count"],
                _statistic_method(row),
                _statistic_result(row),
                _friendly_warning(row),
                analysis_key(row["outcome"], row["fragment_metric"]),
            ]
        )
    return output


def _parse_non_estimable_group(
    row: dict[str, object],
) -> tuple[str, str]:
    grouping = str(row["grouping_or_bucket"])
    prefix = "non_estimable_stratum:"
    remainder = grouping[len(prefix) :].strip()
    match = re.match(r"grade=([6789])\s*\|\s*(.*)", remainder)
    if match:
        return f"Lớp {match.group(1)}", match.group(2)
    return _scope_label(row["grade"]), remainder


def _non_estimable_rows(
    rows: Sequence[dict[str, object]],
) -> list[list[object]]:
    parsed: list[tuple[dict[str, object], str, str]] = []
    for row in rows:
        if not str(row.get("grouping_or_bucket")).startswith(
            "non_estimable_stratum:"
        ):
            continue
        grade, technical_group = _parse_non_estimable_group(row)
        parsed.append((row, grade, technical_group))
    group_labels: dict[str, str] = {}
    for _, _, technical_group in parsed:
        if technical_group not in group_labels:
            group_labels[technical_group] = (
                f"Nhóm chấm {len(group_labels) + 1:02d}"
            )
    output: list[list[object]] = []
    for row, grade, technical_group in parsed:
        output.append(
            [
                grade,
                group_labels[technical_group],
                FRAGMENT_LABELS[str(row["fragment_metric"])],
                OUTCOME_LABELS[str(row["outcome"])],
                row["sample_count"],
                (
                    "Nhóm này không có đủ biến thiên đồng thời về kết quả chấm "
                    "và cách đo fragment."
                ),
                analysis_key(row["outcome"], row["fragment_metric"]),
            ]
        )
    return output


_FIELD_DICTIONARY = {
    "grade": ("Phạm vi", "Khối lớp hoặc phạm vi dữ liệu được phân tích."),
    "analysis_family": (
        "Nhóm câu hỏi phân tích",
        "Cho biết dòng thuộc nhóm phân tích nào.",
    ),
    "outcome": ("Kết quả chấm", "Kết quả được đặt cạnh cách đo fragment."),
    "fragment_metric": (
        "Cách đo fragment",
        "Cách biểu diễn số lượng hoặc độ phủ fragment.",
    ),
    "adjustment": ("Cách so sánh", "Phạm vi nhóm được dùng khi so sánh."),
    "grouping_or_bucket": (
        "Nhóm giá trị",
        "Mức fragment hoặc nhóm dữ liệu của dòng kết quả.",
    ),
    "sample_count": ("Số mẫu", "Số mẫu thực sự có trong dòng phân tích."),
    "pass_count": ("Số đạt", "Số mẫu có trạng thái đạt chính thức."),
    "non_pass_count": (
        "Số không đạt",
        "Số mẫu không có trạng thái đạt chính thức.",
    ),
    "pass_rate": ("Tỷ lệ đạt", "Số đạt chia cho tổng số mẫu."),
    "pass_rate_difference_from_lowest_bucket": (
        "Chênh lệch tỷ lệ đạt",
        "Chênh lệch so với nhóm fragment thấp nhất.",
    ),
    "statistic_name": ("Phép tính", "Tên phép tính được sử dụng."),
    "statistic_value": ("Kết quả phép tính", "Giá trị chính của phép tính."),
    "p_value": ("p-value", "Giá trị phục vụ kiểm tra kỹ thuật."),
    "effect_size_name": (
        "Tên thước đo độ lớn",
        "Tên thước đo mức độ khác biệt hoặc liên hệ.",
    ),
    "effect_size_value": (
        "Giá trị độ lớn",
        "Giá trị của thước đo mức độ khác biệt hoặc liên hệ.",
    ),
    "estimable": (
        "Có đủ dữ liệu để tính",
        "Cho biết dòng có đủ biến thiên để thực hiện phép tính hay không.",
    ),
    "interpretation": (
        "Diễn giải kỹ thuật",
        "Diễn giải đầy đủ được sinh cho dòng kỹ thuật.",
    ),
    "warning": ("Cảnh báo", "Giới hạn cần lưu ý khi đọc kết quả."),
    "metric_min": ("Nhỏ nhất", "Giá trị nhỏ nhất của cách đo fragment."),
    "metric_q1": ("Tứ phân vị 25%", "Mốc mà 25% mẫu không vượt quá."),
    "metric_median": ("Trung vị", "Giá trị nằm giữa phân bố."),
    "metric_q3": ("Tứ phân vị 75%", "Mốc mà 75% mẫu không vượt quá."),
    "metric_max": ("Lớn nhất", "Giá trị lớn nhất của cách đo fragment."),
    "metric_mean": ("Trung bình", "Giá trị trung bình của cách đo fragment."),
    "strata_with_variation": (
        "Số nhóm đủ biến thiên",
        "Số nhóm có thể đóng góp vào phép tính theo nhóm.",
    ),
    "strata_total": (
        "Tổng số nhóm",
        "Tổng số nhóm khối lớp và nhóm chấm được xem xét.",
    ),
    "method": ("Phương pháp", "Mô tả chi tiết cách thực hiện phép tính."),
    "analysis_key": (
        "Mã đối chiếu",
        "Mã nối kết quả dễ đọc với dữ liệu kỹ thuật gốc.",
    ),
}


def _dictionary_rows() -> list[list[str]]:
    concepts = [
        (
            "Số tiêu chí có dẫn fragment",
            "Số tiêu chí trong mẫu có ít nhất một dẫn chứng học liệu.",
            "fragment_row_count",
        ),
        (
            "Tổng lượt dẫn fragment",
            "Tổng số lượt fragment được nhắc tới trong mẫu.",
            "fragment_reference_count",
        ),
        (
            "Số fragment khác nhau",
            "Số mã fragment không trùng nhau trong mẫu.",
            "unique_fragment_count",
        ),
        (
            "Tỷ lệ tiêu chí có dẫn fragment",
            "Số tiêu chí có dẫn fragment chia cho số tiêu chí áp dụng.",
            "fragment_criterion_coverage",
        ),
        (
            "Trạng thái đạt chính thức",
            "Mẫu có trạng thái tổng thể chính thức là đạt.",
            "official_pass",
        ),
        (
            "Tỷ lệ tiêu chí đạt",
            "Số tiêu chí đạt chia cho số tiêu chí áp dụng của mẫu.",
            "checklist_pass_rate",
        ),
        (
            "Nhóm chấm 01, 02, ...",
            "Nhãn rút gọn của nhóm chấm; mã đầy đủ được giữ trong dữ liệu kỹ thuật gốc.",
            "auditor_group",
        ),
    ]
    fields = [
        [display, meaning, field]
        for field in TECHNICAL_FIELDS
        for display, meaning in [_FIELD_DICTIONARY[field]]
    ]
    return [list(row) for row in concepts] + fields


def _write_table_sheet(
    worksheet,
    *,
    title: str,
    note: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[object]],
    widths: Sequence[int],
    table_name: str,
    percentage_columns: Sequence[int] = (),
    zoom: int = 90,
) -> None:
    last_column = len(headers)
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    worksheet.cell(1, 1, title)
    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=last_column,
    )
    worksheet.cell(2, 1, note)
    for column, header in enumerate(headers, start=1):
        worksheet.cell(4, column, header)
    for row_number, row in enumerate(rows, start=5):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(row_number, column, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column in percentage_columns and isinstance(value, (int, float)):
                cell.number_format = "0.00%"
    navy = "1F4E78"
    pale_blue = "DDEBF7"
    worksheet["A1"].fill = PatternFill("solid", fgColor=navy)
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    worksheet["A2"].fill = PatternFill("solid", fgColor=pale_blue)
    worksheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    for cell in worksheet[4]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.row_dimensions[1].height = 34
    worksheet.row_dimensions[2].height = 38
    worksheet.row_dimensions[4].height = 36
    if rows:
        table = Table(
            displayName=table_name,
            ref=f"A4:{get_column_letter(last_column)}{4 + len(rows)}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = zoom
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0


def write_readable_root_technical_workbook(
    path: Path,
    rows: Sequence[dict[str, object]],
) -> None:
    summary_rows = build_hnmu_summary_rows(rows, pooled=True)
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        dir=path.parent,
        delete=False,
    ) as handle:
        temporary_path = Path(handle.name)
    try:
        write_technical_appendix_workbook(
            temporary_path,
            "phu_luc_ky_thuat_fragment",
            rows,
            pooled=True,
        )
        workbook = load_workbook(temporary_path)
        raw = workbook["phu_luc_ky_thuat_fragment"]
        raw.title = ROOT_RAW_SHEET
        raw.sheet_properties.tabColor = "808080"
        for index, name in enumerate(ROOT_TECHNICAL_SHEETS[:-1]):
            workbook.create_sheet(name, index=index)

        _write_table_sheet(
            workbook[ROOT_READABLE_SHEET],
            title="KẾT QUẢ PHÂN TÍCH FRAGMENT DỄ ĐỌC",
            note=(
                "Mỗi dòng là một câu hỏi phân tích. Mã đối chiếu dùng để tìm "
                "dòng liên quan trong dữ liệu kỹ thuật gốc."
            ),
            headers=READABLE_HEADERS,
            rows=_readable_rows(summary_rows),
            widths=(25, 23, 23, 32, 45, 15),
            table_name="KetQuaDeDoc",
            zoom=85,
        )
        _write_table_sheet(
            workbook[ROOT_BUCKET_SHEET],
            title="TỶ LỆ ĐẠT THEO TỪNG MỨC FRAGMENT",
            note=(
                "Bảng mô tả tỷ lệ đạt chính thức trong toàn bộ lớp 6–9 theo "
                "từng nhóm giá trị fragment."
            ),
            headers=BUCKET_HEADERS,
            rows=_bucket_rows(rows),
            widths=(29, 35, 12, 12, 15, 14),
            table_name="TyLeDatTheoNhom",
            percentage_columns=(6,),
        )
        _write_table_sheet(
            workbook[ROOT_STATISTICS_SHEET],
            title="KẾT QUẢ THỐNG KÊ PHỤC VỤ KIỂM TRA",
            note=(
                "Sheet này giữ các phép tính chính. Đọc cột cảnh báo cùng kết "
                "quả; kết luận dễ hiểu nằm ở sheet 01."
            ),
            headers=STATISTICS_HEADERS,
            rows=_statistics_rows(rows),
            widths=(18, 28, 24, 31, 16, 35, 42, 45, 15),
            table_name="KetQuaThongKe",
            zoom=80,
        )
        _write_table_sheet(
            workbook[ROOT_NON_ESTIMABLE_SHEET],
            title="CÁC NHÓM KHÔNG ĐỦ ĐIỀU KIỆN TÍNH",
            note=(
                "Các dòng dưới đây được tách riêng để không lẫn với kết quả có "
                "thể tính."
            ),
            headers=NON_ESTIMABLE_HEADERS,
            rows=_non_estimable_rows(rows),
            widths=(14, 58, 28, 24, 12, 42, 15),
            table_name="NhomKhongDuDieuKien",
            zoom=80,
        )
        _write_table_sheet(
            workbook[ROOT_DICTIONARY_SHEET],
            title="TỪ ĐIỂN KHÁI NIỆM VÀ TRƯỜNG DỮ LIỆU",
            note=(
                "Tên tiếng Việt được dùng làm nhãn chính. Tên kỹ thuật chỉ nằm "
                "ở cột cuối để truy vết."
            ),
            headers=DICTIONARY_HEADERS,
            rows=_dictionary_rows(),
            widths=(34, 70, 36),
            table_name="TuDienFragment",
            zoom=90,
        )
        workbook.active = 0
        workbook.save(path)
        workbook.close()
    finally:
        temporary_path.unlink(missing_ok=True)


def read_root_raw_technical_rows(
    path: Path,
) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, data_only=True)
    if tuple(workbook.sheetnames) != ROOT_TECHNICAL_SHEETS:
        raise ValueError(f"Unexpected root technical sheet order: {workbook.sheetnames}")
    if workbook.active.title != ROOT_READABLE_SHEET:
        raise ValueError("Readable sheet must open first")
    raw = workbook[ROOT_RAW_SHEET]
    headers = tuple(
        raw.cell(TECHNICAL_DATA_HEADER_ROW, column).value
        for column in range(1, len(TECHNICAL_HEADERS) + 1)
    )
    if headers != tuple(TECHNICAL_HEADERS):
        raise ValueError("Raw technical headers changed")
    rows = list(
        raw.iter_rows(
            min_row=TECHNICAL_DATA_HEADER_ROW + 1,
            max_col=len(TECHNICAL_HEADERS),
            values_only=True,
        )
    )
    workbook.close()
    return rows


def expected_root_readable_rows(
    rows: Sequence[dict[str, object]],
) -> list[tuple[object, ...]]:
    summary_rows = build_hnmu_summary_rows(rows, pooled=True)
    return [tuple(row) for row in _readable_rows(summary_rows)]


def expected_root_bucket_rows(
    rows: Sequence[dict[str, object]],
) -> list[tuple[object, ...]]:
    return [tuple(row) for row in _bucket_rows(rows)]


def expected_root_statistics_rows(
    rows: Sequence[dict[str, object]],
) -> list[tuple[object, ...]]:
    return [tuple(row) for row in _statistics_rows(rows)]


def expected_root_non_estimable_rows(
    rows: Sequence[dict[str, object]],
) -> list[tuple[object, ...]]:
    return [tuple(row) for row in _non_estimable_rows(rows)]


def expected_root_dictionary_rows() -> list[tuple[object, ...]]:
    return [tuple(row) for row in _dictionary_rows()]
