import csv
import math
from pathlib import Path

import pytest
from openpyxl import load_workbook

from edu_benchmark.dialogue_audit.teacher_bundle import load_canonical_bundle_data
from edu_benchmark.dialogue_audit.teacher_bundle_v2 import COVERAGE_HEADERS, STATUS_HEADERS
from edu_benchmark.dialogue_audit.teacher_bundle_v2_partitioned import (
    GRADE_OUTPUT_NAMES,
    ROOT_OUTPUT_NAMES,
    build_partitioned_phase1_teacher_bundle_v2 as build_phase1_teacher_bundle_v2,
    validate_partitioned_phase1_teacher_bundle_v2 as validate_phase1_teacher_bundle_v2,
)

EXPERIMENT_DIR = Path("experiments/20260709_155523")
EXPECTED_STATUS_COUNTS = {
    "6": {"pass": 106, "need_human_review": 131, "failed": 1},
    "7": {"pass": 132, "need_human_review": 91, "failed": 1},
    "8": {"pass": 209, "need_human_review": 70, "failed": 1},
    "9": {"pass": 218, "need_human_review": 90, "failed": 0},
    "all": {"pass": 665, "need_human_review": 382, "failed": 3},
}


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def test_v2_join_inputs_preserve_supplementary_canonical_rows():
    data = load_canonical_bundle_data(EXPERIMENT_DIR)

    assert len(data.normalized_by_id) == 1050
    assert len(data.missing_rows) == 22
    assert len(data.duplicate_rows) == 1
    assert len(data.read_paths) == 15
    assert "Báo cáo rà soát bước đầu dữ liệu hội thoại" in data.teacher_report_text


def test_build_and_validate_v2_bundle_by_deliverable_type(tmp_path, monkeypatch):
    original_path_open = Path.open

    def guarded_path_open(path, *args, **kwargs):
        if "/shared/" in path.resolve().as_posix():
            raise AssertionError(f"V2 builder dereferenced source_file: {path}")
        return original_path_open(path, *args, **kwargs)

    monkeypatch.setattr(Path, "open", guarded_path_open)
    bundle_dir = tmp_path / "hnmu_dialogue_audit_phase1_v2"
    summary = build_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle_dir)
    validation = validate_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle_dir)

    assert {path.name for path in bundle_dir.iterdir()} == set(ROOT_OUTPUT_NAMES) | {
        "lop_6", "lop_7", "lop_8", "lop_9"
    }
    assert summary.status_counts == EXPECTED_STATUS_COUNTS
    assert validation["status_counts"] == EXPECTED_STATUS_COUNTS
    assert validation["grade_counts"] == {"6": 238, "7": 224, "8": 280, "9": 308}
    assert validation["checklist_row_count"] == 18900
    assert validation["missing_row_count"] == 22
    assert validation["duplicate_row_count"] == 1
    assert validation["source_count"] == 15

    expected_sheets = {
        "02_checklist_tieu_chi.xlsx": "checklist_tieu_chi",
        "03_thong_ke_pass_reject_giua_cac_khoi.xlsx": "pass_reject_giua_cac_khoi",
        "04_thong_ke_do_phu_mau_pass_giua_cac_khoi.xlsx": "do_phu_pass_giua_cac_khoi",
    }
    for name, sheet in expected_sheets.items():
        workbook = load_workbook(bundle_dir / name, read_only=True, data_only=True)
        assert workbook.sheetnames == [sheet]
        workbook.close()

    all_ids = set()
    all_pairs = set()
    expected_counts = {"6": 238, "7": 224, "8": 280, "9": 308}
    for grade, expected_count in expected_counts.items():
        grade_dir = bundle_dir / f"lop_{grade}"
        assert {path.name for path in grade_dir.iterdir()} == set(GRADE_OUTPUT_NAMES)
        normalized = _read_csv(grade_dir / "01_du_lieu_tho_sau_chuan_hoa.csv")
        quality = _read_csv(grade_dir / "03_ket_qua_cham_tong_the_tung_mau.csv")
        checklist = _read_csv(grade_dir / "04_ket_qua_cham_chi_tiet_tung_tieu_chi.csv")
        assert len(normalized) == len(quality) == expected_count
        assert len(checklist) == expected_count * 18
        assert all(row["grade"] == grade for row in normalized + quality + checklist)
        assert all(row["source_file"] for row in normalized + quality + checklist)
        grade_ids = {row["sample_id"] for row in normalized}
        grade_pairs = {(row["sample_id"], row["criterion_id"]) for row in checklist}
        assert not all_ids & grade_ids
        assert not all_pairs & grade_pairs
        all_ids.update(grade_ids)
        all_pairs.update(grade_pairs)

        workbook = load_workbook(
            grade_dir / "02_thong_ke_do_phu_mau_pass.xlsx",
            read_only=True,
            data_only=True,
        )
        assert workbook.sheetnames == ["do_phu_mau_pass"]
        workbook.close()

        duplicate_path = grade_dir / "06_ung_vien_trung_lap.csv"
        with duplicate_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            duplicate_rows = list(reader)
            assert reader.fieldnames
        if grade != "9":
            assert duplicate_rows == []

    assert len(all_ids) == 1050
    assert len(all_pairs) == 18900
    normalized_9 = _read_csv(bundle_dir / "lop_9/01_du_lieu_tho_sau_chuan_hoa.csv")
    formula_sample = next(row for row in normalized_9 if row["sample_id"] == "HNMU-G9-R0157-STT2")
    assert formula_sample["answer_sgv"] == "=IF(logical_test, value_if_true, value_if_false)."
    assert formula_sample["source_file"].endswith("Lớp 9.xlsx")

    with pytest.raises(FileExistsError, match="refusing to overwrite"):
        build_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle_dir)


def test_v2_status_workbook_and_statistical_report(tmp_path):
    bundle_dir = tmp_path / "bundle"
    summary = build_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle_dir)

    assert math.isclose(summary.chi_square.statistic, 61.79509372419273)
    assert summary.chi_square.degrees_of_freedom == 6
    assert math.isclose(summary.chi_square.p_value, 1.9421246523453063e-11)
    assert math.isclose(summary.chi_square.cramers_v, 0.17154076806041893)

    workbook = load_workbook(
        bundle_dir / "03_thong_ke_pass_reject_giua_cac_khoi.xlsx",
        read_only=True,
        data_only=True,
    )
    worksheet = workbook["pass_reject_giua_cac_khoi"]
    assert tuple(cell.value for cell in worksheet[1]) == STATUS_HEADERS
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    workbook.close()
    all_rows = {row[2]: row for row in rows if row[0] == "all"}
    assert all_rows["pass"][4] == 665
    assert all_rows["need_human_review"][4] == 382
    assert all_rows["failed"][4] == 3
    assert all_rows["non_pass"][4] == 385
    assert all_rows["non_pass"][6] == 1050

    report = (bundle_dir / "01_bao_cao_tong_quan.md").read_text(encoding="utf-8")
    assert "Chi-square: 61.795094" in report
    assert "Cramér’s V: 0.171541" in report


def test_v2_pass_coverage_is_recomputed_for_each_grade_and_dimension(tmp_path):
    bundle_dir = tmp_path / "bundle"
    build_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle_dir)

    workbook = load_workbook(
        bundle_dir / "04_thong_ke_do_phu_mau_pass_giua_cac_khoi.xlsx",
        read_only=True,
        data_only=True,
    )
    worksheet = workbook["do_phu_pass_giua_cac_khoi"]
    assert tuple(cell.value for cell in worksheet[1]) == COVERAGE_HEADERS
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    workbook.close()

    for grade, expected_pass in {"6": 106, "7": 132, "8": 209, "9": 218}.items():
        for dimension in ("topic", "lesson", "bloom_band"):
            selected = [row for row in rows if row[0] == grade and row[2] == dimension]
            assert selected
            assert sum(row[6] for row in selected) == expected_pass
            assert all(row[8] == expected_pass for row in selected)
            assert math.isclose(sum(row[7] for row in selected), 1.0)
