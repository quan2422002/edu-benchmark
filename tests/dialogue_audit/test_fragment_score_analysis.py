import hashlib
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from edu_benchmark.dialogue_audit.teacher_bundle_v2_partitioned import build_partitioned_phase1_teacher_bundle_v2

from edu_benchmark.dialogue_audit.fragment_score_analysis import (
    FRAGMENT_METRICS,
    GRADE_ANALYSIS_NAME,
    GRADE_SHEET_NAME,
    ROOT_ANALYSIS_NAME,
    ROOT_SHEET_NAME,
    add_fragment_analysis_outputs,
    build_analysis_rows,
    load_fragment_analysis_data,
    validate_fragment_analysis_outputs,
)

EXPERIMENT_DIR = Path("experiments/20260709_155523")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def test_fragment_analysis_joins_at_sample_level_and_preserves_raw_criterion_counts():
    data = load_fragment_analysis_data(EXPERIMENT_DIR)
    records = data.records

    assert len(records) == len({record.sample_id for record in records}) == 1050
    assert Counter(record.grade for record in records) == {
        "6": 238,
        "7": 224,
        "8": 280,
        "9": 308,
    }
    assert Counter(record.grade for record in records if record.official_pass) == {
        "6": 106,
        "7": 132,
        "8": 209,
        "9": 218,
    }
    assert data.criterion_pair_count == 18284
    assert data.criterion_count_distribution_67 == {16: 308, 18: 154}
    assert all(record.auditor_group == f"{record.checked_by} | {record.unified_shard}" for record in records)


def test_fragment_analysis_contains_eight_pairs_and_no_fake_stratum_p_values():
    data = load_fragment_analysis_data(EXPERIMENT_DIR)
    grade_rows, root_rows = build_analysis_rows(data)
    expected_pairs = {
        ("fragment_vs_official_pass", "official_pass", metric)
        for metric in FRAGMENT_METRICS
    } | {
        ("fragment_vs_checklist_pass_rate", "checklist_pass_rate", metric)
        for metric in FRAGMENT_METRICS
    }

    for grade, rows in grade_rows.items():
        actual_pairs = {
            (row["analysis_family"], row["outcome"], row["fragment_metric"])
            for row in rows
            if row["analysis_family"] != "cross_grade_consistency"
        }
        assert actual_pairs == expected_pairs
        assert all(row["grade"] == grade for row in rows)
        diagnostics = [
            row for row in rows if str(row["grouping_or_bucket"]).startswith("non_estimable_stratum:")
        ]
        assert diagnostics
        assert all(row["estimable"] is False and row["p_value"] == "" for row in diagnostics)

    pooled = [row for row in root_rows if row["grade"] == "all"]
    assert any(row["adjustment"] == "adjusted_for_grade_and_auditor_group" for row in pooled)
    assert sum(row["analysis_family"] == "cross_grade_consistency" for row in pooled) == 8


def test_add_and_validate_fragment_analysis_without_changing_existing_data(tmp_path):
    bundle_dir = tmp_path / "bundle"
    build_partitioned_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle_dir)
    markdown = {
        bundle_dir / "README.md",
        bundle_dir / "01_bao_cao_tong_quan.md",
        *(bundle_dir / f"lop_{grade}/README.md" for grade in "6789"),
    }
    immutable = {
        path: _sha256(path)
        for path in bundle_dir.rglob("*")
        if path.is_file() and path not in markdown
    }

    result = add_fragment_analysis_outputs(EXPERIMENT_DIR, bundle_dir)
    validation = validate_fragment_analysis_outputs(EXPERIMENT_DIR, bundle_dir)

    assert result["status"] == validation["status"] == "ok"
    assert result["join_success_count"] == 1050
    assert result["join_failure_count"] == 0
    assert all(_sha256(path) == digest for path, digest in immutable.items())
    expected = [(bundle_dir / ROOT_ANALYSIS_NAME, ROOT_SHEET_NAME)] + [
        (bundle_dir / f"lop_{grade}" / GRADE_ANALYSIS_NAME, GRADE_SHEET_NAME)
        for grade in "6789"
    ]
    for path, sheet in expected:
        workbook = load_workbook(path, read_only=True, data_only=True)
        assert workbook.sheetnames == [sheet]
        workbook.close()
