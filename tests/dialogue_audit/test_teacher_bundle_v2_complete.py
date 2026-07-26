import csv
import hashlib
import math
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from edu_benchmark.dialogue_audit.fragment_score_analysis import FRAGMENT_METRICS, _summary_row
from edu_benchmark.dialogue_audit.fragment_analysis_hnmu_compact import (
    GRADE_ANALYSIS_NAME,
    GRADE_APPENDIX_NAME,
    GRADE_APPENDIX_SHEET,
    GRADE_SUMMARY_SHEET,
    ROOT_ANALYSIS_NAME,
    ROOT_APPENDIX_NAME,
    SUMMARY_DATA_HEADER_ROW,
    SUMMARY_HEADERS,
    SUMMARY_TITLE,
    TECHNICAL_DATA_HEADER_ROW,
)
from edu_benchmark.dialogue_audit.fragment_analysis_root_deliverables import (
    ROOT_RAW_SHEET,
    ROOT_READABLE_SHEET,
    ROOT_TECHNICAL_SHEETS,
)
from edu_benchmark.dialogue_audit.fragment_score_analysis_repaired import (
    build_repaired_fragment_data,
    prepare_analysis_rows,
)
from edu_benchmark.dialogue_audit.hnmu_audit import RawDialogueRow
from edu_benchmark.dialogue_audit.teacher_bundle import load_canonical_bundle_data
from edu_benchmark.dialogue_audit.teacher_bundle_v2_complete import (
    GRADE_OUTPUT_NAMES,
    ROOT_OUTPUT_NAMES,
    _all_grade_duplicate_candidates,
    rebuild_complete_phase1_teacher_bundle_v2,
    validate_complete_phase1_teacher_bundle_v2,
)

EXPERIMENT_DIR = Path("experiments/20260709_155523")
EXPECTED_GRADES = {"6": 238, "7": 224, "8": 280, "9": 308}
EXPECTED_PASS = {"6": 106, "7": 132, "8": 209, "9": 218}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def test_repaired_fragment_input_has_1050_samples_and_18_criteria_each():
    canonical = load_canonical_bundle_data(EXPERIMENT_DIR)
    repaired = build_repaired_fragment_data(canonical)
    grade_rows, root_rows = prepare_analysis_rows(repaired)

    assert len(repaired.records) == 1050
    assert repaired.criterion_pair_count == 18900
    assert Counter(record.observed_criterion_count for record in repaired.records) == {18: 1050}
    assert Counter(record.grade for record in repaired.records) == EXPECTED_GRADES
    assert Counter(record.grade for record in repaired.records if record.official_pass) == EXPECTED_PASS
    assert all(record.auditor_group == f"{record.checked_by} | {record.unified_shard}" for record in repaired.records)

    for grade, rows in grade_rows.items():
        for metric in FRAGMENT_METRICS:
            assert _summary_row(rows, "fragment_vs_official_pass", metric, "crude")
            assert _summary_row(rows, "fragment_vs_official_pass", metric, "adjusted_for_auditor_group")
            assert _summary_row(rows, "fragment_vs_checklist_pass_rate", metric, "crude")
            assert _summary_row(rows, "fragment_vs_checklist_pass_rate", metric, "adjusted_for_auditor_group")
        diagnostics = [row for row in rows if str(row["grouping_or_bucket"]).startswith("non_estimable_stratum:")]
        assert all(row["estimable"] is False and row["p_value"] == "" for row in diagnostics)
        assert all(row["grade"] == grade for row in rows)

    for metric in FRAGMENT_METRICS:
        assert _summary_row(
            root_rows,
            "fragment_vs_official_pass",
            metric,
            "adjusted_for_grade_and_auditor_group",
        )
        assert _summary_row(
            root_rows,
            "fragment_vs_checklist_pass_rate",
            metric,
            "adjusted_for_grade_and_auditor_group",
        )


def test_global_duplicate_detector_checks_pairs_between_grades():
    common = dict(
        source_file="source.xlsx",
        source_row_number=2,
        grade_label="Lớp",
        stt="1",
        lesson="Bài 1",
        position="SGK",
        question="Câu hỏi trùng giữa khối",
        bloom_level="Nhận biết",
        answer_sgv="Đáp án",
        dialogue="Hội thoại khác nhau nhưng câu hỏi được chuẩn hóa giống nhau.",
    )
    left = RawDialogueRow(sample_id="A", grade="6", **common)
    right = RawDialogueRow(sample_id="B", grade="9", **common)
    candidates = _all_grade_duplicate_candidates([left, right])

    assert any(
        row["duplicate_type"] == "exact_question"
        and {row["sample_id_a"], row["sample_id_b"]} == {"A", "B"}
        for row in candidates
    )


def test_complete_bundle_build_and_validation(tmp_path):
    canonical = load_canonical_bundle_data(EXPERIMENT_DIR)
    source_hashes_before = {path: _sha256(path) for path in canonical.source_hashes}
    bundle = tmp_path / "bundle"

    result = rebuild_complete_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle)
    validation = validate_complete_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle)

    assert result["status"] == validation["status"] == "ok"
    assert validation["grade_counts"] == EXPECTED_GRADES
    assert validation["sample_count"] == 1050
    assert validation["criterion_pair_count"] == 18900
    assert validation["criterion_count_distribution"] == {18: 1050}
    assert validation["root_csv_row_counts"] == {
        "06_ket_qua_cham_tong_the_tung_mau.csv": 1050,
        "07_mau_thieu_sai_truong_du_lieu.csv": 22,
        "08_ung_vien_trung_lap.csv": 1,
        "09_du_lieu_tho_sau_chuan_hoa.csv": 1050,
    }
    assert validation["zero_pass_lessons"] == {"6": 6, "7": 5, "8": 4, "9": 3}
    assert validation["lesson_counts"] == {"6": 17, "7": 16, "8": 20, "9": 22}
    assert validation["path_leak_count"] == 0
    assert {path.name for path in bundle.iterdir()} == set(ROOT_OUTPUT_NAMES) | {
        "lop_6", "lop_7", "lop_8", "lop_9"
    }

    root_ids = {row["sample_id"] for row in _csv_rows(bundle / "09_du_lieu_tho_sau_chuan_hoa.csv")}
    grade_ids: set[str] = set()
    for grade, expected in EXPECTED_GRADES.items():
        grade_dir = bundle / f"lop_{grade}"
        assert {path.name for path in grade_dir.iterdir()} == set(GRADE_OUTPUT_NAMES)
        normalized = _csv_rows(grade_dir / "01_du_lieu_tho_sau_chuan_hoa.csv")
        checklist = _csv_rows(grade_dir / "04_ket_qua_cham_chi_tiet_tung_tieu_chi.csv")
        assert len(normalized) == expected
        assert len(checklist) == expected * 18
        assert all(row["grade"] == grade for row in normalized + checklist)
        grade_ids.update(row["sample_id"] for row in normalized)

        analysis_path = grade_dir / GRADE_ANALYSIS_NAME
        workbook = load_workbook(analysis_path, data_only=True)
        worksheet = workbook.active
        assert workbook.sheetnames == [GRADE_SUMMARY_SHEET]
        assert worksheet.freeze_panes == f"A{SUMMARY_DATA_HEADER_ROW + 1}"
        assert len(worksheet.tables) == 2
        assert len(worksheet.merged_cells.ranges) == 6
        assert worksheet["A1"].value == SUMMARY_TITLE
        assert [worksheet.cell(row, 1).value for row in range(2, 5)] == [
            "KẾT LUẬN CHUNG",
            "KẾT QUẢ ĐÁNG CHÚ Ý",
            "GIỚI HẠN KHI DIỄN GIẢI",
        ]
        assert worksheet.max_column == 5
        assert worksheet.max_row == 18
        for header_row in (7, 14):
            assert tuple(
                worksheet.cell(header_row, column).value
                for column in range(1, 6)
            ) == SUMMARY_HEADERS
        visible_text = " ".join(
            str(cell.value or "")
            for row in worksheet.iter_rows()
            for cell in row
        ).casefold()
        assert "khi xem tất cả mẫu" in visible_text
        assert "khi so các mẫu trong cùng nhóm chấm" in visible_text
        assert "cùng khối lớp" not in visible_text
        assert not any(
            term in visible_text
            for term in ("p-value", "r =", "crude", "adjusted", "estimable")
        )
        workbook.close()

        appendix = load_workbook(grade_dir / GRADE_APPENDIX_NAME, data_only=True)
        appendix_sheet = appendix.active
        assert appendix.sheetnames == [GRADE_APPENDIX_SHEET]
        assert appendix_sheet.freeze_panes == f"A{TECHNICAL_DATA_HEADER_ROW + 1}"
        assert appendix_sheet.auto_filter.ref
        assert not appendix_sheet.merged_cells.ranges
        appendix.close()

    assert grade_ids == root_ids and len(root_ids) == 1050
    assert not (bundle / "04_ket_qua_cham_chi_tiet_tung_tieu_chi.csv").exists()
    assert {path: _sha256(path) for path in source_hashes_before} == source_hashes_before

    leak = re.compile(r"experiments/|outputs/|shared/|\b[A-Za-z]:[\\/]", re.IGNORECASE)
    for path in bundle.rglob("*.csv"):
        assert not leak.search(path.read_text(encoding="utf-8-sig"))

    report = (bundle / ROOT_ANALYSIS_NAME).read_text(encoding="utf-8")
    assert report.startswith("# Kết quả phân tích fragment và tỷ lệ đạt\n")
    assert "**Chưa thể khẳng định.**" in report
    assert not any(line.lstrip().startswith("|") for line in report.splitlines())

    appendix = load_workbook(bundle / ROOT_APPENDIX_NAME, data_only=True)
    assert tuple(appendix.sheetnames) == ROOT_TECHNICAL_SHEETS
    assert appendix.active.title == ROOT_READABLE_SHEET
    readable = appendix[ROOT_READABLE_SHEET]
    assert readable.max_column == 6
    assert readable.max_row == 12
    assert readable.freeze_panes == "A5"
    assert len(readable.tables) == 1
    raw = appendix[ROOT_RAW_SHEET]
    assert raw.max_row == 396
    assert raw.max_column == 29
    assert raw.freeze_panes == f"A{TECHNICAL_DATA_HEADER_ROW + 1}"
    assert raw.auto_filter.ref
    for row in raw.iter_rows(min_row=TECHNICAL_DATA_HEADER_ROW + 1, values_only=True):
        estimable = row[16]
        if estimable == "Không":
            assert row[13] in (None, "Không thể ước lượng")
        for value in row:
            assert not (isinstance(value, float) and math.isnan(value))
    appendix.close()
