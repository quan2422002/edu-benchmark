from pathlib import Path

import pytest
from openpyxl import load_workbook

from edu_benchmark.dialogue_audit.teacher_bundle import (
    AllowlistedSourceReader,
    EXPECTED_GRADE_COUNTS,
    EXPECTED_REVIEW_COUNTS,
    SHEET_NAMES,
    build_phase1_teacher_bundle,
    canonical_source_paths,
    load_canonical_bundle_data,
    validate_phase1_teacher_bundle,
)

EXPERIMENT_DIR = Path("experiments/20260709_155523")


def test_canonical_allowlist_contains_exactly_fifteen_local_inputs():
    sources = canonical_source_paths(EXPERIMENT_DIR)
    paths = sources.all_paths()

    assert len(paths) == 15
    assert len(set(paths)) == 15
    assert all(path.is_file() for path in paths)
    assert all("shared/" not in path.as_posix() for path in paths)
    assert not any(
        token in path.as_posix()
        for path in paths
        for token in ("/pilot_", "/shard_", ".pre_", "/regex_repair/")
    )


def test_allowlisted_reader_rejects_source_file_provenance_path():
    sources = canonical_source_paths(EXPERIMENT_DIR)
    reader = AllowlistedSourceReader(sources.all_paths())

    with pytest.raises(ValueError, match="not in the canonical allowlist"):
        reader.read_text(
            Path("shared/raw_data/HNMU-teacher_dialog_samples/Lớp 6.xlsx")
        )


def test_canonical_data_is_a_lossless_four_grade_partition():
    data = load_canonical_bundle_data(EXPERIMENT_DIR)
    counts = {
        grade: sum(row["grade"] == grade for row in data.normalized_by_id.values())
        for grade in EXPECTED_GRADE_COUNTS
    }
    review_counts = {
        grade: sum(
            data.normalized_by_id[sample_id]["grade"] == grade
            for sample_id in data.review_by_id
        )
        for grade in EXPECTED_REVIEW_COUNTS
    }

    assert counts == EXPECTED_GRADE_COUNTS
    assert review_counts == EXPECTED_REVIEW_COUNTS
    assert len(data.normalized_by_id) == 1050
    assert len(data.read_paths) == 15
    assert all(len(data.checklist_by_id[sample_id]) == 18 for sample_id in data.normalized_by_id)


def test_build_and_validate_teacher_bundle_preserves_traceability(tmp_path, monkeypatch):
    original_path_open = Path.open

    def guarded_path_open(path, *args, **kwargs):
        if "/shared/" in path.resolve().as_posix():
            raise AssertionError(f"Builder dereferenced source_file: {path}")
        return original_path_open(path, *args, **kwargs)

    monkeypatch.setattr(Path, "open", guarded_path_open)
    bundle_dir = tmp_path / "teacher_bundle"
    summary = build_phase1_teacher_bundle(EXPERIMENT_DIR, bundle_dir)
    validation = validate_phase1_teacher_bundle(EXPERIMENT_DIR, bundle_dir)

    assert summary.grade_counts == EXPECTED_GRADE_COUNTS
    assert summary.review_counts == EXPECTED_REVIEW_COUNTS
    assert len(summary.read_paths) == 15
    assert validation["status"] == "ok"
    assert validation["source_count"] == 15

    for grade in EXPECTED_GRADE_COUNTS:
        workbook_path = (
            bundle_dir / f"lop_{grade}" / f"01_ket_qua_ra_soat_lop_{grade}.xlsx"
        )
        wb = load_workbook(workbook_path, data_only=True)
        assert wb.sheetnames == SHEET_NAMES
        assert wb["PL_Nguon_du_lieu"].max_row == 16
        wb.close()

    wb = load_workbook(
        bundle_dir / "lop_9/01_ket_qua_ra_soat_lop_9.xlsx",
        data_only=True,
    )
    found = None
    for sheet_name in ("02_Can_ra_soat", "03_Da_dat"):
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            if row["Mã mẫu (sample_id)"] == "HNMU-G9-R0157-STT2":
                found = row
                break
    wb.close()

    assert found is not None
    assert found["Đáp án SGV"] == "=IF(logical_test, value_if_true, value_if_false)."
    assert found["Tệp nguồn (source_file)"].endswith("Lớp 9.xlsx")

    with pytest.raises(FileExistsError):
        build_phase1_teacher_bundle(EXPERIMENT_DIR, bundle_dir)
