from pathlib import Path

from openpyxl import load_workbook

from edu_benchmark.dialogue_audit.teacher_bundle import load_canonical_bundle_data
from edu_benchmark.dialogue_audit.teacher_bundle_v2 import (
    _status_rows,
    format_percentage_half_up,
    round_percentage_fraction_half_up,
)
from edu_benchmark.dialogue_audit.teacher_bundle_v2_complete import (
    _coverage_rows,
    _read_lesson_catalog,
    rebuild_complete_phase1_teacher_bundle_v2,
)

EXPERIMENT_DIR = Path("experiments/20260709_155523")


def test_percentage_helpers_use_round_half_up_at_two_decimal_places():
    assert format_percentage_half_up(91, 224) == "40.63%"
    assert round_percentage_fraction_half_up(91, 224) == 0.4063


def test_count_derived_rows_are_pre_rounded_for_workbook_display():
    data = load_canonical_bundle_data(EXPERIMENT_DIR)
    catalog, _ = _read_lesson_catalog()

    for row in _status_rows(data):
        assert row[5] == round_percentage_fraction_half_up(row[4], row[6])
    for row in _coverage_rows(data, catalog):
        expected = (
            None
            if row[4] == 0
            else round_percentage_fraction_half_up(row[5], row[4])
        )
        assert row[7] == expected


def test_rebuilt_bundle_uses_half_up_in_markdown_and_workbooks(tmp_path):
    bundle = tmp_path / "bundle"
    rebuild_complete_phase1_teacher_bundle_v2(EXPERIMENT_DIR, bundle)

    report = (bundle / "01_bao_cao_tong_quan.md").read_text(encoding="utf-8")
    assert "91 (40.63%)" in report
    assert "91 (40.62%)" not in report

    workbook = load_workbook(
        bundle / "03_thong_ke_pass_reject_giua_cac_khoi.xlsx",
        data_only=True,
    )
    worksheet = workbook["pass_reject_giua_cac_khoi"]
    for row in worksheet.iter_rows(min_row=2):
        count = int(row[4].value)
        total = int(row[6].value)
        assert row[5].value == round_percentage_fraction_half_up(count, total)
        assert row[5].number_format == "0.00%"
    workbook.close()
