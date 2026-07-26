import math
import re
from pathlib import Path

from openpyxl import load_workbook

from edu_benchmark.dialogue_audit.fragment_analysis_hnmu_compact import (
    TECHNICAL_FIELDS,
    build_hnmu_summary_rows,
    grade_summary_display_rows,
    hnmu_main_conclusion,
    validate_summary_keys,
)
from edu_benchmark.dialogue_audit.fragment_analysis_root_deliverables import (
    ROOT_BUCKET_SHEET,
    ROOT_DICTIONARY_SHEET,
    ROOT_NON_ESTIMABLE_SHEET,
    ROOT_RAW_SHEET,
    ROOT_READABLE_SHEET,
    ROOT_STATISTICS_SHEET,
    ROOT_TECHNICAL_SHEETS,
    expected_root_bucket_rows,
    expected_root_dictionary_rows,
    expected_root_non_estimable_rows,
    expected_root_readable_rows,
    expected_root_statistics_rows,
    read_root_raw_technical_rows,
    root_report_content,
    root_report_markdown,
    write_readable_root_technical_workbook,
)
from edu_benchmark.dialogue_audit.fragment_analysis_hnmu import (
    expected_technical_display_rows,
)
from edu_benchmark.dialogue_audit.fragment_score_analysis_repaired import (
    build_repaired_fragment_data,
    prepare_analysis_rows,
)
from edu_benchmark.dialogue_audit.teacher_bundle import load_canonical_bundle_data

EXPERIMENT_DIR = Path("experiments/20260709_155523")


def _analysis_rows():
    canonical = load_canonical_bundle_data(EXPERIMENT_DIR)
    repaired = build_repaired_fragment_data(canonical)
    return prepare_analysis_rows(repaired)


def test_internal_analysis_and_grade_summaries_remain_complete_and_traceable():
    grade_rows, root_rows = _analysis_rows()
    summary = build_hnmu_summary_rows(root_rows, pooled=True)

    assert len(summary) == 8
    assert len({row["analysis_key"] for row in summary}) == 8
    validate_summary_keys(summary, root_rows)
    assert TECHNICAL_FIELDS[-1] == "analysis_key"
    assert len(TECHNICAL_FIELDS) == 29

    for grade, rows in grade_rows.items():
        grade_summary = build_hnmu_summary_rows(rows, pooled=False, grade=grade)
        assert len(grade_summary) == 8
        validate_summary_keys(grade_summary, rows)
        display_rows = grade_summary_display_rows(grade_summary)
        assert len(display_rows) == 8
        assert all(len(row) == 5 for row in display_rows)
        assert [row[0] for row in display_rows[:4]] == [
            "Số tiêu chí có dẫn fragment",
            "Tổng lượt dẫn fragment",
            "Số fragment khác nhau",
            "Tỷ lệ tiêu chí có dẫn fragment",
        ]
        assert all(len(str(row[4]).split()) <= 25 for row in display_rows)


def test_root_report_is_plain_markdown_and_uses_focused_pair():
    _, root_rows = _analysis_rows()
    summary = build_hnmu_summary_rows(root_rows, pooled=True)
    report = root_report_markdown(summary)
    content = root_report_content(summary)

    assert report.startswith("# Kết quả phân tích fragment và tỷ lệ đạt\n")
    assert "Tỷ lệ tiêu chí có dẫn fragment" in report
    assert "350 trong tổng số 1.050 mẫu" in report
    assert "các mẫu có nhiều tiêu chí được dẫn fragment hơn" not in report
    assert content["short_answer"] == "**Chưa thể khẳng định.**"
    assert "Khi xem toàn bộ dữ liệu" in content["explanation"]
    assert "cùng khối lớp và cùng nhóm chấm" in content["explanation"]
    assert content["data_limit"].startswith(
        "Phép so sánh trong cùng khối lớp và nhóm chấm chỉ sử dụng được 350"
    )
    assert not any(line.lstrip().startswith("|") for line in report.splitlines())
    forbidden = (
        "hệ số tương quan", "p-value", "ý nghĩa thống kê",
        "liên hệ dương", "liên hệ âm", "trước điều chỉnh",
        "sau điều chỉnh", "kiểm soát biến", "effect size",
        "strata", "estimable", "fragment_criterion_coverage",
        "official_pass", "checklist_pass_rate",
    )
    visible = report.casefold()
    assert not any(term in visible for term in forbidden)


def test_root_report_changes_when_focused_result_changes():
    _, root_rows = _analysis_rows()
    summary = [dict(row) for row in build_hnmu_summary_rows(root_rows, pooled=True)]
    focused = next(row for row in summary if row["analysis_key"] == "FRG-OP-04")
    focused["adjusted_estimable"] = True
    focused["adjusted_evidence"] = True
    focused["adjusted_statistic"] = 0.18

    content = root_report_content(summary)
    assert content["short_answer"] == "**Có dấu hiệu cho thấy có.**"


def test_readable_root_technical_workbook_preserves_raw_and_splits_views(tmp_path):
    _, root_rows = _analysis_rows()
    path = tmp_path / "technical.xlsx"
    write_readable_root_technical_workbook(path, root_rows)

    workbook = load_workbook(path, data_only=True)
    assert tuple(workbook.sheetnames) == ROOT_TECHNICAL_SHEETS
    assert workbook.active.title == ROOT_READABLE_SHEET
    assert workbook[ROOT_READABLE_SHEET].max_column == 6
    assert workbook[ROOT_READABLE_SHEET].max_row == 12
    assert workbook[ROOT_READABLE_SHEET].freeze_panes == "A5"
    assert len(workbook[ROOT_READABLE_SHEET].tables) == 1
    bucket_rows = expected_root_bucket_rows(root_rows)
    assert workbook[ROOT_BUCKET_SHEET].max_row - 4 == len(bucket_rows)
    reference_labels = {
        row[1] for row in bucket_rows if row[0] == "Tổng lượt dẫn fragment"
    }
    assert {
        "Không quá 5 lượt dẫn fragment",
        "Từ 6 đến 7 lượt dẫn fragment",
        "Trên 7 lượt dẫn fragment",
    } <= reference_labels
    assert "Không quá 7 lượt dẫn fragment" not in reference_labels
    assert workbook[ROOT_STATISTICS_SHEET].max_row - 4 == len(expected_root_statistics_rows(root_rows))
    non_estimable_rows = expected_root_non_estimable_rows(root_rows)
    assert workbook[ROOT_NON_ESTIMABLE_SHEET].max_row - 4 == len(non_estimable_rows)
    assert all(re.fullmatch(r"Nhóm chấm \d{2}", str(row[1])) for row in non_estimable_rows)
    assert workbook[ROOT_DICTIONARY_SHEET].max_row - 4 == len(expected_root_dictionary_rows())
    raw = workbook[ROOT_RAW_SHEET]
    assert raw.max_row == 396
    assert raw.max_column == 29
    workbook.close()

    actual_raw = read_root_raw_technical_rows(path)
    expected_raw = expected_technical_display_rows(root_rows)
    assert len(actual_raw) == len(expected_raw) == 379
    for actual_row, expected_row in zip(actual_raw, expected_raw):
        for actual, expected in zip(actual_row, expected_row):
            if expected == "" and actual is None:
                continue
            if isinstance(expected, float):
                assert isinstance(actual, (int, float))
                assert math.isclose(float(actual), expected, rel_tol=1e-12, abs_tol=1e-12)
            else:
                assert actual == expected
    readable = expected_root_readable_rows(root_rows)
    assert len(readable) == 8
    assert [row[0] for row in readable[:4]] == [
        "Số tiêu chí có dẫn fragment",
        "Tổng lượt dẫn fragment",
        "Số fragment khác nhau",
        "Tỷ lệ tiêu chí có dẫn fragment",
    ]
    focus = next(row for row in readable if row[5] == "FRG-OP-04")
    assert focus[2] == "Có xu hướng cao hơn"
    assert focus[3] == "Không thấy khác biệt rõ ràng"
    assert "Chưa thể khẳng định" in focus[4]
    assert all(len(str(row[4]).split()) <= 25 for row in readable)


def test_pooled_technical_interpretations_still_follow_evidence_rules():
    _, root_rows = _analysis_rows()
    summary = build_hnmu_summary_rows(root_rows, pooled=True)
    interpretations = {row["analysis_key"]: row["plain_interpretation"] for row in summary}

    assert interpretations["FRG-CR-01"] == "Không thấy bằng chứng về mối liên hệ ở cả hai phân tích."
    for key in ("FRG-OP-01", "FRG-OP-04"):
        assert interpretations[key] == "Mối liên hệ quan sát được không còn sau điều chỉnh."
    assert interpretations["FRG-CR-04"] == "Mối liên hệ chỉ xuất hiện sau điều chỉnh và cần được diễn giải thận trọng."
    for key in ("FRG-OP-03", "FRG-CR-03"):
        assert interpretations[key] == "Mối liên hệ đổi chiều sau điều chỉnh nên kết quả chưa ổn định."


def test_grade_8_summary_keeps_all_adjusted_results_non_estimable():
    grade_rows, _ = _analysis_rows()
    summary = build_hnmu_summary_rows(grade_rows["8"], pooled=False, grade="8")

    assert all(row["adjusted_sample_count"] == 0 for row in summary)
    assert all(row["adjusted_estimable"] is False for row in summary)
    conclusion = hnmu_main_conclusion(summary, pooled=False, grade="8")
    assert conclusion == (
        "Trong lớp 8, không đủ dữ liệu phù hợp để kết luận khi so sánh các "
        "mẫu trong cùng nhóm chấm."
    )
